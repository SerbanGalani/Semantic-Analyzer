"""
===============================================================================
AISA - AI Semantic Analyzer
16_viz.py - Visualization Engine
===============================================================================

Three independent output types — each works even if the other two deps are absent.

    1. export_excel_charts(db_path, output_path, year=None)
       → aisa_charts.xlsx  (openpyxl — already installed)
       Charts: Rankings bar · Trend lines · 7-dim Radar ·
               Sentiment pie · Taxonomy bars · Language bar

    2. export_html_dashboard(db_path, output_path, year=None)
       → aisa_dashboard.html  (pip install plotly)
       Charts: Animated leaderboard · Radar profiles ·
               Industry heatmap · Sentiment trend ·
               Category co-occurrence · Language breakdown

    3. export_pdf_report(db_path, output_path, year=None)
       → aisa_report.pdf  (pip install matplotlib)
       Pages: Distribution · Top-25 · Trend lines ·
              Taxonomy heatmap · Industry bars · Sentiment + Language

Usage:
    m = importlib.import_module("16_viz")
    m.export_excel_charts(db_path, "Results_AISA/aisa_charts.xlsx")
    m.export_html_dashboard(db_path, "Results_AISA/aisa_dashboard.html")
    m.export_pdf_report(db_path, "Results_AISA/aisa_report.pdf")

CHANGELOG:
    v1.0.0 (2026-04) - Initial release

Author: TeRa0
License: MIT
===============================================================================
"""

from __future__ import annotations

import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from version import AISA_VERSION

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

DIMS = [
    "volume_index", "depth_index", "breadth_index",
    "tone_index", "specificity_index", "forward_looking_index", "salience_index",
]
DIM_LABELS = ["Volume", "Depth", "Breadth", "Tone", "Specificity", "Fwd-Looking", "Salience"]

CAT_A: Dict[str, str] = {
    "A1": "Product Innovation",   "A2": "Operational Excellence",
    "A3": "Customer Experience",  "A4": "Risk & Compliance",
    "A5": "Data & Analytics",     "A6": "Strategy & Investment",
    "A7": "Governance & Ethics",  "A8": "Talent & Workforce",
}
CAT_B: Dict[str, str] = {
    "B1": "Traditional ML",  "B2": "Deep Learning",
    "B3": "NLP",             "B4": "GenAI & LLMs",
    "B5": "Computer Vision", "B6": "Robotics & Auto",
    "B7": "Infrastructure",  "B8": "General AI",
}

_PALETTE = [
    "#1f4e79", "#2e75b6", "#4aa3df", "#70c1e8",
    "#a9d6f5", "#1a936f", "#88d498", "#f4a261",
    "#e76f51", "#9467bd", "#c5b0d5", "#8c564b",
]


# ─────────────────────────────────────────────────────────────────────────────
# Shared data helpers
# ─────────────────────────────────────────────────────────────────────────────

def _connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _df(conn: sqlite3.Connection, sql: str, params: tuple = ()) -> pd.DataFrame:
    try:
        rows = conn.execute(sql, params).fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _load_buzz(conn: sqlite3.Connection, year: Optional[int] = None) -> pd.DataFrame:
    df = _df(conn, "SELECT * FROM adoption_index ORDER BY year, rank_in_year")
    if not df.empty and year:
        df = df[df["year"] == year]
    return df


def _load_refs(conn: sqlite3.Connection, year: Optional[int] = None) -> pd.DataFrame:
    sql = (
        "SELECT company, year, category_a, category_b, sentiment, "
        "semantic_score, language, product_name, product_vendor "
        "FROM ai_references_raw"
    )
    params: tuple = ()
    if year:
        sql += " WHERE year = ?"
        params = (year,)
    return _df(conn, sql, params)


def _load_industry(conn: sqlite3.Connection, year: Optional[int] = None) -> pd.DataFrame:
    df = _df(conn, "SELECT * FROM adoption_index_industry ORDER BY year, rank_among_industries")
    if not df.empty and year:
        df = df[df["year"] == year]
    return df


def _latest_year(df: pd.DataFrame) -> Optional[int]:
    if df.empty or "year" not in df.columns:
        return None
    return int(df["year"].max())


def _short_name(name: str, maxlen: int = 22) -> str:
    """Truncate long company names for chart labels."""
    return name if len(name) <= maxlen else name[:maxlen - 1] + "…"


def _cat_code(raw: str) -> str:
    """Extract 'A1', 'B4' etc. from category strings like 'A1_Product_Innovation'."""
    m = re.match(r"^([AB]\d)", str(raw or ""))
    return m.group(1) if m else str(raw or "")[:2]


# ─────────────────────────────────────────────────────────────────────────────
# 1. EXCEL CHARTS  (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def export_excel_charts(
    db_path: str,
    output_path: str,
    year: Optional[int] = None,
) -> str:
    """
    Create aisa_charts.xlsx with 6 chart sheets embedded via openpyxl.

    Args:
        db_path:     Path to the AISA SQLite database.
        output_path: Destination .xlsx file.
        year:        If set, filter all data to this year only.

    Returns:
        Absolute path to the written file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, RadarChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    conn = _connect(db_path)
    df_buzz = _load_buzz(conn, year)
    df_refs = _load_refs(conn, year)
    df_ind  = _load_industry(conn, year)
    conn.close()

    latest = year or _latest_year(df_buzz)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _xl_rankings(wb, df_buzz, latest)
    _xl_trend(wb, df_buzz)
    _xl_radar(wb, df_buzz, latest)
    _xl_sentiment(wb, df_refs)
    _xl_taxonomy(wb, df_refs)
    _xl_language(wb, df_refs)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(Path(output_path).resolve())


# ── Excel helper: style header row ───────────────────────────────────────────

def _xl_header(ws, cols: List[str]):
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


# ── Chart 1: Top-25 Rankings (horizontal bar) ────────────────────────────────

def _xl_rankings(wb, df_buzz: pd.DataFrame, year: Optional[int]):
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Rankings")
    if df_buzz.empty:
        ws["A1"] = "No AI Buzz Index data available."
        return

    yr_label = str(year) if year else "All Years"
    df = df_buzz.copy()
    if year and "year" in df.columns:
        df = df[df["year"] == year]
    if df.empty:
        ws["A1"] = f"No data for year {year}."
        return

    df = df.nlargest(25, "ai_buzz_index")[["company", "ai_buzz_index"]].reset_index(drop=True)

    _xl_header(ws, ["Company", "AI Buzz Index"])
    for r, row in df.iterrows():
        ws.cell(row=r + 2, column=1, value=_short_name(row["company"]))
        ws.cell(row=r + 2, column=2, value=round(float(row["ai_buzz_index"]), 4))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    chart = BarChart()
    chart.type         = "bar"          # horizontal
    chart.grouping     = "clustered"
    chart.title        = f"Top 25 Companies — AI Buzz Index ({yr_label})"
    chart.y_axis.title = "Company"
    chart.x_axis.title = "AI Buzz Index"
    chart.width        = 24
    chart.height       = 18

    n = len(df) + 1
    data   = Reference(ws, min_col=2, min_row=1, max_row=n)
    cats   = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


# ── Chart 2: Trend lines (top-10 companies over years) ───────────────────────

def _xl_trend(wb, df_buzz: pd.DataFrame):
    from openpyxl.chart import LineChart, Reference

    ws = wb.create_sheet("Trend")
    if df_buzz.empty or "year" not in df_buzz.columns:
        ws["A1"] = "No trend data available."
        return

    years = sorted(df_buzz["year"].unique())
    if len(years) < 2:
        ws["A1"] = "Need at least 2 years for trend chart."
        return

    # Pick top-10 companies by average buzz
    top10 = (
        df_buzz.groupby("company")["ai_buzz_index"]
        .mean()
        .nlargest(10)
        .index.tolist()
    )
    pivot = (
        df_buzz[df_buzz["company"].isin(top10)]
        .pivot_table(index="year", columns="company", values="ai_buzz_index")
        .reindex(years)
        .fillna(0)
    )
    pivot.columns = [_short_name(c) for c in pivot.columns]

    # Write data
    cols = ["Year"] + list(pivot.columns)
    _xl_header(ws, cols)
    for r, yr in enumerate(pivot.index, 2):
        ws.cell(row=r, column=1, value=int(yr))
        for c, val in enumerate(pivot.loc[yr], 2):
            ws.cell(row=r, column=c, value=round(float(val), 4))
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 16

    chart = LineChart()
    chart.title        = "AI Buzz Index Trend — Top 10 Companies"
    chart.y_axis.title = "AI Buzz Index"
    chart.x_axis.title = "Year"
    chart.width        = 26
    chart.height       = 16
    chart.style        = 10

    n_rows = len(pivot) + 1
    n_cols = len(pivot.columns) + 1
    data = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A{}".format(len(pivot) + 5))


# ── Chart 3: 7-dimension Radar (top-5 companies) ─────────────────────────────

def _xl_radar(wb, df_buzz: pd.DataFrame, year: Optional[int]):
    from openpyxl.chart import RadarChart, Reference

    ws = wb.create_sheet("Profiles")
    if df_buzz.empty:
        ws["A1"] = "No profile data available."
        return

    df = df_buzz.copy()
    if year and "year" in df.columns:
        df = df[df["year"] == year]
    if df.empty or not all(d in df.columns for d in DIMS):
        ws["A1"] = "Dimension data not available."
        return

    top5 = df.nlargest(5, "ai_buzz_index")[["company"] + DIMS].reset_index(drop=True)

    # Layout: rows = dimensions, columns = companies
    _xl_header(ws, ["Dimension"] + [_short_name(c) for c in top5["company"]])
    for r, (dim, label) in enumerate(zip(DIMS, DIM_LABELS), 2):
        ws.cell(row=r, column=1, value=label)
        for c, val in enumerate(top5[dim], 2):
            ws.cell(row=r, column=c, value=round(float(val), 4))
    ws.column_dimensions["A"].width = 18
    for i in range(2, len(top5) + 2):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 16

    chart = RadarChart()
    chart.type   = "filled"
    chart.title  = f"7-Dimension AI Profile — Top 5 ({year or 'All'})"
    chart.width  = 20
    chart.height = 16

    cats = Reference(ws, min_col=1, min_row=2, max_row=8)
    chart.set_categories(cats)
    for col in range(2, len(top5) + 2):
        values = Reference(ws, min_col=col, min_row=1, max_row=8)
        series = chart.series.append(__import__("openpyxl").chart.Series(values, title_from_data=True))
    ws.add_chart(chart, "A12")


# ── Chart 4: Sentiment distribution (pie) ────────────────────────────────────

def _xl_sentiment(wb, df_refs: pd.DataFrame):
    from openpyxl.chart import PieChart, Reference

    ws = wb.create_sheet("Sentiment")
    if df_refs.empty or "sentiment" not in df_refs.columns:
        ws["A1"] = "No sentiment data available."
        return

    counts = df_refs["sentiment"].value_counts()
    _xl_header(ws, ["Sentiment", "Count"])
    for r, (label, cnt) in enumerate(counts.items(), 2):
        ws.cell(row=r, column=1, value=str(label).capitalize())
        ws.cell(row=r, column=2, value=int(cnt))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12

    n = len(counts) + 1
    chart = PieChart()
    chart.title  = "Sentiment Distribution of AI References"
    chart.width  = 16
    chart.height = 14

    data  = Reference(ws, min_col=2, min_row=1, max_row=n)
    cats  = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


# ── Chart 5: Taxonomy category usage (bar) ───────────────────────────────────

def _xl_taxonomy(wb, df_refs: pd.DataFrame):
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Taxonomy")
    if df_refs.empty:
        ws["A1"] = "No reference data available."
        return

    row_offset = 1
    for dim_col, cat_map, label in [
        ("category_a", CAT_A, "Application (A)"),
        ("category_b", CAT_B, "Technology (B)"),
    ]:
        if dim_col not in df_refs.columns:
            continue

        codes  = df_refs[dim_col].dropna().apply(_cat_code)
        counts = codes.value_counts()
        ordered_codes = sorted(cat_map.keys())

        # Header
        ws.cell(row=row_offset, column=1, value=f"Category ({label})")
        ws.cell(row=row_offset, column=2, value="References")
        from openpyxl.styles import Font, PatternFill
        for c in [1, 2]:
            cell = ws.cell(row=row_offset, column=c)
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF")

        data_start = row_offset + 1
        for code in ordered_codes:
            ws.cell(row=row_offset + 1, column=1,
                    value=f"{code} — {cat_map.get(code, code)}")
            ws.cell(row=row_offset + 1, column=2,
                    value=int(counts.get(code, 0)))
            row_offset += 1
        row_offset += 1  # blank separator

        n_data = len(ordered_codes)
        chart = BarChart()
        chart.type         = "col"
        chart.title        = f"Reference Count by {label} Category"
        chart.y_axis.title = "References"
        chart.width        = 20
        chart.height       = 12

        data_ref = Reference(ws, min_col=2, min_row=data_start - 1,
                             max_row=data_start + n_data - 1)
        cats_ref = Reference(ws, min_col=1, min_row=data_start,
                             max_row=data_start + n_data - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        anchor = "D{}".format(data_start - 1)
        ws.add_chart(chart, anchor)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14


# ── Chart 6: Language distribution (bar) ─────────────────────────────────────

def _xl_language(wb, df_refs: pd.DataFrame):
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Languages")
    if df_refs.empty or "language" not in df_refs.columns:
        ws["A1"] = "No language data available."
        return

    lang_labels = {"en": "English", "zh": "Chinese", "ja": "Japanese",
                   "ko": "Korean", "other": "Other"}
    counts = df_refs["language"].fillna("en").value_counts()

    _xl_header(ws, ["Language", "References"])
    for r, (code, cnt) in enumerate(counts.items(), 2):
        ws.cell(row=r, column=1, value=lang_labels.get(str(code), str(code)))
        ws.cell(row=r, column=2, value=int(cnt))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14

    n = len(counts) + 1
    chart = BarChart()
    chart.type         = "col"
    chart.title        = "References by Document Language"
    chart.y_axis.title = "References"
    chart.width        = 16
    chart.height       = 12

    data = Reference(ws, min_col=2, min_row=1, max_row=n)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


# ─────────────────────────────────────────────────────────────────────────────
# 2. PLOTLY HTML DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def export_html_dashboard(
    db_path: str,
    output_path: str,
    year: Optional[int] = None,
) -> str:
    """
    Create a standalone interactive HTML dashboard using Plotly.

    Args:
        db_path:     Path to the AISA SQLite database.
        output_path: Destination .html file.
        year:        If set, filter static charts to this year.

    Returns:
        Absolute path to the written file.
    """
    try:
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        import plotly.io as pio
    except ImportError:
        raise ImportError("plotly is required: pip install plotly")

    conn    = _connect(db_path)
    df_buzz = _load_buzz(conn)          # full history for animations
    df_refs = _load_refs(conn, year)
    df_ind  = _load_industry(conn)
    conn.close()

    latest = year or _latest_year(df_buzz)
    charts_html: List[str] = []

    def _fig_html(fig, title: str, desc: str = "") -> str:
        fig.update_layout(
            template="plotly_white",
            font=dict(family="Segoe UI, Arial, sans-serif", size=13),
            margin=dict(l=40, r=40, t=60, b=40),
            paper_bgcolor="#ffffff",
            plot_bgcolor="#f8f9fa",
        )
        h = pio.to_html(fig, full_html=False, include_plotlyjs=False,
                        config={"displayModeBar": True, "responsive": True})
        return f"""
        <div class="chart-card">
          <div class="chart-title">{title}</div>
          {"<div class='chart-desc'>" + desc + "</div>" if desc else ""}
          {h}
        </div>"""

    # ── Chart 1: Animated AI Buzz Leaderboard ────────────────────────────────
    if not df_buzz.empty and "year" in df_buzz.columns:
        top_n = 20
        companies = (
            df_buzz.groupby("company")["ai_buzz_index"]
            .mean().nlargest(top_n).index.tolist()
        )
        df_anim = df_buzz[df_buzz["company"].isin(companies)].copy()
        df_anim["company_short"] = df_anim["company"].apply(lambda x: _short_name(x, 28))
        df_anim = df_anim.sort_values(["year", "ai_buzz_index"], ascending=[True, True])

        fig = px.bar(
            df_anim, x="ai_buzz_index", y="company_short",
            animation_frame="year", orientation="h",
            color="ai_buzz_index",
            color_continuous_scale=[[0, "#e8f4fd"], [0.5, "#2e75b6"], [1, "#1f4e79"]],
            labels={"ai_buzz_index": "AI Buzz Index", "company_short": ""},
            range_x=[0, df_anim["ai_buzz_index"].max() * 1.1],
        )
        fig.update_layout(
            height=520, showlegend=False,
            coloraxis_showscale=False,
            yaxis=dict(tickfont=dict(size=11)),
        )
        charts_html.append(_fig_html(
            fig,
            "🏆 AI Buzz Index Leaderboard",
            "Top 20 companies — use the Play button to animate across years",
        ))

    # ── Chart 2: 7-Dimension Radar with company selector ─────────────────────
    if not df_buzz.empty and all(d in df_buzz.columns for d in DIMS):
        df_latest = df_buzz[df_buzz["year"] == latest] if latest else df_buzz
        top30 = df_latest.nlargest(30, "ai_buzz_index")["company"].tolist()

        fig = go.Figure()
        for i, company in enumerate(top30[:10]):
            row = df_latest[df_latest["company"] == company]
            if row.empty:
                continue
            vals = [float(row.iloc[0].get(d, 0)) for d in DIMS]
            vals.append(vals[0])  # close polygon
            fig.add_trace(go.Scatterpolar(
                r=vals,
                theta=DIM_LABELS + [DIM_LABELS[0]],
                fill="toself",
                opacity=0.6,
                name=_short_name(company, 25),
                visible=(i == 0),
            ))

        buttons = []
        for i, company in enumerate(top30[:10]):
            vis = [j == i for j in range(len(top30[:10]))]
            buttons.append(dict(
                label=_short_name(company, 22),
                method="update",
                args=[{"visible": vis}, {"title": f"7-Dimension Profile: {_short_name(company, 30)}"}],
            ))

        fig.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 1])),
            updatemenus=[dict(
                buttons=buttons, direction="down",
                showactive=True, x=0.01, xanchor="left", y=1.15, yanchor="top",
            )],
            height=500,
            title=f"7-Dimension Profile: {_short_name(top30[0], 30) if top30 else ''}",
        )
        charts_html.append(_fig_html(
            fig,
            "📡 7-Dimension AI Profile",
            f"Radar of Volume · Depth · Breadth · Tone · Specificity · Fwd-Looking · Salience — {latest}",
        ))

    # ── Chart 3: Industry heatmap (industry × year) ───────────────────────────
    if not df_ind.empty and "industry" in df_ind.columns and "year" in df_ind.columns:
        pivot = df_ind.pivot_table(
            index="industry", columns="year", values="ai_buzz_index_industry"
        ).fillna(0)
        pivot = pivot.loc[pivot.max(axis=1).nlargest(15).index]

        fig = go.Figure(go.Heatmap(
            z=pivot.values.tolist(),
            x=[str(y) for y in pivot.columns],
            y=list(pivot.index),
            colorscale=[[0, "#e8f4fd"], [0.4, "#4aa3df"], [1, "#1f4e79"]],
            hovertemplate="Industry: %{y}<br>Year: %{x}<br>AI Buzz: %{z:.3f}<extra></extra>",
        ))
        fig.update_layout(
            height=440,
            xaxis_title="Year",
            yaxis=dict(tickfont=dict(size=11)),
        )
        charts_html.append(_fig_html(
            fig,
            "🏭 Industry AI Buzz Heatmap",
            "Average AI Buzz Index per industry across years — darker = higher AI engagement",
        ))

    # ── Chart 4: Sentiment trend (stacked area) ───────────────────────────────
    if not df_refs.empty and "sentiment" in df_refs.columns and "year" in df_refs.columns:
        sent_yr = (
            df_refs.groupby(["year", "sentiment"])
            .size().reset_index(name="count")
        )
        order = ["positive", "neutral", "negative"]
        colors = {"positive": "#1a936f", "neutral": "#4aa3df", "negative": "#e76f51"}

        fig = go.Figure()
        for s in order:
            sub = sent_yr[sent_yr["sentiment"] == s]
            if sub.empty:
                continue
            fig.add_trace(go.Scatter(
                x=sub["year"], y=sub["count"],
                name=s.capitalize(), mode="lines",
                stackgroup="one", fillcolor=colors.get(s, "#aaa"),
                line=dict(color=colors.get(s, "#aaa")),
                hovertemplate=f"{s.capitalize()}: %{{y:,}}<extra></extra>",
            ))
        fig.update_layout(
            height=380, xaxis_title="Year", yaxis_title="References",
            legend=dict(orientation="h", y=1.1),
        )
        charts_html.append(_fig_html(
            fig,
            "💬 Sentiment Trend",
            "Positive / Neutral / Negative AI reference distribution over time",
        ))

    # ── Chart 5: Category co-occurrence heatmap (A × B) ──────────────────────
    if not df_refs.empty and "category_a" in df_refs.columns and "category_b" in df_refs.columns:
        df_c = df_refs.copy()
        df_c["ca"] = df_c["category_a"].apply(_cat_code)
        df_c["cb"] = df_c["category_b"].apply(_cat_code)

        cross = df_c.groupby(["ca", "cb"]).size().reset_index(name="count")
        a_codes = sorted([k for k in CAT_A if k in df_c["ca"].values])
        b_codes = sorted([k for k in CAT_B if k in df_c["cb"].values])

        matrix = np.zeros((len(a_codes), len(b_codes)))
        for _, row in cross.iterrows():
            if row["ca"] in a_codes and row["cb"] in b_codes:
                matrix[a_codes.index(row["ca"])][b_codes.index(row["cb"])] = row["count"]

        fig = go.Figure(go.Heatmap(
            z=matrix.tolist(),
            x=[f"{c} {CAT_B.get(c,'')}" for c in b_codes],
            y=[f"{c} {CAT_A.get(c,'')}" for c in a_codes],
            colorscale=[[0, "#f8f9fa"], [0.3, "#88d498"], [1, "#1a936f"]],
            hovertemplate="App: %{y}<br>Tech: %{x}<br>Refs: %{z:,}<extra></extra>",
        ))
        fig.update_layout(
            height=400,
            xaxis=dict(tickangle=-30, tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=10)),
        )
        charts_html.append(_fig_html(
            fig,
            "🔀 Application × Technology Co-occurrence",
            "How often each use-case (A) combines with a technology type (B)",
        ))

    # ── Chart 6: Language breakdown per year ─────────────────────────────────
    if not df_refs.empty and "language" in df_refs.columns and "year" in df_refs.columns:
        lang_map = {"en": "English", "zh": "Chinese", "ja": "Japanese",
                    "ko": "Korean", "other": "Other"}
        df_l = df_refs.copy()
        df_l["lang_label"] = df_l["language"].fillna("en").map(
            lambda x: lang_map.get(str(x), str(x))
        )
        lang_yr = df_l.groupby(["year", "lang_label"]).size().reset_index(name="count")
        lang_colors = {
            "English": "#1f4e79", "Chinese": "#e63946",
            "Japanese": "#f4a261", "Korean": "#1a936f", "Other": "#9467bd",
        }

        fig = go.Figure()
        for lang in ["English", "Chinese", "Japanese", "Korean", "Other"]:
            sub = lang_yr[lang_yr["lang_label"] == lang]
            if sub.empty:
                continue
            fig.add_trace(go.Bar(
                x=sub["year"], y=sub["count"],
                name=lang, marker_color=lang_colors.get(lang, "#aaa"),
            ))
        fig.update_layout(
            barmode="stack", height=360,
            xaxis_title="Year", yaxis_title="References",
            legend=dict(orientation="h", y=1.1),
        )
        charts_html.append(_fig_html(
            fig,
            "🌏 References by Document Language",
            "Multilingual coverage — non-English references use the multilingual semantic model",
        ))

    # ── Build final HTML ──────────────────────────────────────────────────────
    total_charts = len(charts_html)
    if total_charts == 0:
        charts_html.append("<p style='padding:2rem;color:#888'>No data found in the database.</p>")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>AISA Dashboard v{AISA_VERSION}</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    background: #f0f4f8;
    color: #1a2332;
  }}
  header {{
    background: linear-gradient(135deg, #1f4e79 0%, #2e75b6 100%);
    color: white;
    padding: 1.5rem 2rem;
    display: flex; align-items: center; gap: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.2);
  }}
  header .logo {{ font-size: 2rem; }}
  header .title-block h1 {{ font-size: 1.6rem; font-weight: 700; }}
  header .title-block p  {{ font-size: 0.85rem; opacity: 0.8; margin-top: 0.2rem; }}
  header .badge {{
    margin-left: auto;
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.3);
    padding: 0.3rem 0.8rem;
    border-radius: 1rem;
    font-size: 0.8rem;
  }}
  main {{
    max-width: 1300px;
    margin: 0 auto;
    padding: 1.5rem;
    display: grid;
    grid-template-columns: 1fr;
    gap: 1.5rem;
  }}
  .chart-card {{
    background: white;
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
    border: 1px solid #e2e8f0;
  }}
  .chart-title {{
    font-size: 1.1rem;
    font-weight: 600;
    color: #1f4e79;
    margin-bottom: 0.25rem;
  }}
  .chart-desc {{
    font-size: 0.8rem;
    color: #64748b;
    margin-bottom: 0.75rem;
  }}
  footer {{
    text-align: center;
    padding: 1.5rem;
    color: #94a3b8;
    font-size: 0.8rem;
  }}
</style>
</head>
<body>
<header>
  <span class="logo">📊</span>
  <div class="title-block">
    <h1>AISA Dashboard</h1>
    <p>AI Semantic Analyzer — Fortune 500 AI Disclosure Analysis</p>
  </div>
  <div class="badge">v{AISA_VERSION} &nbsp;·&nbsp; {datetime.now().strftime('%Y-%m-%d')}</div>
</header>
<main>
{''.join(charts_html)}
</main>
<footer>
  Generated by AISA v{AISA_VERSION} &nbsp;·&nbsp; {total_charts} charts &nbsp;·&nbsp;
  {datetime.now().strftime('%Y-%m-%d %H:%M')}
</footer>
</body>
</html>"""

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(html, encoding="utf-8")
    return str(Path(output_path).resolve())


# ─────────────────────────────────────────────────────────────────────────────
# 3. MATPLOTLIB PDF REPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_pdf_report(
    db_path: str,
    output_path: str,
    year: Optional[int] = None,
) -> str:
    """
    Create a multi-page publication-quality PDF using Matplotlib.

    Pages:
        1. Title + summary stats
        2. AI Buzz Index distribution (box plots per year)
        3. Top-25 companies — latest year (horizontal bar)
        4. Top-10 trend lines over years
        5. Taxonomy A×B co-occurrence heatmap
        6. Industry comparison (bar)
        7. Sentiment + Language breakdown (2 subplots)

    Args:
        db_path:     Path to the AISA SQLite database.
        output_path: Destination .pdf file.
        year:        If set, use this as the reference "latest" year.

    Returns:
        Absolute path to the written file.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        from matplotlib.backends.backend_pdf import PdfPages
        from matplotlib.gridspec import GridSpec
        import matplotlib.ticker as mticker
    except ImportError:
        raise ImportError("matplotlib is required: pip install matplotlib")

    conn    = _connect(db_path)
    df_buzz = _load_buzz(conn)
    df_refs = _load_refs(conn, year)
    df_ind  = _load_industry(conn, year)
    conn.close()

    latest = year or _latest_year(df_buzz)

    # Apply a clean academic style
    plt.rcParams.update({
        "font.family":       "DejaVu Sans",
        "font.size":         10,
        "axes.titlesize":    12,
        "axes.titleweight":  "bold",
        "axes.spines.top":   False,
        "axes.spines.right": False,
        "axes.grid":         True,
        "axes.grid.axis":    "y",
        "grid.alpha":        0.4,
        "figure.dpi":        150,
        "savefig.dpi":       200,
        "savefig.bbox":      "tight",
    })

    BLUE   = "#1f4e79"
    ACCENT = "#2e75b6"
    LIGHT  = "#e8f4fd"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with PdfPages(output_path) as pdf:

        # ── Page 1: Title & summary stats ────────────────────────────────────
        fig = plt.figure(figsize=(11, 8.5))
        ax  = fig.add_subplot(111)
        ax.axis("off")

        n_companies = df_buzz["company"].nunique() if not df_buzz.empty else 0
        n_refs_raw  = len(df_refs)
        years_range = ""
        if not df_buzz.empty and "year" in df_buzz.columns:
            y_min, y_max = int(df_buzz["year"].min()), int(df_buzz["year"].max())
            years_range = f"{y_min} – {y_max}"

        fig.patch.set_facecolor(BLUE)
        ax.set_facecolor(BLUE)

        ax.text(0.5, 0.80, "AISA — AI Semantic Analyzer",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=28, fontweight="bold", color="white")
        ax.text(0.5, 0.70, "Fortune 500 AI Disclosure Analysis",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=16, color="#a9d6f5")
        ax.text(0.5, 0.60, f"v{AISA_VERSION}   ·   {datetime.now().strftime('%B %Y')}",
                ha="center", va="center", transform=ax.transAxes,
                fontsize=11, color="#a9d6f5")

        stats_text = (
            f"Companies:  {n_companies:,}       "
            f"References:  {n_refs_raw:,}       "
            f"Years:  {years_range}"
        )
        ax.text(0.5, 0.42, stats_text,
                ha="center", va="center", transform=ax.transAxes,
                fontsize=13, color="white",
                bbox=dict(boxstyle="round,pad=0.6", facecolor="rgba(255,255,255,0.1)",
                          edgecolor="rgba(255,255,255,0.3)"))

        pdf.savefig(fig, facecolor=BLUE)
        plt.close(fig)

        # ── Page 2: Distribution (box plots per year) ─────────────────────────
        if not df_buzz.empty and "year" in df_buzz.columns:
            fig, ax = plt.subplots(figsize=(11, 6))
            years = sorted(df_buzz["year"].unique())
            data_by_year = [
                df_buzz[df_buzz["year"] == yr]["ai_buzz_index"].dropna().values
                for yr in years
            ]
            bp = ax.boxplot(
                data_by_year,
                labels=[str(y) for y in years],
                patch_artist=True,
                medianprops=dict(color="white", linewidth=2),
                whiskerprops=dict(color=ACCENT),
                capprops=dict(color=ACCENT),
                flierprops=dict(marker="o", markerfacecolor=ACCENT, markersize=3, alpha=0.5),
            )
            for patch in bp["boxes"]:
                patch.set_facecolor(ACCENT)
                patch.set_alpha(0.7)

            ax.set_title("AI Buzz Index Distribution by Year")
            ax.set_xlabel("Year")
            ax.set_ylabel("AI Buzz Index")
            ax.set_facecolor("#f8f9fa")
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        # ── Page 3: Top-25 horizontal bar (latest year) ───────────────────────
        if not df_buzz.empty and latest:
            df_yr = df_buzz[df_buzz["year"] == latest].copy()
            if not df_yr.empty:
                df_yr = df_yr.nlargest(25, "ai_buzz_index")
                df_yr = df_yr.sort_values("ai_buzz_index")

                fig, ax = plt.subplots(figsize=(11, 9))
                colors_bar = [
                    BLUE if i >= len(df_yr) - 3 else ACCENT
                    for i in range(len(df_yr))
                ]
                bars = ax.barh(
                    [_short_name(c, 30) for c in df_yr["company"]],
                    df_yr["ai_buzz_index"],
                    color=colors_bar, edgecolor="none", height=0.7,
                )
                # Value labels
                for bar in bars:
                    w = bar.get_width()
                    ax.text(w + 0.002, bar.get_y() + bar.get_height() / 2,
                            f"{w:.3f}", va="center", fontsize=8, color="#444")

                ax.set_title(f"Top 25 Companies — AI Buzz Index ({latest})")
                ax.set_xlabel("AI Buzz Index")
                ax.set_facecolor("#f8f9fa")
                ax.xaxis.grid(True, alpha=0.4)
                ax.yaxis.grid(False)
                ax.spines["bottom"].set_visible(True)
                ax.tick_params(axis="y", labelsize=9)
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)

        # ── Page 4: Top-10 trend lines ────────────────────────────────────────
        if not df_buzz.empty and "year" in df_buzz.columns:
            years = sorted(df_buzz["year"].unique())
            if len(years) >= 2:
                top10 = (
                    df_buzz.groupby("company")["ai_buzz_index"]
                    .mean().nlargest(10).index.tolist()
                )
                fig, ax = plt.subplots(figsize=(11, 6))
                for i, company in enumerate(top10):
                    sub = df_buzz[df_buzz["company"] == company].sort_values("year")
                    ax.plot(
                        sub["year"], sub["ai_buzz_index"],
                        marker="o", markersize=4, linewidth=1.8,
                        color=_PALETTE[i % len(_PALETTE)],
                        label=_short_name(company, 22),
                    )
                ax.set_title("AI Buzz Index Trend — Top 10 Companies")
                ax.set_xlabel("Year")
                ax.set_ylabel("AI Buzz Index")
                ax.set_xticks(years)
                ax.set_facecolor("#f8f9fa")
                ax.legend(fontsize=8, bbox_to_anchor=(1.01, 1), loc="upper left",
                          frameon=True, framealpha=0.9)
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)

        # ── Page 5: Taxonomy A×B co-occurrence heatmap ────────────────────────
        if (not df_refs.empty
                and "category_a" in df_refs.columns
                and "category_b" in df_refs.columns):

            df_c = df_refs.copy()
            df_c["ca"] = df_c["category_a"].apply(_cat_code)
            df_c["cb"] = df_c["category_b"].apply(_cat_code)

            a_codes = sorted(CAT_A.keys())
            b_codes = sorted(CAT_B.keys())
            matrix  = np.zeros((len(a_codes), len(b_codes)), dtype=int)

            cross = df_c.groupby(["ca", "cb"]).size()
            for (ca, cb), cnt in cross.items():
                if ca in a_codes and cb in b_codes:
                    matrix[a_codes.index(ca)][b_codes.index(cb)] = int(cnt)

            fig, ax = plt.subplots(figsize=(11, 6))
            im = ax.imshow(matrix, cmap="Blues", aspect="auto")
            ax.set_xticks(range(len(b_codes)))
            ax.set_xticklabels(
                [f"{c}\n{CAT_B[c]}" for c in b_codes], fontsize=8, rotation=15, ha="right"
            )
            ax.set_yticks(range(len(a_codes)))
            ax.set_yticklabels(
                [f"{c} {CAT_A[c]}" for c in a_codes], fontsize=8
            )
            for i in range(len(a_codes)):
                for j in range(len(b_codes)):
                    v = matrix[i, j]
                    if v > 0:
                        ax.text(j, i, f"{v:,}", ha="center", va="center",
                                fontsize=7, color="white" if v > matrix.max() * 0.5 else "#333")
            plt.colorbar(im, ax=ax, shrink=0.7, label="Reference count")
            ax.set_title("Application × Technology Category Co-occurrence")
            ax.set_xlabel("Technology Category (B)")
            ax.set_ylabel("Application Category (A)")
            ax.grid(False)
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        # ── Page 6: Industry comparison ───────────────────────────────────────
        if not df_ind.empty and "industry" in df_ind.columns:
            df_i = df_ind.copy()
            if latest and "year" in df_i.columns:
                df_i = df_i[df_i["year"] == latest]
            if not df_i.empty:
                df_i = df_i.nlargest(15, "ai_buzz_index_industry").sort_values(
                    "ai_buzz_index_industry"
                )
                fig, ax = plt.subplots(figsize=(11, 7))
                ax.barh(
                    df_i["industry"],
                    df_i["ai_buzz_index_industry"],
                    color=ACCENT, edgecolor="none", height=0.65, alpha=0.85,
                )
                ax.set_title(
                    f"Industry AI Buzz Index ({latest or 'all years'})"
                )
                ax.set_xlabel("Average AI Buzz Index")
                ax.set_facecolor("#f8f9fa")
                ax.tick_params(axis="y", labelsize=9)
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)

        # ── Page 7: Sentiment + Language (2 subplots) ─────────────────────────
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 5))

        # Sentiment pie
        if not df_refs.empty and "sentiment" in df_refs.columns:
            counts = df_refs["sentiment"].value_counts()
            sent_colors = {
                "positive": "#1a936f", "neutral": "#4aa3df", "negative": "#e76f51"
            }
            labels = [s.capitalize() for s in counts.index]
            colors_pie = [sent_colors.get(s, "#aaa") for s in counts.index]
            wedges, texts, autotexts = ax1.pie(
                counts.values, labels=labels, colors=colors_pie,
                autopct="%1.1f%%", startangle=90,
                wedgeprops=dict(edgecolor="white", linewidth=2),
            )
            for at in autotexts:
                at.set_fontsize(9)
            ax1.set_title("Sentiment Distribution")
        else:
            ax1.text(0.5, 0.5, "No sentiment data", ha="center", va="center")
            ax1.axis("off")

        # Language bar
        if not df_refs.empty and "language" in df_refs.columns:
            lang_map = {"en": "English", "zh": "Chinese", "ja": "Japanese",
                        "ko": "Korean", "other": "Other"}
            lang_counts = df_refs["language"].fillna("en").map(
                lambda x: lang_map.get(str(x), str(x))
            ).value_counts()
            lang_colors2 = [
                "#1f4e79", "#e63946", "#f4a261", "#1a936f", "#9467bd"
            ]
            bars = ax2.bar(
                lang_counts.index, lang_counts.values,
                color=lang_colors2[:len(lang_counts)], edgecolor="none", width=0.6,
            )
            for bar in bars:
                ax2.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height() + lang_counts.max() * 0.01,
                         f"{int(bar.get_height()):,}",
                         ha="center", va="bottom", fontsize=9)
            ax2.set_title("References by Document Language")
            ax2.set_ylabel("References")
            ax2.set_facecolor("#f8f9fa")
        else:
            ax2.text(0.5, 0.5, "No language data", ha="center", va="center")
            ax2.axis("off")

        fig.suptitle("Sentiment & Language Overview", fontsize=13, fontweight="bold", y=1.02)
        fig.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Attach metadata
        d = pdf.infodict()
        d["Title"]   = f"AISA Analysis Report v{AISA_VERSION}"
        d["Author"]  = "AISA — AI Semantic Analyzer"
        d["Subject"] = "Fortune 500 AI Disclosure Analysis"
        d["Creator"] = f"AISA v{AISA_VERSION}"

    return str(Path(output_path).resolve())


# ─────────────────────────────────────────────────────────────────────────────
# CLI smoke test
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=f"AISA v{AISA_VERSION} — Visualization Engine")
    parser.add_argument("db",   help="Path to aisa_results.db")
    parser.add_argument("--out", default="Results_AISA", help="Output directory")
    parser.add_argument("--year", type=int, default=None, help="Filter to a specific year")
    parser.add_argument("--excel", action="store_true", help="Export Excel charts")
    parser.add_argument("--html",  action="store_true", help="Export HTML dashboard")
    parser.add_argument("--pdf",   action="store_true", help="Export PDF report")
    args = parser.parse_args()

    do_all = not (args.excel or args.html or args.pdf)
    out    = Path(args.out)

    if args.excel or do_all:
        try:
            p = export_excel_charts(args.db, str(out / "aisa_charts.xlsx"), args.year)
            print(f"  Excel charts : {p}")
        except Exception as e:
            print(f"  Excel charts FAILED: {e}")

    if args.html or do_all:
        try:
            p = export_html_dashboard(args.db, str(out / "aisa_dashboard.html"), args.year)
            print(f"  HTML dashboard: {p}")
        except Exception as e:
            print(f"  HTML dashboard FAILED: {e}")

    if args.pdf or do_all:
        try:
            p = export_pdf_report(args.db, str(out / "aisa_report.pdf"), args.year)
            print(f"  PDF report   : {p}")
        except Exception as e:
            print(f"  PDF report FAILED: {e}")
