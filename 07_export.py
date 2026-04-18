from __future__ import annotations

import importlib
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

_m1 = importlib.import_module("01_models")
logger = _m1.logger
_m2 = importlib.import_module("02_db")
DatabaseManager = _m2.DatabaseManager
_m3 = importlib.import_module("03_taxonomy_base")
TaxonomyProvider = _m3.TaxonomyProvider
from version import AISA_VERSION

_AI_MARKER_RE = re.compile(r">>>(.*?)<<<", re.DOTALL)
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ufffe\uffff\ufffd]")
_CHAR_MAP = {
    "\u2022": "-", "\u25cf": "-", "\u25cb": "-", "\u25aa": "-",
    "\u2192": "->", "\u2013": "-", "\u2014": "-",
    "\u201c": '"',  "\u201d": '"', "\u2018": "'", "\u2019": "'",
    "\u2026": "...", "\u3000": " ",
}


def _sanitize(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    for src, dst in _CHAR_MAP.items():
        text = text.replace(src, dst)
    text = _ILLEGAL_CHARS_RE.sub("", text)
    text = re.sub(r" {2,}", " ", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if len(text) > 32000:
        text = text[:32000] + "..."
    return text.strip()


def strip_markers(text: str) -> str:
    if not text:
        return ""
    return _AI_MARKER_RE.sub(r"\1", text)


def _sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].map(lambda v: _sanitize(v) if isinstance(v, str) else v)
    return out


def _dimension_names(taxonomy: Optional[TaxonomyProvider], rows: List[dict]) -> List[str]:
    if taxonomy is not None:
        try:
            dims = list((taxonomy.get_dimensions() or {}).keys())
            if dims:
                return dims
        except Exception:
            pass
    names: List[str] = []
    seen = set()
    for row in rows:
        payload = row.get("dimensions_json") or ""
        if not payload:
            continue
        try:
            dims = json.loads(payload)
        except Exception:
            continue
        for name in dims.keys():
            if name not in seen:
                seen.add(name)
                names.append(name)
    return names


def _english_text(row: dict) -> str:
    lang = (row.get("language") or "en").lower()
    detected = row.get("text") or ""
    translated = row.get("text_translated") or ""
    if lang == "en":
        return detected or ""
    return translated or detected or ""


def _english_context(row: dict) -> str:
    lang = (row.get("language") or "en").lower()
    detected = row.get("context") or ""
    translated = row.get("context_translated") or ""
    if lang == "en":
        return detected or ""
    return translated or detected or ""


def _pages_identified(row: dict) -> str:
    val = row.get("pages")
    if val not in (None, ""):
        return str(val)
    val = row.get("page")
    if val not in (None, ""):
        return str(val)
    return ""


def _occurrences(row: dict) -> int:
    for key in ("occurrence_count", "total_occurrences", "Occurrences"):
        val = row.get(key)
        if val not in (None, ""):
            try:
                return int(val)
            except Exception:
                return 0
    return 0


def _build_simple_references_sheet(
    db: DatabaseManager,
    deduplicated: bool = False,
    year: Optional[int] = None,
    taxonomy: Optional[TaxonomyProvider] = None,
    company: Optional[str] = None,
) -> pd.DataFrame:
    rows = db.get_references(year=year, deduplicated=deduplicated, company=company)
    if not rows:
        base_cols = [
            "Company", "Year", "Industry", "Sector", "Country", "Document Type",
            "Language", "Detected Text", "Text (English)", "Context", "Context (English)",
            "Pages Identified", "Occurrences",
        ]
        dims = _dimension_names(taxonomy, [])
        for dim in dims:
            base_cols.extend([dim, f"{dim} Confidence"])
        return pd.DataFrame(columns=base_cols)

    dims = _dimension_names(taxonomy, rows)
    records: List[dict] = []
    for row in rows:
        rec = {
            "Company": row.get("company") or "",
            "Year": (int(row.get("year")) if row.get("year") not in (None, "") else ""),
            "Industry": row.get("industry") or "",
            "Sector": row.get("sector") or "",
            "Country": row.get("country") or "",
            "Document Type": row.get("doc_type") or "",
            "Language": row.get("language") or "en",
            "Detected Text": row.get("text") or "",
            "Text (English)": _english_text(row),
            "Context": row.get("context") or "",
            "Context (English)": _english_context(row),
            "Pages Identified": _pages_identified(row),
            "Occurrences": _occurrences(row),
        }
        payload = row.get("dimensions_json") or ""
        parsed = {}
        if payload:
            try:
                parsed = json.loads(payload) or {}
            except Exception:
                parsed = {}
        for dim in dims:
            value = parsed.get(dim)
            code = ""
            conf = ""
            if isinstance(value, (list, tuple)) and value:
                code = value[0] or ""
                conf = value[1] if len(value) > 1 and value[1] is not None else ""
            rec[dim] = code
            rec[f"{dim} Confidence"] = conf
        records.append(rec)

    df = pd.DataFrame(records)
    ordered = [
        "Company", "Year", "Industry", "Sector", "Country", "Document Type",
        "Language", "Detected Text", "Text (English)", "Context", "Context (English)",
        "Pages Identified", "Occurrences",
    ]
    for dim in dims:
        ordered.extend([dim, f"{dim} Confidence"])
    return df[[c for c in ordered if c in df.columns]]


def export_excel(
    db: DatabaseManager,
    output_path: str,
    year: Optional[int] = None,
    taxonomy: Optional[TaxonomyProvider] = None,
    export_type: str = "full",
    company: Optional[str] = None,
) -> str:
    output_path = str(Path(output_path).with_suffix(".xlsx"))
    logger.info(f"Exporting Excel -> {output_path}")
    dedup = export_type == "dedup"
    refs_df = _build_simple_references_sheet(db, deduplicated=dedup, year=year, taxonomy=taxonomy, company=company)
    safe_df = _sanitize_df_for_excel(refs_df)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        safe_df.to_excel(writer, sheet_name="References", index=False)
        ws = writer.sheets["References"]
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            fill = PatternFill("solid", fgColor="1F4E79")
            font = Font(color="FFFFFF", bold=True)
            align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = align
            ws.freeze_panes = "A2"
            widths = {
                "Company": 24, "Industry": 22, "Sector": 22, "Country": 16,
                "Document Type": 18, "Language": 10,
                "Detected Text": 60, "Text (English)": 60, "Context": 90, "Context (English)": 90,
                "Pages Identified": 16, "Occurrences": 12,
            }
            for idx, name in enumerate(safe_df.columns, start=1):
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(name, 22)
        except Exception as exc:
            logger.debug(f"Excel formatting partial: {exc}")
    size_kb = Path(output_path).stat().st_size / 1024
    logger.info(f"Excel export complete: {output_path} ({size_kb:.0f} KB)")
    return os.path.abspath(output_path)


def export_json(
    db: DatabaseManager,
    output_path: str,
    year: Optional[int] = None,
    pretty: bool = True,
    taxonomy: Optional[TaxonomyProvider] = None,
    export_type: str = "full",
    company: Optional[str] = None,
) -> str:
    output_path = str(Path(output_path).with_suffix(".json"))
    dedup = export_type == "dedup"
    df = _build_simple_references_sheet(db, deduplicated=dedup, year=year, taxonomy=taxonomy, company=company)
    payload = {
        "meta": {
            "aisa_version": AISA_VERSION,
            "export_date": datetime.now().isoformat(),
            "year_filter": year,
            "database": db.db_path,
            "deduplicated": dedup,
        },
        "references": [{k: (None if pd.isna(v) else v) for k, v in row.items()} for row in df.to_dict(orient="records")],
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2 if pretty else None, default=str)
    return os.path.abspath(output_path)


def export_csv(
    db: DatabaseManager,
    output_dir: str,
    year: Optional[int] = None,
    prefix: str = "aisa",
    taxonomy: Optional[TaxonomyProvider] = None,
    export_type: str = "full",
    company: Optional[str] = None,
) -> List[str]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    dedup = export_type == "dedup"
    df = _build_simple_references_sheet(db, deduplicated=dedup, year=year, taxonomy=taxonomy, company=company)
    suffix = "references_dedup" if dedup else "references_raw"
    path = out_dir / f"{prefix}_{suffix}.csv"
    _sanitize_df_for_excel(df).to_csv(path, index=False, encoding="utf-8-sig")
    return [str(path.absolute())]


def export_all(
    db: DatabaseManager,
    output_dir: str,
    base_name: str = "aisa_results",
    year: Optional[int] = None,
    formats: Optional[List[str]] = None,
    taxonomy: Optional[TaxonomyProvider] = None,
    export_type: str = "full",
    company: Optional[str] = None,
) -> Dict[str, object]:
    if formats is None:
        formats = ["excel", "json", "csv"]
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    result: Dict[str, object] = {"excel": None, "json": None, "csv": None}
    suffix = "dedup" if export_type == "dedup" else "raw"
    stem = f"{base_name}_{suffix}" if export_type in {"raw", "dedup"} else base_name
    if "excel" in formats:
        result["excel"] = export_excel(db, str(out_dir / f"{stem}.xlsx"), year=year, taxonomy=taxonomy, export_type=export_type, company=company)
    if "json" in formats:
        result["json"] = export_json(db, str(out_dir / f"{stem}.json"), year=year, taxonomy=taxonomy, export_type=export_type, company=company)
    if "csv" in formats:
        result["csv"] = export_csv(db, str(out_dir / "csv"), year=year, prefix=stem, taxonomy=taxonomy, export_type=export_type, company=company)
    return result
