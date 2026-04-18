"""
===============================================================================
AISA - AI Semantic Analyzer
09_tpdi.py - Technology-Product Diffusion Index (TPDI) v1.0
===============================================================================

Measures how widely SPECIFIC AI products (ChatGPT, Copilot, Gemini, Claude,
TensorFlow, Proprietary AI, etc.) are adopted across Fortune 500 companies.

Methodology (product-centric, consistent with v6.3):
    For each specific product (ex: ChatGPT) and each year:
        TPDI(product, year) = (active_adopters x 1pt) + (mature_adopters x 2pt)

    A company is "mature" if: year - first_seen_year >= MATURITY_THRESHOLD (5 ani)

    Aggregation:
        category_scores = suma TPDI tuturor produselor din categoria B
        vendor_scores   = suma TPDI tuturor produselor unui vendor

    Product extraction uses ai_products_v1 (100+ known products) with temporal
    validation -- references where year < product_release_year are skipped.

DB reads from:
    adoption_portfolio_b (company, category_b, product_name, product_vendor,
                          first_seen_year, last_confirmed_year, discontinued_year, status)
    NOTE: there is NO 'year' column in this table. The active interval
    [first_seen_year, last_confirmed_year] is expanded row-by-row into one
    synthetic entry per year inside calculate_from_db() before scoring.

CHANGELOG:
    v1.0.0 (2026-02) - AISA v1.0 -- rewritten from category-based to product-centric.
                        Methodology aligned with v6.3 (ai_tpdi.py).
                        Integrated ai_products_v1 for extraction + temporal validation.
                        Added calculate_from_db(db) reading from DatabaseManager directly.

Author: TeRa0
License: MIT
===============================================================================
"""

from __future__ import annotations

import importlib
import json
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from version import TPDI_VERSION

_m1 = importlib.import_module("01_models")
logger          = _m1.logger

_m2 = importlib.import_module("02_db")
DatabaseManager = _m2.DatabaseManager

# ---------------------------------------------------------------------------
# ai_products_v1 integration — HARD FAIL if not present.
# TPDI without product extraction is methodologically meaningless
# (it would only count category-level entries, which Buzz Index already does).
# ---------------------------------------------------------------------------
try:
    _ap = importlib.import_module("14_ai_products_v1")
except ImportError as e:
    raise RuntimeError(
        "14_ai_products_v1 is required for TPDI but was not found. "
        "Ensure 14_ai_products_v1.py is in the same directory as 09_tpdi.py. "
        f"Original error: {e}"
    ) from e

try:
    extract_product_info     = _ap.extract_product_info
    GranularityLevel         = _ap.GranularityLevel
    KNOWN_PRODUCTS           = _ap.KNOWN_PRODUCTS
    VENDOR_PATTERNS          = _ap.VENDOR_PATTERNS
    get_vendor_for_product   = _ap.get_vendor_for_product
    get_product_category     = _ap.get_product_category
    get_product_release_year = _ap.get_product_release_year
except AttributeError as e:
    raise RuntimeError(
        f"14_ai_products_v1 is missing required attributes: {e}. "
        "Ensure you are using the correct version of ai_products_v1.py."
    ) from e

logger.info("14_ai_products_v1 loaded OK (%d known products)", len(KNOWN_PRODUCTS))

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MATURITY_THRESHOLD_YEARS = 5
MATURITY_BONUS_POINTS    = 2
ADOPTION_BASE_POINTS     = 1
PROPRIETARY_LABEL        = "Proprietary AI"
PROPRIETARY_VENDOR       = "Proprietary"

DISCONTINUATION_PATTERNS = [
    r'\b(?:discontinued|stopped|ceased|ended)\s+(?:using|utilizing|leveraging)\b',
    r'\bno\s+longer\s+(?:use|using|utilize|leveraging)\b',
    r'\bphased?\s+out\b',
    r'\bdeprecated\b',
    r'\bsunset(?:ting|ted)?\b',
]

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ProductAdoption:
    """Tracks a single company's adoption history for a specific product."""
    company:              str
    product:              str
    vendor:               str           = ""
    category:             str           = ""
    granularity:          str           = ""
    first_seen_year:      int           = 0
    last_confirmed_year:  int           = 0
    status:               str           = "ACTIVE"
    discontinued_year:    Optional[int] = None
    years_active:         List[int]     = field(default_factory=list)


@dataclass
class TPDIScore:
    """TPDI score for a product (or category / vendor) in a given year."""
    product:             str
    year:                int
    vendor:              str       = ""
    category:            str       = ""
    adopter_count:       int       = 0
    mature_count:        int       = 0
    discontinued_count:  int       = 0
    base_points:         float     = 0.0
    maturity_points:     float     = 0.0
    tpdi:                float     = 0.0
    adopters:            List[str] = field(default_factory=list)
    mature_adopters:     List[str] = field(default_factory=list)
    discontinued_by:     List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# TPDICalculator
# ---------------------------------------------------------------------------

class TPDICalculator:
    """
    Product-centric TPDI calculator.

    Usage:
        calc    = TPDICalculator()
        results = calc.calculate_from_db(db)
        calc.export_to_excel(results, "tpdi_output.xlsx")
    """

    def __init__(self, maturity_years: int = MATURITY_THRESHOLD_YEARS):
        self.maturity_years = maturity_years

    # ------------------------------------------------------------------
    # Product extraction helper
    # ------------------------------------------------------------------

    def _extract_product(
        self,
        text: str,
        context: str = "",
        category_b: Optional[str] = None,
        report_year: Optional[int] = None,
    ) -> Tuple[Optional[str], Optional[str], str, str]:
        """
        Wrapper over 14_ai_products_v1.extract_product_info().

        Returns:
            (product, vendor, category_b_resolved, granularity_str)
        """
        product, vendor, granularity = extract_product_info(
            text, category_b, context, report_year=report_year
        )
        if granularity == GranularityLevel.SPECIFIC:
            cat = get_product_category(product) or category_b or ""
            return product, vendor, cat, "SPECIFIC"
        elif granularity == GranularityLevel.INTERNAL:
            return PROPRIETARY_LABEL, PROPRIETARY_VENDOR, category_b or "", "INTERNAL"
        elif granularity == GranularityLevel.VENDOR_ONLY:
            if vendor and vendor in VENDOR_PATTERNS:
                defaults = VENDOR_PATTERNS[vendor].get("default_products", [])
                if defaults:
                    cat = VENDOR_PATTERNS[vendor].get("default_category_b", category_b or "")
                    return defaults[0], vendor, cat, "VENDOR_ONLY"
            return (f"{vendor} AI" if vendor else None), vendor, category_b or "", "VENDOR_ONLY"

        # Proprietary / internal detection (CATEGORY_ONLY from extract_product_info)
        full = f"{text} {context}".lower()
        for pat in [
            r'\b(?:proprietary|in[\s-]?house|internal|custom[\s-]?built)\s+(?:AI|ML|model|platform|solution)\b',
            r'\b(?:our|their)\s+(?:own|proprietary)\s+(?:AI|ML|algorithm|model)\b',
            r'\bdeveloped\s+(?:internally|in[\s-]?house)\b.*(?:AI|ML|model)',
        ]:
            if re.search(pat, full, re.IGNORECASE):
                return PROPRIETARY_LABEL, PROPRIETARY_VENDOR, category_b or "", "INTERNAL"

        return None, None, category_b or "", "CATEGORY_ONLY"

    # ------------------------------------------------------------------
    # Main calculation methods
    # ------------------------------------------------------------------

    def calculate_from_db(self, db: "DatabaseManager") -> Dict:
        """
        Read adoption_portfolio_b directly from DatabaseManager and calculate TPDI.
        """
        rows = db.conn.execute(
            "SELECT company, category_b, product_name, product_vendor,"
            "       first_seen_year, last_confirmed_year, discontinued_year, status"
            " FROM adoption_portfolio_b"
        ).fetchall()

        if not rows:
            logger.warning("adoption_portfolio_b is empty — TPDI has no data")
            return self._empty_result()

        import pandas as pd
        records = []
        for row in rows:
            first = int(row["first_seen_year"])
            last  = int(row["last_confirmed_year"])
            disc  = row["discontinued_year"]
            if disc is not None:
                last = min(last, int(disc) - 1)
            # When no specific product, use category_b as the tracking unit
            product_name = row["product_name"] or row["category_b"] or ""
            for yr in range(first, last + 1):
                records.append({
                    "Company":             row["company"],
                    "Year":                yr,
                    "Text":                product_name,
                    "Product_Name":        product_name,
                    "Product_Vendor":      row["product_vendor"] or "",
                    "Category_Technology": row["category_b"] or "",
                })

        if not records:
            logger.warning("adoption_portfolio_b has entries but no active year ranges")
            return self._empty_result()

        return self.calculate_from_df(pd.DataFrame(records))

    def calculate_from_df(self, df) -> Dict:
        """
        Core TPDI calculation from a DataFrame.

        Expected columns:
            Company, Year, Text, Category_Technology
            Optional: Product_Name, Product_Vendor, Context

        Returns dict with keys:
            product_scores, category_scores, vendor_scores,
            adopter_detail, extraction_log, adoptions, metadata
        """
        import pandas as pd

        col_map = {
            "Category_Tech":  "Category_Technology",
            "Category_App":   "Category_Application",
            "Reference_Text": "Text",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        text_col        = "Text" if "Text" in df.columns else "Reference_Text"
        context_col     = "Context" if "Context" in df.columns else None
        cat_col         = "Category_Technology" if "Category_Technology" in df.columns else None
        has_product_col = "Product_Name" in df.columns and df["Product_Name"].notna().any()

        all_years       = sorted(int(y) for y in df["Year"].unique())
        global_min_year = int(min(all_years))
        global_max_year = int(max(all_years))

        adoptions:        Dict[Tuple[str, str], ProductAdoption] = {}
        extraction_log:   List[Dict]       = []
        extraction_stats: Dict[str, int]   = defaultdict(int)

        # ── Pass 1: build adoptions ───────────────────────────────────────
        for _, row in df.iterrows():
            company = str(row["Company"])
            year    = int(row["Year"])
            text    = str(row.get(text_col, ""))
            context = str(row.get(context_col, "")) if context_col else ""
            cat_b   = str(row.get(cat_col, "")) if cat_col and pd.notna(row.get(cat_col)) else ""

            if has_product_col and pd.notna(row.get("Product_Name")):
                product     = str(row["Product_Name"])
                vendor      = str(row.get("Product_Vendor", "")) if pd.notna(row.get("Product_Vendor")) else ""
                granularity = "PRE_EXTRACTED"
                if product:
                    # Skip release-year check for category codes (e.g. "B4_GenAI_LLMs")
                    # — these are not real product names and have no release year.
                    is_category_code = bool(product and len(product) > 1 and product[1:2] == "_" and product[0] in "AB")
                    if not is_category_code:
                        release_yr = get_product_release_year(product)
                        if release_yr and year < release_yr:
                            product = None
                            vendor  = None
                            granularity = "BLOCKED_ANACHRONISM"
            else:
                product, vendor, cat_b_ext, granularity = self._extract_product(
                    text, context, cat_b, report_year=year
                )
                if cat_b_ext:
                    cat_b = cat_b_ext

            extraction_log.append({
                "Company":           company,
                "Year":              year,
                "Text":              text[:100],
                "Context":           context[:150],
                "Category_B":        cat_b,
                "Extracted_Product": product or "(none)",
                "Extracted_Vendor":  vendor or "",
                "Granularity":       granularity,
            })
            extraction_stats[granularity] += 1

            if not product:
                continue

            key = (company, product)
            if key not in adoptions:
                adoptions[key] = ProductAdoption(
                    company=company, product=product, vendor=vendor or "",
                    category=cat_b, granularity=granularity,
                    first_seen_year=year, last_confirmed_year=year,
                    years_active=[year],
                )
            else:
                a = adoptions[key]
                a.last_confirmed_year = max(a.last_confirmed_year, year)
                if year not in a.years_active:
                    a.years_active.append(year)
                    a.years_active.sort()
                if vendor and not a.vendor:
                    a.vendor = vendor
                if cat_b and not a.category:
                    a.category = cat_b

        # ── Pass 2: discontinuation detection ────────────────────────────
        self._detect_discontinuations(df, adoptions, text_col, context_col, cat_col)

        # ── Pass 3: score per product per year ────────────────────────────
        product_scores: Dict[str, Dict[int, TPDIScore]] = defaultdict(dict)
        adopter_detail: List[Dict] = []
        all_products = sorted({a.product for a in adoptions.values()})

        for year in range(global_min_year, global_max_year + 1):
            for product in all_products:
                pa = [a for a in adoptions.values() if a.product == product]
                if not pa:
                    continue
                adopters, mature_adopters, discontinued_by = [], [], []
                vendor, category = "", ""

                for adopt in pa:
                    vendor   = adopt.vendor   or vendor
                    category = adopt.category or category

                    is_active = (
                        adopt.first_seen_year <= year
                        and adopt.status == "ACTIVE"
                        and (adopt.discontinued_year is None or adopt.discontinued_year > year)
                    )
                    if not is_active and year in adopt.years_active:
                        is_active = True
                    is_disc = (
                        adopt.status == "DISCONTINUED"
                        and adopt.discontinued_year is not None
                        and adopt.discontinued_year <= year
                    )

                    if is_active:
                        adopters.append(adopt.company)
                        years_using = year - adopt.first_seen_year
                        is_mature   = years_using >= self.maturity_years
                        if is_mature:
                            mature_adopters.append(adopt.company)
                        adopter_detail.append({
                            "Product":              product,
                            "Vendor":               vendor,
                            "Category":             category,
                            "Granularity":          adopt.granularity,
                            "Company":              adopt.company,
                            "Year":                 year,
                            "First_Seen":           adopt.first_seen_year,
                            "Years_Using":          years_using,
                            "Is_Mature":            is_mature,
                            "Status":               "ACTIVE",
                            "Base_Points":          ADOPTION_BASE_POINTS,
                            "Maturity_Points":      MATURITY_BONUS_POINTS if is_mature else 0,
                            "Company_Contribution": ADOPTION_BASE_POINTS + (MATURITY_BONUS_POINTS if is_mature else 0),
                        })
                    elif is_disc:
                        discontinued_by.append(adopt.company)
                        adopter_detail.append({
                            "Product":              product,
                            "Vendor":               vendor,
                            "Category":             category,
                            "Granularity":          adopt.granularity,
                            "Company":              adopt.company,
                            "Year":                 year,
                            "First_Seen":           adopt.first_seen_year,
                            "Years_Using":          0,
                            "Is_Mature":            False,
                            "Status":               "DISCONTINUED",
                            "Base_Points":          0,
                            "Maturity_Points":      0,
                            "Company_Contribution": 0,
                        })

                bp = len(adopters)        * ADOPTION_BASE_POINTS
                mp = len(mature_adopters) * MATURITY_BONUS_POINTS
                if bp + mp > 0 or discontinued_by:
                    product_scores[product][year] = TPDIScore(
                        product=product, year=year, vendor=vendor, category=category,
                        adopter_count=len(adopters), mature_count=len(mature_adopters),
                        discontinued_count=len(discontinued_by),
                        base_points=bp, maturity_points=mp, tpdi=bp + mp,
                        adopters=sorted(adopters),
                        mature_adopters=sorted(mature_adopters),
                        discontinued_by=sorted(discontinued_by),
                    )

        category_scores = self._aggregate_scores(product_scores, adoptions, "category")
        vendor_scores   = self._aggregate_scores(product_scores, adoptions, "vendor")

        return {
            "product_scores":  dict(product_scores),
            "category_scores": dict(category_scores),
            "vendor_scores":   dict(vendor_scores),
            "adopter_detail":  adopter_detail,
            "extraction_log":  extraction_log,
            "adoptions":       adoptions,
            "metadata": {
                "version":              TPDI_VERSION,
                "date":                 datetime.now().isoformat(),
                "total_references":     len(df),
                "total_products_found": len(all_products),
                "total_companies":      len({a.company for a in adoptions.values()}),
                "total_adoptions":      len(adoptions),
                "year_range":           f"{global_min_year}-{global_max_year}",
                "maturity_threshold":   f"{self.maturity_years} years",
                "products_module":      "14_ai_products_v1",
                "discontinuations":     sum(1 for a in adoptions.values() if a.status == "DISCONTINUED"),
                "proprietary_ai_count": sum(1 for a in adoptions.values() if a.product == PROPRIETARY_LABEL),
                "extraction_stats":     dict(extraction_stats),
            },
        }

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _detect_discontinuations(self, df, adoptions, text_col, context_col, cat_col):
        import pandas as pd
        for _, row in df.iterrows():
            text    = str(row.get(text_col, ""))
            context = str(row.get(context_col, "")) if context_col else ""
            combined = f"{text} {context}".lower()
            if not any(re.search(p, combined, re.IGNORECASE) for p in DISCONTINUATION_PATTERNS):
                continue
            company = str(row["Company"])
            year    = int(row["Year"])
            cat_b   = str(row.get(cat_col, "")) if cat_col and pd.notna(row.get(cat_col)) else ""
            product, _, _, _ = self._extract_product(text, context, cat_b, report_year=year)
            if product:
                key = (company, product)
                if key in adoptions:
                    adoptions[key].status = "DISCONTINUED"
                    adoptions[key].discontinued_year = year

    def _aggregate_scores(
        self,
        product_scores: Dict,
        adoptions: Dict,
        group_by: str,
    ) -> Dict[str, Dict[int, "TPDIScore"]]:
        product_group = {}
        for a in adoptions.values():
            product_group[a.product] = (a.category or a.product) if group_by == "category" else (a.vendor or "Unknown")

        grouped: Dict = defaultdict(lambda: defaultdict(lambda: {"adopters": set(), "mature": set(), "disc": set()}))
        for product, yd in product_scores.items():
            grp = product_group.get(product, product)
            if not grp or grp == "Unknown":
                continue
            for year, s in yd.items():
                g = grouped[grp][year]
                g["adopters"].update(s.adopters)
                g["mature"].update(s.mature_adopters)
                g["disc"].update(s.discontinued_by)

        result: Dict[str, Dict[int, TPDIScore]] = {}
        for grp, yd in grouped.items():
            result[grp] = {}
            for year, d in yd.items():
                ad = sorted(d["adopters"])
                ma = sorted(d["mature"])
                di = sorted(d["disc"] - d["adopters"])
                bp = len(ad) * ADOPTION_BASE_POINTS
                mp = len(ma) * MATURITY_BONUS_POINTS
                result[grp][year] = TPDIScore(
                    product=grp, year=year,
                    adopter_count=len(ad), mature_count=len(ma), discontinued_count=len(di),
                    base_points=bp, maturity_points=mp, tpdi=bp + mp,
                    adopters=ad, mature_adopters=ma, discontinued_by=di,
                )
        return result

    @staticmethod
    def _empty_result() -> Dict:
        return {
            "product_scores":  {},
            "category_scores": {},
            "vendor_scores":   {},
            "adopter_detail":  [],
            "extraction_log":  [],
            "adoptions":       {},
            "metadata":        {},
        }

    # ------------------------------------------------------------------
    # Excel export (3 heatmap matrices + detail sheets)
    # ------------------------------------------------------------------

    def export_to_excel(self, results: Dict, filepath: str) -> None:
        """
        Export TPDI results to Excel.

        Sheets:
            TPDI_Product_Matrix   - products x years heatmap
            TPDI_Category_Matrix  - categories x years heatmap
            TPDI_Vendor_Matrix    - vendors x years heatmap
            TPDI_Product_Detail   - per-product per-year scores
            TPDI_Adopter_Detail   - per-company per-product per-year
            TPDI_Extraction_Log   - extraction audit log
            TPDI_Formula          - methodology documentation
            TPDI_Assumptions      - assumptions documentation
        """
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter

        ps = results["product_scores"]
        cs = results["category_scores"]
        vs = results["vendor_scores"]

        df_pm = self._build_matrix(ps, "Product")
        df_cm = self._build_matrix(cs, "Category")
        df_vm = self._build_matrix(vs, "Vendor")

        detail_data = []
        for product, yd in sorted(ps.items()):
            for year in sorted(yd.keys()):
                s = yd[year]
                detail_data.append({
                    "Product":            product,
                    "Vendor":             s.vendor,
                    "Category":           s.category,
                    "Year":               year,
                    "Adopter_Count":      s.adopter_count,
                    "Mature_Count":       s.mature_count,
                    "Discontinued_Count": s.discontinued_count,
                    "Base_Points":        s.base_points,
                    "Maturity_Points":    s.maturity_points,
                    "TPDI":               s.tpdi,
                    "Adopters":           ", ".join(s.adopters),
                    "Mature_Adopters":    ", ".join(s.mature_adopters),
                })

        df_formula = pd.DataFrame([
            {"Component": "TPDI Formula",    "Calculation": "TPDI(p,t) = adopters x 1 + mature_adopters x 2",               "Description": "Score for product p in year t"},
            {"Component": "Base Adoption",   "Calculation": "+1 pt per active adopter company",                              "Description": "Each company using product = 1 pt"},
            {"Component": "Maturity Bonus",  "Calculation": f"+{MATURITY_BONUS_POINTS} pts after {self.maturity_years} years", "Description": "Embedded technology signal"},
            {"Component": "Decommission",    "Calculation": "discontinued -> 0 pts",                                         "Description": "Product no longer in use"},
            {"Component": "Product Extract", "Calculation": "14_ai_products_v1.extract_product_info(text, context)",            "Description": "SPECIFIC/INTERNAL/VENDOR_ONLY/CATEGORY_ONLY"},
            {"Component": "Proprietary AI",  "Calculation": "In-house patterns -> 'Proprietary AI'",                         "Description": "Custom-built AI solutions tracked separately"},
            {"Component": "Temporal Valid",  "Calculation": "skip if year < get_product_release_year(product)",              "Description": "No anachronistic references allowed"},
        ])
        df_assumptions = pd.DataFrame([
            {"Assumption": "Silence != Discontinuation", "Justification": "Products persist unless explicitly discontinued",       "Risk_Level": "Low"},
            {"Assumption": f"Maturity = {self.maturity_years} years", "Justification": "Technology embedded after continuous use", "Risk_Level": "Medium"},
            {"Assumption": "Equal company weight",        "Justification": "Each company = 1 pt regardless of size",              "Risk_Level": "Medium"},
            {"Assumption": "Proprietary AI as single",    "Justification": "All in-house solutions grouped together",             "Risk_Level": "Low"},
        ])

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_pm.to_excel(writer, sheet_name="TPDI_Product_Matrix",   index=False)
            df_cm.to_excel(writer, sheet_name="TPDI_Category_Matrix",  index=False)
            df_vm.to_excel(writer, sheet_name="TPDI_Vendor_Matrix",    index=False)
            if detail_data:
                pd.DataFrame(detail_data).to_excel(writer, sheet_name="TPDI_Product_Detail",  index=False)
            if results.get("adopter_detail"):
                pd.DataFrame(results["adopter_detail"]).to_excel(writer, sheet_name="TPDI_Adopter_Detail", index=False)
            if results.get("extraction_log"):
                pd.DataFrame(results["extraction_log"]).to_excel(writer, sheet_name="TPDI_Extraction_Log", index=False)
            df_formula.to_excel(writer,     sheet_name="TPDI_Formula",     index=False)
            df_assumptions.to_excel(writer, sheet_name="TPDI_Assumptions", index=False)

        # Green heatmap (white -> light green -> dark green)
        wb = load_workbook(filepath)
        for sn, dm in [("TPDI_Product_Matrix", df_pm), ("TPDI_Category_Matrix", df_cm), ("TPDI_Vendor_Matrix", df_vm)]:
            if sn not in wb.sheetnames or dm.empty:
                continue
            ws = wb[sn]
            yc = [c for c in dm.columns if str(c).isdigit()]
            if not yc:
                continue
            fc  = dm.columns.get_loc(yc[0])  + 2
            lc  = dm.columns.get_loc(yc[-1]) + 2
            rule = ColorScaleRule(
                start_type="num",      start_value=0,  start_color="FFFFFF",
                mid_type="percentile", mid_value=50,   mid_color="C6EFCE",
                end_type="max",                        end_color="1E7B34",
            )
            ws.conditional_formatting.add(
                f"{get_column_letter(fc)}2:{get_column_letter(lc)}{len(dm)+1}", rule
            )
        wb.save(filepath)
        logger.info("TPDI exported to %s", filepath)

    def _build_matrix(self, scores: Dict, label: str):
        import pandas as pd
        if not scores:
            return pd.DataFrame()
        all_years = sorted({y for yd in scores.values() for y in yd.keys()})
        rows = []
        for entity in sorted(scores.keys()):
            row = {label: entity}
            for year in all_years:
                row[year] = scores[entity].get(year, TPDIScore(entity, year)).tpdi
            rows.append(row)
        return pd.DataFrame(rows)


# ============================================================================
# SMOKE TEST
# ============================================================================

if __name__ == "__main__":
    import tempfile
    from version import get_version_string

    print(get_version_string())
    print(f"  TPDI version: {TPDI_VERSION}")
    print()

    try:
        import pandas as pd
    except ImportError:
        print("pandas not available -- skipping smoke test")
        raise SystemExit(0)

    test_data = [
        {"Company": "Microsoft", "Year": 2023, "Text": "ChatGPT",    "Context": "We deployed ChatGPT for customer support",    "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Microsoft", "Year": 2023, "Text": "Copilot",    "Context": "Microsoft Copilot across finance",             "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Microsoft", "Year": 2024, "Text": "Copilot",    "Context": "Expanded Copilot usage",                       "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Apple",     "Year": 2024, "Text": "ChatGPT",    "Context": "Integrated ChatGPT into products",             "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Alphabet",  "Year": 2024, "Text": "Gemini",     "Context": "Gemini powers our search",                     "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Alphabet",  "Year": 2023, "Text": "AI",         "Context": "Google AI platform analytics",                 "Category_Technology": "B7_Infrastructure_Platforms"},
        {"Company": "Toyota",    "Year": 2023, "Text": "ML",         "Context": "Our proprietary ML platform for manufacturing", "Category_Technology": "B1_Traditional_ML"},
        {"Company": "Shell",     "Year": 2023, "Text": "AI",         "Context": "Developed in-house AI solution",               "Category_Technology": "B8_General_AI"},
        {"Company": "Shell",     "Year": 2024, "Text": "Claude",     "Context": "Using Claude from Anthropic for analysis",     "Category_Technology": "B4_GenAI_LLMs"},
        {"Company": "Toyota",    "Year": 2018, "Text": "TensorFlow", "Context": "TensorFlow in production ML",                  "Category_Technology": "B1_Traditional_ML"},
        {"Company": "Toyota",    "Year": 2024, "Text": "TensorFlow", "Context": "TensorFlow still in production",               "Category_Technology": "B1_Traditional_ML"},
    ]

    df   = pd.DataFrame(test_data)
    calc = TPDICalculator()
    results = calc.calculate_from_df(df)

    print(f"Extraction Stats:         {results['metadata']['extraction_stats']}")
    print(f"Proprietary AI companies: {results['metadata']['proprietary_ai_count']}")
    print()
    print("Product TPDI (latest year):")
    for product, yd in sorted(results["product_scores"].items()):
        ly = max(yd.keys())
        s  = yd[ly]
        mat = f" [mature: {', '.join(s.mature_adopters)}]" if s.mature_adopters else ""
        print(f"  {product:<22} TPDI={s.tpdi:<5} ({s.adopter_count} adopters){mat}")

    print()
    print("Extraction Log:")
    for e in results["extraction_log"]:
        print(f"  {e['Company']:12} | {e['Text'][:25]:25} -> {e['Extracted_Product']:20} [{e['Granularity']}]")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out = tmp.name
    calc.export_to_excel(results, out)
    df_pm = pd.read_excel(out, sheet_name="TPDI_Product_Matrix")
    print(f"\nProduct Matrix ({len(df_pm)} rows):")
    print(df_pm.to_string())

    assert results["metadata"]["total_products_found"] > 0, "No products found"
    assert any(yd for yd in results["product_scores"].values()), "No product scores"
    print()
    print("  09_tpdi.py all checks passed.")


# ============================================================================
# REPORT WRAPPER  (consumed by 11_cli.py → cmd_tpdi)
# ============================================================================

@dataclass
class TPDICurve:
    key:             str
    kind:            str
    tpdi_score:      float
    max_penetration: float
    lifecycle_stage: str
    years:           List[int]        = field(default_factory=list)
    scores_by_year:  Dict[int, float] = field(default_factory=dict)

    def label(self) -> str:
        return self.key


@dataclass
class TPDIReport:
    curves:          List[TPDICurve]
    analysis_years:  List[int]
    total_companies: int
    raw:             Dict = field(default_factory=dict)

    def top_n(self, n: int = 10) -> List[TPDICurve]:
        return sorted(self.curves, key=lambda c: c.tpdi_score, reverse=True)[:n]

    def to_records(self) -> List[Dict]:
        return [
            {
                "key":             c.key,
                "kind":            c.kind,
                "tpdi_score":      round(c.tpdi_score, 4),
                "max_penetration": round(c.max_penetration, 4),
                "lifecycle_stage": c.lifecycle_stage,
                "first_year":      c.years[0]  if c.years else None,
                "last_year":       c.years[-1] if c.years else None,
            }
            for c in self.curves
        ]

    def yearly_to_records(self) -> List[Dict]:
        rows = []
        for c in self.curves:
            for yr, score in sorted(c.scores_by_year.items()):
                rows.append({"key": c.key, "kind": c.kind, "year": yr, "tpdi": round(score, 4)})
        return rows


def build_tpdi_report(
    db,
    config=None,
    include_products: bool = True,
    min_adopters: int = 3,
) -> "TPDIReport":
    calc    = TPDICalculator()
    results = calc.calculate_from_db(db)

    metadata        = results.get("metadata", {})
    total_companies = max(metadata.get("total_companies", 0), 1)
    product_scores  = results.get("product_scores", {})
    category_scores = results.get("category_scores", {})

    all_years: set = set()
    for yd in list(product_scores.values()) + list(category_scores.values()):
        all_years.update(yd.keys())
    analysis_years = sorted(all_years) if all_years else []

    max_pts = total_companies * (ADOPTION_BASE_POINTS + MATURITY_BONUS_POINTS)

    def _norm(raw: float) -> float:
        return min(raw / max_pts, 1.0) if max_pts > 0 else 0.0

    def _lifecycle(penetration: float, norm_tpdi: float) -> str:
        if penetration < 0.05:   return "emerging"
        if penetration >= 0.50 or norm_tpdi >= 0.60: return "mature"
        if penetration >= 0.20 or norm_tpdi >= 0.30: return "growth"
        return "declining"

    def _build_curve(key: str, kind: str, year_dict: Dict) -> "Optional[TPDICurve]":
        if not year_dict:
            return None
        latest        = year_dict[max(year_dict.keys())]
        adopter_count = getattr(latest, "adopter_count", 0)
        if adopter_count < min_adopters:
            return None
        norm_tpdi = _norm(float(getattr(latest, "tpdi", 0)))
        max_pen   = max(
            getattr(s, "adopter_count", 0) / total_companies
            for s in year_dict.values()
        )
        return TPDICurve(
            key=key, kind=kind,
            tpdi_score=norm_tpdi,
            max_penetration=max_pen,
            lifecycle_stage=_lifecycle(max_pen, norm_tpdi),
            years=sorted(year_dict.keys()),
            scores_by_year={
                yr: _norm(float(getattr(s, "tpdi", 0)))
                for yr, s in year_dict.items()
            },
        )

    curves: List[TPDICurve] = []
    for cat_key, yd in category_scores.items():
        c = _build_curve(cat_key, "category", yd)
        if c:
            curves.append(c)
    if include_products:
        for prod_key, yd in product_scores.items():
            c = _build_curve(prod_key, "product", yd)
            if c:
                curves.append(c)

    return TPDIReport(
        curves=curves,
        analysis_years=analysis_years,
        total_companies=total_companies,
        raw=results,
    )
