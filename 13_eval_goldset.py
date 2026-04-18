"""
===============================================================================
AISA - AI Semantic Analyzer
13_eval_goldset.py - Gold set evaluation: precision / recall / F1
===============================================================================

Evaluates the FP detection performance of the taxonomy on a gold set extracted from
the manual review file (S3-false_positives_only_manual.xlsx or equivalent).

What it does:
    1. Loads gold set (Excel or previously exported JSON/CSV)
    2. Runs TAXONOMY.check_false_positive() on each fragment
    3. (Optional) also runs semantic score via SentenceTransformer
    4. Calculates Precision / Recall / F1 for FP detection
    5. Per-category FP report (what types the taxonomy catches / misses)
    6. Saves JSON report + CSV with per-fragment predictions

Usage:
    # Quick evaluation (rule-only, without semantic):
    python 13_eval_goldset.py --input S3-false_positives_only_manual.xlsx --mode rules

    # Full evaluation (rules + semantic scoring):
    python 13_eval_goldset.py --input S3-false_positives_only_manual.xlsx --mode full

    # Evaluation on stratified subset (n fragments per FP category):
    python 13_eval_goldset.py --input S3-false_positives_only_manual.xlsx \\
        --mode rules --sample 50 --seed 42

    # A/B test semantic threshold:
    python 13_eval_goldset.py --input S3-false_positives_only_manual.xlsx \\
        --mode full --ab-threshold 0.60,0.68

    # Evaluation on JSON gold set (previously exported):
    python 13_eval_goldset.py --input goldset.json --mode rules

Accepted input formats:
    - .xlsx: Excel with columns from S3-false_positives_only_manual.xlsx
    - .json: list of dicts with keys: text, context, label (0=TP, 1=FP),
             fp_reason (optional), category (optional)
    - .csv:  columns: text, context, label, fp_reason, category

Output:
    - reports/eval_TIMESTAMP.json  — full report
    - reports/eval_TIMESTAMP.csv   — per-fragment predictions (for audit)

CHANGELOG:
    v1.0.0 (2026-03) - Initial release

Author: TeRa0
===============================================================================
"""

from __future__ import annotations

import argparse
import csv
import importlib
import json
import os
import sys
import time
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Path setup — allows running from any directory
# ---------------------------------------------------------------------------
_HERE = Path(__file__).parent.resolve()
sys.path.insert(0, str(_HERE))


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class GoldFragment:
    """A fragment from the gold set with manual label."""
    id:         str
    text:       str           # detected term (e.g.: "intelligent")
    context:    str           # context window (~5 sentences)
    label:      int           # 0 = True Positive (real AI), 1 = False Positive
    fp_reason:  str = ""      # FP reason from manual review
    category:   str = ""      # AISA category (A-dim)
    company:    str = ""
    year:       int = 0
    source:     str = ""      # Annual Report / Proxy / Sustainability
    # Fields filled by evaluator:
    pred_is_fp: Optional[bool] = None   # taxonomy prediction
    pred_fp_cat: str = ""               # detected FP category
    semantic_score: float = -1.0        # semantic score (if mode=full)
    correct:    Optional[bool] = None   # pred == label


@dataclass
class EvalReport:
    """Final evaluation report."""
    timestamp:      str
    taxonomy_version: str
    mode:           str           # "rules" or "full"
    n_total:        int
    n_fp:           int           # fragments with label=FP in gold set
    n_tp:           int           # fragments with label=TP in gold set
    # Global metrics (from the perspective of detecting FP as positive class)
    true_pos:       int = 0       # FP in gold set and detected as FP  (correct)
    false_neg:      int = 0       # FP in gold set but NOT caught (escaped)
    true_neg:       int = 0       # TP in gold set and not retained as FP (correct)
    false_pos:      int = 0       # TP in gold set but wrongly marked as FP
    precision:      float = 0.0   # TP / (TP + FP_pred)  [how precise the filters are]
    recall:         float = 0.0   # TP / (TP + FN)       [how many FPs are caught]
    f1:             float = 0.0
    accuracy:       float = 0.0
    fp_escape_rate: float = 0.0   # FPs that escape (false negatives) / total FP
    # Per-category FP
    by_fp_reason:   Dict[str, Dict] = field(default_factory=dict)
    by_fp_category: Dict[str, Dict] = field(default_factory=dict)
    # A/B threshold results (if --ab-threshold)
    ab_results:     Dict[str, Dict] = field(default_factory=dict)
    # Top misses (FPs that escape undetected — for debugging)
    top_misses:     List[Dict] = field(default_factory=list)
    # Top false alarms (TPs wrongly marked as FP)
    top_false_alarms: List[Dict] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Gold set loader
# ---------------------------------------------------------------------------

def load_xlsx(path: Path, sample: int = 0, seed: int = 42) -> List[GoldFragment]:
    """
    Loads from the Excel file S3-false_positives_only_manual.xlsx.

    The file contains ONLY FPs (Is_false_positive=True), with the exception of 2
    rows marked "Positive" in the Manual review Alexandra/Serban column.

    Gold set = all rows with complete manual review (both columns filled).
    Label:
        - "FP" / "FP"     → label=1 (confirmed false positive)
        - "Positive"       → label=0 (true positive wrongly classified)
        - "Borderline"/"?" → excluded from evaluation (ambiguous)
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Read header
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col = {h: i for i, h in enumerate(headers) if h}

    # Check required columns
    for req in ["Text", "Context"]:
        if req not in col:
            raise ValueError(f"Column '{req}' is missing from the Excel file.")

    has_manual = "Manual review Alexandra" in col or 18 < len(headers)
    ci_text    = col.get("Text", 7)
    ci_ctx     = col.get("Context", 8)
    ci_is_fp   = col.get("Is_false_positive", 14)
    ci_reason  = col.get("False_positive_reason", 15)
    ci_cat     = col.get("Category", 6)
    ci_company = col.get("Company name", 1)
    ci_year    = col.get("Year", 2)
    ci_source  = col.get("Sources", 9)
    # Manual review columns are fixed at 18 and 19 (0-indexed)
    CI_REV_A   = 18
    CI_REV_S   = 19

    fragments = []
    skipped_ambiguous = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) <= ci_text:
            continue

        text    = str(row[ci_text]    or "").strip()
        context = str(row[ci_ctx]     or "").strip() if len(row) > ci_ctx else ""

        if not text or not context:
            continue

        # Label from manual review
        rev_a = str(row[CI_REV_A] or "").strip() if len(row) > CI_REV_A else ""
        rev_s = str(row[CI_REV_S] or "").strip() if len(row) > CI_REV_S else ""

        if has_manual and (rev_a or rev_s):
            # Exclude rows without both reviews
            if not rev_a or not rev_s or rev_a in ("None", "") or rev_s in ("None", ""):
                continue
            # Exclude ambiguous
            if any(x in (rev_a, rev_s) for x in ["??", "?", "Borderline"]):
                skipped_ambiguous += 1
                continue
            label = 0 if "Positive" in (rev_a, rev_s) else 1
        else:
            # Without manual review — use Is_false_positive
            is_fp_raw = row[ci_is_fp] if len(row) > ci_is_fp else True
            label = 1 if str(is_fp_raw).lower() in ("true", "1", "yes") else 0

        fp_reason = str(row[ci_reason] or "").strip() if len(row) > ci_reason else ""
        category  = str(row[ci_cat]    or "").strip() if len(row) > ci_cat    else ""
        company   = str(row[ci_company] or "").strip() if len(row) > ci_company else ""
        year_raw  = row[ci_year] if len(row) > ci_year else None
        year      = int(year_raw) if year_raw and str(year_raw).isdigit() else 0
        source    = str(row[ci_source] or "").strip() if len(row) > ci_source else ""

        fragments.append(GoldFragment(
            id        = f"row_{row_idx}",
            text      = text,
            context   = context,
            label     = label,
            fp_reason = fp_reason,
            category  = category,
            company   = company,
            year      = year,
            source    = source,
        ))

    wb.close()

    if skipped_ambiguous:
        print(f"  [loader] Excluded {skipped_ambiguous} ambiguous rows (Borderline/??)")

    return _maybe_sample(fragments, sample, seed)


def load_json(path: Path, sample: int = 0, seed: int = 42) -> List[GoldFragment]:
    """
    Loads gold set from JSON.
    Expected format: list of objects with keys:
        text, context, label (0/1), fp_reason?, category?, company?, year?, source?
    """
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    fragments = []
    for i, d in enumerate(data):
        fragments.append(GoldFragment(
            id        = d.get("id", f"json_{i}"),
            text      = str(d.get("text", "")).strip(),
            context   = str(d.get("context", "")).strip(),
            label     = int(d.get("label", 1)),
            fp_reason = str(d.get("fp_reason", "")),
            category  = str(d.get("category", "")),
            company   = str(d.get("company", "")),
            year      = int(d.get("year", 0)),
            source    = str(d.get("source", "")),
        ))

    return _maybe_sample(fragments, sample, seed)


def load_csv(path: Path, sample: int = 0, seed: int = 42) -> List[GoldFragment]:
    """Loads gold set from CSV with header."""
    fragments = []
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            fragments.append(GoldFragment(
                id        = row.get("id", f"csv_{i}"),
                text      = str(row.get("text", "")).strip(),
                context   = str(row.get("context", "")).strip(),
                label     = int(row.get("label", 1)),
                fp_reason = str(row.get("fp_reason", "")),
                category  = str(row.get("category", "")),
                company   = str(row.get("company", "")),
                year      = int(row.get("year", 0)) if row.get("year", "").isdigit() else 0,
                source    = str(row.get("source", "")),
            ))
    return _maybe_sample(fragments, sample, seed)


def _maybe_sample(
    fragments: List[GoldFragment],
    sample: int,
    seed: int,
) -> List[GoldFragment]:
    """Stratified sampling by fp_reason if --sample is specified."""
    if sample <= 0 or sample >= len(fragments):
        return fragments

    import random
    rng = random.Random(seed)

    # Group by fp_reason
    by_reason: Dict[str, List] = defaultdict(list)
    for f in fragments:
        by_reason[f.fp_reason or "unknown"].append(f)

    # Distribute sample proportionally
    total = len(fragments)
    result = []
    for reason, items in by_reason.items():
        n = max(1, round(sample * len(items) / total))
        result.extend(rng.sample(items, min(n, len(items))))

    # Adjust to exact sample size
    if len(result) > sample:
        result = rng.sample(result, sample)

    return result


# ---------------------------------------------------------------------------
# Rule-based evaluation (TAXONOMY.check_false_positive)
# ---------------------------------------------------------------------------

def run_rules_eval(
    fragments: List[GoldFragment],
    taxonomy,
) -> None:
    """Runs check_false_positive on each fragment and fills in pred_*."""
    for frag in fragments:
        result = taxonomy.check_false_positive(frag.text, frag.context)
        frag.pred_is_fp  = result.is_fp
        frag.pred_fp_cat = result.category if result.is_fp else ""
        frag.correct     = (int(result.is_fp) == frag.label)


# ---------------------------------------------------------------------------
# Semantic evaluation (optional, mode=full)
# ---------------------------------------------------------------------------

def run_semantic_eval(
    fragments: List[GoldFragment],
    threshold: float,
) -> None:
    """
    Adds the semantic score for each fragment.
    A fragment is TP if semantic_score >= threshold (and not caught by rules).
    """
    try:
        from sentence_transformers import SentenceTransformer, util as st_util
    except ImportError:
        print("  [semantic] sentence_transformers not available — skip semantic scoring.")
        return

    _VERSION = importlib.import_module("version")
    model = SentenceTransformer(_VERSION.SEMANTIC_MODEL_NAME)

    # Anchor phrases (from 05_detect.py)
    AI_ANCHORS = [
        "artificial intelligence implementation",
        "machine learning model deployment",
        "deep learning neural network",
        "natural language processing NLP",
        "computer vision AI system",
        "generative AI large language model",
        "AI-powered automation",
        "predictive AI model",
        "AI investment strategy",
        "responsible AI governance",
    ]
    anchor_embs = model.encode(AI_ANCHORS, convert_to_tensor=True, normalize_embeddings=True)

    contexts = [f.context for f in fragments]
    print(f"  [semantic] Encoding {len(contexts)} fragments...")
    t0 = time.time()
    ctx_embs = model.encode(contexts, convert_to_tensor=True,
                            normalize_embeddings=True, batch_size=64,
                            show_progress_bar=True)
    print(f"  [semantic] Done in {time.time()-t0:.1f}s")

    import torch
    sims = st_util.cos_sim(ctx_embs, anchor_embs)  # [N, A]
    max_sims = sims.max(dim=1).values.cpu().tolist()

    for frag, score in zip(fragments, max_sims):
        frag.semantic_score = round(float(score), 4)
        # Combined decision: FP if rules say FP OR semantic below threshold
        if frag.pred_is_fp is None:
            # rules have not run yet — semantic only
            frag.pred_is_fp = (score < threshold)
        else:
            # rules + semantic: FP if rules catch it OR semantic below threshold
            frag.pred_is_fp = frag.pred_is_fp or (score < threshold)
        frag.correct = (int(frag.pred_is_fp) == frag.label)


# ---------------------------------------------------------------------------
# Metrics calculation
# ---------------------------------------------------------------------------

def compute_metrics(fragments: List[GoldFragment]) -> EvalReport:
    """Calculates all metrics from the completed predictions."""
    _ver = importlib.import_module("version")

    report = EvalReport(
        timestamp        = datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        taxonomy_version = _ver.TAXONOMY_VERSION,
        mode             = "rules",  # overwritten by caller
        n_total          = len(fragments),
        n_fp             = sum(1 for f in fragments if f.label == 1),
        n_tp             = sum(1 for f in fragments if f.label == 0),
    )

    for frag in fragments:
        pred = int(frag.pred_is_fp or False)
        actual = frag.label
        if actual == 1 and pred == 1:
            report.true_pos  += 1
        elif actual == 1 and pred == 0:
            report.false_neg += 1   # FP escaped (miss)
        elif actual == 0 and pred == 0:
            report.true_neg  += 1
        else:  # actual == 0 and pred == 1
            report.false_pos += 1   # TP wrongly filtered

    tp = report.true_pos
    fp = report.false_pos
    fn = report.false_neg
    tn = report.true_neg

    report.precision     = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    report.recall        = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    report.f1            = (2 * report.precision * report.recall /
                            (report.precision + report.recall)
                            if (report.precision + report.recall) > 0 else 0.0)
    report.accuracy      = (tp + tn) / len(fragments) if fragments else 0.0
    report.fp_escape_rate = fn / report.n_fp if report.n_fp > 0 else 0.0

    # Per fp_reason
    by_reason: Dict[str, Dict[str, int]] = defaultdict(lambda: {"total": 0, "caught": 0, "missed": 0})
    for frag in fragments:
        if frag.label == 1:  # are real FPs
            reason = frag.fp_reason or "unknown"
            by_reason[reason]["total"]  += 1
            if frag.pred_is_fp:
                by_reason[reason]["caught"] += 1
            else:
                by_reason[reason]["missed"] += 1
    for reason, d in by_reason.items():
        d["recall"] = round(d["caught"] / d["total"], 4) if d["total"] > 0 else 0.0
    report.by_fp_reason = dict(by_reason)

    # Per fp_category (AISA category)
    by_cat: Dict[str, Dict[str, int]] = defaultdict(lambda: {"total": 0, "caught": 0, "missed": 0})
    for frag in fragments:
        if frag.label == 1:
            cat = frag.category or "unknown"
            by_cat[cat]["total"]  += 1
            if frag.pred_is_fp:
                by_cat[cat]["caught"] += 1
            else:
                by_cat[cat]["missed"] += 1
    for cat, d in by_cat.items():
        d["recall"] = round(d["caught"] / d["total"], 4) if d["total"] > 0 else 0.0
    report.by_fp_category = dict(by_cat)

    # Top misses — escaped FPs (useful for debugging / pattern extension)
    misses = [f for f in fragments if f.label == 1 and not f.pred_is_fp]
    misses_sorted = sorted(misses, key=lambda f: f.fp_reason)
    report.top_misses = [
        {
            "id":        f.id,
            "text":      f.text,
            "fp_reason": f.fp_reason,
            "category":  f.category,
            "context":   f.context[:300],
            "company":   f.company,
            "year":      f.year,
        }
        for f in misses_sorted[:50]
    ]

    # Top false alarms — TPs wrongly marked as FP
    alarms = [f for f in fragments if f.label == 0 and f.pred_is_fp]
    report.top_false_alarms = [
        {
            "id":         f.id,
            "text":       f.text,
            "pred_fp_cat": f.pred_fp_cat,
            "category":   f.category,
            "context":    f.context[:300],
            "company":    f.company,
            "year":       f.year,
        }
        for f in alarms[:20]
    ]

    return report


# ---------------------------------------------------------------------------
# A/B test on semantic threshold
# ---------------------------------------------------------------------------

def run_ab_threshold(
    fragments: List[GoldFragment],
    thresholds: List[float],
    taxonomy,
) -> Dict[str, Dict]:
    """
    Runs evaluation for multiple semantic thresholds and
    returns comparative metrics.
    """
    results = {}
    # Save rules-only predictions to reuse them
    rules_preds = [(f.pred_is_fp, f.pred_fp_cat) for f in fragments]

    for thresh in thresholds:
        # Reset predictions to rules-only
        for frag, (pred, cat) in zip(fragments, rules_preds):
            frag.pred_is_fp  = pred
            frag.pred_fp_cat = cat

        run_semantic_eval(fragments, thresh)
        report = compute_metrics(fragments)

        results[str(thresh)] = {
            "threshold":     thresh,
            "precision":     round(report.precision, 4),
            "recall":        round(report.recall, 4),
            "f1":            round(report.f1, 4),
            "accuracy":      round(report.accuracy, 4),
            "fp_escape_rate": round(report.fp_escape_rate, 4),
            "true_pos":      report.true_pos,
            "false_neg":     report.false_neg,
            "true_neg":      report.true_neg,
            "false_pos":     report.false_pos,
        }

    # Restore rules-only predictions
    for frag, (pred, cat) in zip(fragments, rules_preds):
        frag.pred_is_fp  = pred
        frag.pred_fp_cat = cat

    return results


# ---------------------------------------------------------------------------
# Print report to console
# ---------------------------------------------------------------------------

def print_report(report: EvalReport) -> None:
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"  AISA Gold Set Evaluation — {report.timestamp}")
    print(f"  Taxonomy v{report.taxonomy_version}  |  Mode: {report.mode}")
    print(sep)
    print(f"\n  Gold set:  {report.n_total} fragments  "
          f"({report.n_fp} FP + {report.n_tp} TP)")
    print()
    print(f"  {'Metric':<28} {'Value':>10}")
    print(f"  {'-'*40}")
    print(f"  {'Precision (FP detection)':<28} {report.precision:>10.4f}")
    print(f"  {'Recall (FP detection)':<28} {report.recall:>10.4f}")
    print(f"  {'F1 score':<28} {report.f1:>10.4f}")
    print(f"  {'Accuracy':<28} {report.accuracy:>10.4f}")
    print(f"  {'FP escape rate (misses)':<28} {report.fp_escape_rate:>10.4f}")
    print()
    print(f"  Confusion matrix (FP = positive class):")
    print(f"    True Positives  (FP correctly caught):   {report.true_pos:>6}")
    print(f"    False Negatives (FP escaped):            {report.false_neg:>6}")
    print(f"    True Negatives  (TP correctly kept):     {report.true_neg:>6}")
    print(f"    False Positives (TP wrongly filtered):   {report.false_pos:>6}")

    print(f"\n  Recall per FP type (fp_reason):")
    by_r = sorted(report.by_fp_reason.items(), key=lambda x: -x[1]["total"])
    for reason, d in by_r:
        bar = "█" * int(d["recall"] * 20)
        print(f"    {reason[:45]:<45} {d['caught']:>4}/{d['total']:<4}  "
              f"{d['recall']:>5.1%}  {bar}")

    print(f"\n  Recall per AISA category (Category):")
    by_c = sorted(report.by_fp_category.items(), key=lambda x: -x[1]["total"])
    for cat, d in by_c:
        bar = "█" * int(d["recall"] * 20)
        print(f"    {cat[:40]:<40} {d['caught']:>4}/{d['total']:<4}  "
              f"{d['recall']:>5.1%}  {bar}")

    if report.top_misses:
        print(f"\n  Top missed FP ({len(report.top_misses)}) — "
              f"fragments that should be added to patterns:")
        for m in report.top_misses[:15]:
            print(f"    [{m['fp_reason'][:30]}] text={m['text']!r:<30}  "
                  f"ctx={m['context'][:80]!r}")

    if report.top_false_alarms:
        print(f"\n  False alarms ({len(report.top_false_alarms)}) — "
              f"TPs wrongly filtered by taxonomy:")
        for a in report.top_false_alarms[:10]:
            print(f"    [cat={a['pred_fp_cat'][:25]}] text={a['text']!r:<25}  "
                  f"ctx={a['context'][:80]!r}")

    if report.ab_results:
        print(f"\n  A/B Threshold comparison:")
        print(f"    {'Threshold':>10}  {'Precision':>10}  {'Recall':>10}  "
              f"{'F1':>8}  {'Escape%':>8}")
        print(f"    {'-'*55}")
        for thresh, r in sorted(report.ab_results.items(), key=lambda x: float(x[0])):
            print(f"    {float(thresh):>10.2f}  {r['precision']:>10.4f}  "
                  f"{r['recall']:>10.4f}  {r['f1']:>8.4f}  "
                  f"{r['fp_escape_rate']:>8.2%}")

    print(f"\n{sep}\n")


# ---------------------------------------------------------------------------
# Save report
# ---------------------------------------------------------------------------

def save_report(
    report: EvalReport,
    fragments: List[GoldFragment],
    out_dir: Path,
) -> Tuple[Path, Path]:
    """Saves JSON report + CSV with per-fragment predictions."""
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # JSON
    json_path = out_dir / f"eval_{ts}.json"
    report_dict = asdict(report)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report_dict, f, ensure_ascii=False, indent=2)

    # CSV with per-fragment predictions
    csv_path = out_dir / f"eval_{ts}_predictions.csv"
    fieldnames = [
        "id", "company", "year", "source", "text", "category",
        "fp_reason", "label", "pred_is_fp", "pred_fp_cat",
        "semantic_score", "correct", "context",
    ]
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for frag in fragments:
            writer.writerow({
                "id":            frag.id,
                "company":       frag.company,
                "year":          frag.year,
                "source":        frag.source,
                "text":          frag.text,
                "category":      frag.category,
                "fp_reason":     frag.fp_reason,
                "label":         frag.label,
                "pred_is_fp":    int(frag.pred_is_fp or False),
                "pred_fp_cat":   frag.pred_fp_cat,
                "semantic_score": frag.semantic_score,
                "correct":       int(frag.correct or False),
                "context":       frag.context[:500],
            })

    return json_path, csv_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="AISA Gold Set Evaluator — precision/recall for FP detection",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--input", "-i",
        required=True,
        help="Gold set file: .xlsx, .json or .csv",
    )
    p.add_argument(
        "--mode",
        choices=["rules", "full"],
        default="rules",
        help="'rules' = taxonomy rules only; 'full' = rules + semantic scoring",
    )
    p.add_argument(
        "--sample", "-n",
        type=int,
        default=0,
        help="If >0: stratified sampling to N fragments (default: all)",
    )
    p.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Seed for sampling reproducibility (default: 42)",
    )
    p.add_argument(
        "--ab-threshold",
        type=str,
        default="",
        help="A/B test: comma-separated list of thresholds (e.g.: 0.60,0.65,0.68)",
    )
    p.add_argument(
        "--out-dir",
        type=str,
        default="reports",
        help="Output directory for reports (default: reports/)",
    )
    p.add_argument(
        "--no-save",
        action="store_true",
        help="Do not save the report to disk (console display only)",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"[ERROR] File '{input_path}' does not exist.", file=sys.stderr)
        return 1

    # --- Load TAXONOMY ---
    print(f"\n  Loading taxonomy...")
    try:
        _m4 = importlib.import_module("04_taxonomy_builtin")
        taxonomy = _m4.TAXONOMY
        _ver = importlib.import_module("version")
        print(f"  Taxonomy v{_ver.TAXONOMY_VERSION} OK")
    except Exception as e:
        print(f"[ERROR] Cannot load taxonomy: {e}", file=sys.stderr)
        return 1

    # --- Load gold set ---
    print(f"  Loading gold set from {input_path}...")
    suffix = input_path.suffix.lower()
    try:
        if suffix == ".xlsx":
            fragments = load_xlsx(input_path, args.sample, args.seed)
        elif suffix == ".json":
            fragments = load_json(input_path, args.sample, args.seed)
        elif suffix == ".csv":
            fragments = load_csv(input_path, args.sample, args.seed)
        else:
            print(f"[ERROR] Unsupported format: {suffix}", file=sys.stderr)
            return 1
    except Exception as e:
        print(f"[ERROR] Loading error: {e}", file=sys.stderr)
        return 1

    if not fragments:
        print("[ERROR] Empty gold set — check the input file.", file=sys.stderr)
        return 1

    n_fp_gold = sum(1 for f in fragments if f.label == 1)
    n_tp_gold = sum(1 for f in fragments if f.label == 0)
    print(f"  Gold set: {len(fragments)} fragments ({n_fp_gold} FP + {n_tp_gold} TP)")

    # --- Rules evaluation ---
    print(f"  Rules evaluation ({len(fragments)} fragments)...")
    t0 = time.time()
    run_rules_eval(fragments, taxonomy)
    print(f"  Rules done in {time.time()-t0:.2f}s")

    # --- Semantic evaluation (mode=full) ---
    if args.mode == "full":
        _ver = importlib.import_module("version")
        run_semantic_eval(fragments, _ver.SEMANTIC_THRESHOLD)

    # --- A/B threshold ---
    ab_results = {}
    if args.ab_threshold:
        thresholds = [float(t.strip()) for t in args.ab_threshold.split(",") if t.strip()]
        print(f"  A/B test thresholds: {thresholds}")
        ab_results = run_ab_threshold(fragments, thresholds, taxonomy)

    # --- Metrics calculation ---
    report = compute_metrics(fragments)
    report.mode       = args.mode
    report.ab_results = ab_results

    # --- Display ---
    print_report(report)

    # --- Save ---
    if not args.no_save:
        out_dir = Path(args.out_dir)
        json_path, csv_path = save_report(report, fragments, out_dir)
        print(f"  Report saved:")
        print(f"    JSON: {json_path}")
        print(f"    CSV:  {csv_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
