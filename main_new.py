"""
===============================================================================
AISA - AI Semantic Analyzer
main_new.py - Interactive menu v2 (restructured)
===============================================================================

Entry point with text menu v2. Restructured with taxonomy Layer 0 + 9 chapters.

Structure:
    Layer 0  - Taxonomy setup (wizard la primul start, retinut in config)
    Cap 1    - Data extraction (ingest PDFs)
    Cap 2    - Deduplication
    Cap 3    - Export (bruta + deduplicata, cuvinte cheie colorate)
    Cap 4    - Indici compoziti (AI: Memory+AITI+Buzz+TPDI / non-AI: Relational)
    Cap 5    - Analize statistice [rezervat]
    Cap 6/7  - Vizualizari & grafice (Excel / HTML / PDF)
    Cap 8    - Coverage & statistics (unificat)
    Cap 9    - Configuratie & status (extins multi-taxonomie)

Taxonomy naming:
    output_folder : [NumeTaxonomie]_results/
    database_name : [NumeTaxonomie]_results.db
    export files  : [IndexName]_YYYYMMDD_HHMM.xlsx

Usage:
    python main_new.py
    python main_new.py --config aisa_config.json
    python main_new.py --db AI_Disclosure_results/AI_Disclosure_results.db

Author: TeRa0
License: MIT
===============================================================================
"""

from __future__ import annotations

import argparse
import glob
import importlib
import json
import logging
import os
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ── Force UTF-8 output on Windows ────────────────────────────────────────────
if sys.platform == "win32" and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from version import get_version_string, AISA_VERSION

# ── Suppress AISA console handler (logs go to file only) ─────────────────────
def _silence_aisa_console():
    aisa_log = logging.getLogger("AISA")
    aisa_log.handlers = [
        h for h in aisa_log.handlers
        if isinstance(h, logging.FileHandler)
    ]
_silence_aisa_console()

# ── Suppress noisy third-party library warnings ───────────────────────────────
logging.getLogger("pdfminer.pdfinterp").setLevel(logging.ERROR)
logging.getLogger("pdfminer.pdfpage").setLevel(logging.ERROR)

# ─────────────────────────────────────────────────────────────────────────────
# Lazy imports (numeric prefix → importlib)
# ─────────────────────────────────────────────────────────────────────────────

def _cli():
    m = importlib.import_module("11_cli")
    _silence_aisa_console()
    return m

def _models():
    m = importlib.import_module("01_models")
    _silence_aisa_console()
    return m

def _db_mod():
    m = importlib.import_module("02_db")
    _silence_aisa_console()
    return m

# ─────────────────────────────────────────────────────────────────────────────
# Display constants
# ─────────────────────────────────────────────────────────────────────────────

W    = 78
SEP  = "═" * W
SEP2 = "─" * W


# ─────────────────────────────────────────────────────────────────────────────
# Visual width helpers (emoji takes 2 columns in terminal)
# ─────────────────────────────────────────────────────────────────────────────

def vlen(s: str) -> int:
    """Real visual width of a string (emoji = 2 columns)."""
    w = 0
    for c in s:
        ew = unicodedata.east_asian_width(c)
        w += 2 if ew in ("W", "F") else 1
    return w


def vpad(s: str, width: int) -> str:
    """Pad string to exact visual width."""
    return s + " " * max(0, width - vlen(s))


# ─────────────────────────────────────────────────────────────────────────────
# Helpers UI
# ─────────────────────────────────────────────────────────────────────────────

def clr():
    os.system("cls" if os.name == "nt" else "clear")


def ask(prompt: str, default: str = "") -> str:
    hint = f" [{default}]" if default else ""
    val  = input(f"  {prompt}{hint}: ").strip()
    return val if val else default


def ask_int(prompt: str, default: int, lo: int = 1, hi: int = 99999) -> int:
    while True:
        raw = ask(prompt, str(default))
        try:
            v = int(raw)
            if lo <= v <= hi:
                return v
            print(f"  ✗ Enter a number between {lo} and {hi}.")
        except ValueError:
            print("  ✗ Invalid value.")


def ask_yn(prompt: str, default: bool = True) -> bool:
    hint = "Yes/no" if default else "yes/No"
    raw  = ask(f"{prompt} ({hint})", "").lower()
    if not raw:
        return default
    return raw[0] in ("y", "d")


def pause():
    input(f"\n  ⏎  Press Enter to continue...")


def section(title: str):
    print(f"\n{SEP}")
    print(f"  {title}")
    print(SEP)


def result_ok(title: str):
    print(f"\n{SEP2}")
    print(f"  ✓ {title}")
    print(SEP2)


def _fmt_sec(s: float) -> str:
    if s < 60:   return f"{s:.0f}s"
    if s < 3600: return f"{int(s//60)}m{int(s%60):02d}s"
    return f"{int(s//3600)}h{int((s%3600)//60):02d}m"


# ─────────────────────────────────────────────────────────────────────────────
# Global session state
# ─────────────────────────────────────────────────────────────────────────────

class Session:
    """Global state: config, paths, current session statistics."""

    def __init__(self, config_path: str = "", db_override: str = "",
                 input_folder: str = "", output_folder: str = ""):
        self.config_path      = config_path
        self.db_override      = db_override
        self.input_folder     = input_folder    # override → AnalyzerConfig.input_folder
        self.output_folder    = output_folder   # override → AnalyzerConfig.output_folder
        self._stats_cache     = None
        self._last_proc_stats = None   # ProcessingStats from last ingest
        self._last_doc_stats  = None   # List[Dict] stats per doc from last ingest

        # ── Legacy fields (backward compat) ──────────────────────────────────
        self.taxonomy_name    = "AI_Disclosure"   # taxonomia activă în sesiunea curentă
        self.taxonomy_excel   = None              # path la Excel taxonomy (None = folosește modul Python)
        self.extra_taxonomies: List[Dict] = []    # multi-taxonomy: [{"taxonomy_name": str, "taxonomy_excel": Optional[str]}]

        # ── Extended fields (v2) ──────────────────────────────────────────────
        # Lista tuturor taxonomiilor configurate (primary + extra).
        # Fiecare entry: {
        #   "taxonomy_name": str,
        #   "taxonomy_excel": Optional[str],   # None dacă e built-in Python
        #   "output_folder": str,              # ex: "AI_Disclosure_results/"
        #   "database_name": str,              # ex: "AI_Disclosure_results.db"
        # }
        self.taxonomies: List[Dict] = []

        # Indexul în self.taxonomies al taxonomiei primare (de obicei 0)
        self.primary_taxonomy_idx: int = 0

    # ── Computed properties ───────────────────────────────────────────────────

    @property
    def primary_taxonomy(self) -> Dict:
        """Returnează dictionarul taxonomiei primare."""
        if self.taxonomies:
            return self.taxonomies[self.primary_taxonomy_idx]
        return {
            "taxonomy_name":  self.taxonomy_name,
            "taxonomy_excel": self.taxonomy_excel,
            "output_folder":  self.output_folder or f"{self.taxonomy_name}_results",
            "database_name":  f"{self.taxonomy_name}_results.db",
        }

    @property
    def is_ai_taxonomy(self) -> bool:
        """True dacă taxonomia activă este AI_Disclosure (legacy)."""
        pt = self.primary_taxonomy
        return (
            pt.get("taxonomy_name") == "AI_Disclosure"
            and not pt.get("taxonomy_excel")
        )

    @property
    def is_multi_taxonomy(self) -> bool:
        """True dacă sunt configurate 2+ taxonomii."""
        return len(self.taxonomies) > 1

    # ── Sync legacy ───────────────────────────────────────────────────────────

    def sync_legacy_fields(self):
        """
        Sincronizează câmpurile legacy (taxonomy_name, taxonomy_excel,
        output_folder, extra_taxonomies) din self.taxonomies.
        Apelează după orice modificare a self.taxonomies.
        """
        if not self.taxonomies:
            return
        pt = self.taxonomies[0]
        self.taxonomy_name    = pt["taxonomy_name"]
        self.taxonomy_excel   = pt.get("taxonomy_excel")
        self.output_folder    = pt["output_folder"]
        self.extra_taxonomies = [
            {
                "taxonomy_name":  t["taxonomy_name"],
                "taxonomy_excel": t.get("taxonomy_excel"),
            }
            for t in self.taxonomies[1:]
        ]
        # Sincronizează și _extra_config pentru resolve()
        extra = self.get_extra_config()
        extra["database_name"] = pt["database_name"]
        self._extra_config = extra

    # ── Config helpers ────────────────────────────────────────────────────────

    def base_args(self, **extra) -> SimpleNamespace:
        ns = SimpleNamespace(
            config = self.config_path  or None,
            db     = self.db_override  or None,
            input  = self.input_folder  or None,   # → base["input_folder"]
            output = self.output_folder or None,   # → base["output_folder"]
        )
        for k, v in extra.items():
            setattr(ns, k, v)
        return ns

    def get_extra_config(self) -> dict:
        """Returns extra fields (fortune500_csv, database_name) set in wizard."""
        return getattr(self, "_extra_config", {})

    def resolve(self) -> Tuple:
        """Returns (AnalyzerConfig, db_path)."""
        cli  = _cli()
        args = self.base_args()
        # Inject extra fields from wizard (fortune500_csv, database_name)
        extra = self.get_extra_config()
        for k, v in extra.items():
            if not hasattr(args, k):
                setattr(args, k, v)
        # Injectează taxonomy în args dacă e setat în sesiune
        if not hasattr(args, "taxonomy_name") or not args.taxonomy_name:
            setattr(args, "taxonomy_name", self.taxonomy_name)
        if not hasattr(args, "taxonomy_excel") or not args.taxonomy_excel:
            setattr(args, "taxonomy_excel", self.taxonomy_excel)
        # If _extra_config has database_name, pass it as args.db
        if "database_name" in extra and not args.db:
            args.db = extra["database_name"]
        # If _extra_config has fortune500_csv, pass it
        if "fortune500_csv" in extra:
            args.csv = extra.get("fortune500_csv")
        return cli._resolve_config(args)

    def open_db(self, apply_migrations: bool = True):
        cli = _cli()
        _, db_path = self.resolve()
        return cli._open_db(db_path, apply_migrations=apply_migrations)

    # ── Stats live ───────────────────────────────────────────────────────────

    def get_stats(self, force: bool = False) -> Dict:
        if self._stats_cache and not force:
            return self._stats_cache

        s = dict(
            total_pdfs=0, processed=0, new=0,
            refs_raw=0, refs_dedup=0, total_occurrences=0,
            companies_with_index=0, docs_with_text_issues=0,
            reprocessed=0, by_text_status={},
        )

        try:
            config, db_path = self.resolve()

            pdf_folder = Path(config.input_folder)
            if pdf_folder.exists():
                s["total_pdfs"] = len(list(pdf_folder.glob("*.pdf")))

            if not Path(db_path).exists():
                s["new"] = s["total_pdfs"]
                return s

            with self.open_db(apply_migrations=False) as db:
                c = db.conn
                s["processed"]   = c.execute("SELECT COUNT(*) FROM processed_documents").fetchone()[0]
                s["reprocessed"] = c.execute("SELECT COUNT(*) FROM processed_documents WHERE process_count > 1").fetchone()[0]
                s["refs_raw"]    = c.execute("SELECT COUNT(*) FROM ai_references_raw").fetchone()[0]
                s["total_occurrences"] = c.execute(
                    "SELECT COALESCE(SUM(occurrence_count),0) FROM ai_references_raw"
                ).fetchone()[0]
                s["refs_dedup"]  = c.execute("SELECT COUNT(*) FROM ai_references_deduplicated").fetchone()[0]
                s["companies_with_index"] = c.execute(
                    "SELECT COUNT(DISTINCT company) FROM adoption_index"
                ).fetchone()[0]
                rows = c.execute(
                    "SELECT text_status, COUNT(*) FROM processed_documents GROUP BY text_status"
                ).fetchall()
                s["by_text_status"] = {r[0] or "unknown": r[1] for r in rows}
                s["docs_with_text_issues"] = c.execute(
                    "SELECT COUNT(*) FROM processed_documents "
                    "WHERE text_status IN ('corrupted_ocr_failed','ocr_needed','empty','error')"
                ).fetchone()[0]
        except Exception:
            pass

        s["new"] = max(0, s["total_pdfs"] - s["processed"])
        self._stats_cache = s
        return s

    def invalidate(self):
        self._stats_cache = None


# ─────────────────────────────────────────────────────────────────────────────
# HEADER helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_header_taxonomy_label(sess: Session) -> str:
    """Returnează labelul scurt pentru header."""
    if sess.is_multi_taxonomy:
        return f"Multi-taxonomy ({len(sess.taxonomies)} taxonomii active)"
    pt = sess.primary_taxonomy
    name = pt.get("taxonomy_excel") or pt.get("taxonomy_name", "Unknown")
    return f"{name} — Fortune 500"


# ─────────────────────────────────────────────────────────────────────────────
# HEADER (display)
# ─────────────────────────────────────────────────────────────────────────────

def display_header(sess: Session):
    s = sess.get_stats()
    BW = W - 2   # interior width between ╔ and ╗
    print()
    print("╔" + "═" * BW + "╗")
    t1 = f"  AISA v{AISA_VERSION} — {_get_header_taxonomy_label(sess)}"
    t2 = f"  Taxonomy v1.0 | Schema v2 | Model: all-MiniLM-L6-v2 | Threshold 0.60"
    print(f"║{vpad(t1, BW)}║")
    print(f"║{vpad(t2, BW)}║")
    print("╚" + "═" * BW + "╝")

    print(
        f"\n  📂 PDFs    : {s['total_pdfs']} total  |  "
        f"✅ {s['processed']} processed  |  "
        f"🆕 {s['new']} new"
    )
    if s["reprocessed"] > 0:
        print(f"     🔄 {s['reprocessed']} re-analyzed in previous sessions")
    if s["docs_with_text_issues"] > 0:
        print(f"     ⚠️  {s['docs_with_text_issues']} with text issues  →  option 1.5")
    if s["refs_raw"] > 0:
        print(
            f"  🔍 References: {s['refs_raw']:,} raw  |  "
            f"{s['total_occurrences']:,} occurrences  |  "
            f"{s['refs_dedup']:,} deduplicated  |  "
            f"{s['companies_with_index']} companies with index"
        )
    # Paths line — always show input + output
    try:
        config, _ = sess.resolve()
        inp = config.input_folder
        out = config.output_folder
    except Exception:
        inp = sess.input_folder or "Fortune500_PDFs"
        out = sess.output_folder or "Results_AISA"
    print(f"  📥 Input  : {inp}")
    print(f"  📤 Output : {out}")

    # Afișează taxonomia/taxonomiile active
    if sess.is_multi_taxonomy:
        print(f"  🏷️  Taxonomii: [{len(sess.taxonomies)} total]")
        for i, t in enumerate(sess.taxonomies):
            label  = f"{t['output_folder']} — {t['database_name']}"
            marker = "  [PRIMAR]" if i == sess.primary_taxonomy_idx else ""
            print(f"       [{i+1}] {label}{marker}")
    else:
        pt         = sess.primary_taxonomy
        tax_label  = pt.get("taxonomy_excel") or pt.get("taxonomy_name", "AI_Disclosure")
        tax_source = "📊 Excel" if pt.get("taxonomy_excel") else "🐍 Python"
        print(f"  🏷️  Taxonomy: {tax_source}  ->  {tax_label}")

    # CSV Fortune 500
    try:
        _csv = config.fortune500_csv
        if _csv:
            print(f"  📋 CSV F500: {_csv}")
    except Exception:
        pass
    if sess.config_path:
        print(f"  ⚙️  Config : {sess.config_path}")


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 0 — HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _setup_file_logging(output_folder: str) -> str:
    """
    Configure FileHandler for AISA logger in output_folder/aisa.log.
    Called after output_folder is known — logs go ONLY to file,
    not to console (console stays clean for menu).
    """
    log_path = str(Path(output_folder) / "aisa.log")
    aisa_log = logging.getLogger("AISA")
    aisa_log.handlers.clear()
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    aisa_log.addHandler(fh)
    aisa_log.setLevel(logging.DEBUG)
    return log_path


def _detect_builtin_taxonomies() -> List[Dict]:
    """
    Detectează automat toate taxonomiile built-in disponibile.
    Caută fișiere 04_taxonomy_*.py în folderul curent.
    Returnează listă de dictionare {"name": str, "module": str, "label": str}.
    """
    _HARDCODED = [
        {
            "name":   "AI_Disclosure",
            "module": "04_taxonomy_builtin",
            "label":  "AI Disclosure (legacy — Fortune 500 AI)",
        },
        {
            "name":   "Digitalization_Relational_v2",
            "module": "04_taxonomy_digitalization",
            "label":  "Digitalization Relational v2.2 (D/T/G tri-axial)",
        },
        {
            "name":   "Digitalization_Relational_v2_2_ZH",
            "module": "04_taxonomy_digitalization_zh",
            "label":  "Digitalization Relational v2.2 ZH (bilingual EN+ZH)",
        },
    ]
    known_modules = {t["module"] for t in _HARDCODED}

    base_dir = os.path.dirname(os.path.abspath(__file__))
    extra = []
    for path in glob.glob(os.path.join(base_dir, "04_taxonomy_*.py")):
        stem   = Path(path).stem            # e.g. "04_taxonomy_custom"
        module = stem                       # importlib name
        name   = stem[len("04_taxonomy_"):] # e.g. "custom"
        if module not in known_modules:
            extra.append({
                "name":   name,
                "module": module,
                "label":  f"{name} (detected)",
            })
            known_modules.add(module)

    return _HARDCODED + extra


def _build_taxonomy_entry(name: str, excel: Optional[str],
                           out_folder: str = "", db_name: str = "") -> Dict:
    """
    Construiește un entry standard pentru sess.taxonomies.
    Dacă out_folder sau db_name sunt vide, folosește naming standard.
    """
    if not out_folder:
        out_folder = f"{name}_results/"
    if not db_name:
        db_name = f"{name}_results.db"
    return {
        "taxonomy_name":  name,
        "taxonomy_excel": excel,
        "output_folder":  out_folder,
        "database_name":  db_name,
    }


def _display_taxonomy_config(taxonomies: List[Dict]):
    """Afișează frumos lista de taxonomii configurate."""
    BW = W - 2
    n  = len(taxonomies)
    print(f"\n  Taxonomii configurate ({n}):")
    print("  ╔" + "═" * (BW - 2) + "╗")
    for i, t in enumerate(taxonomies):
        marker = "  [PRIMAR]" if i == 0 else ""
        name   = t.get("taxonomy_name", "?")
        folder = t.get("output_folder", "")
        db     = t.get("database_name", "")
        excel  = t.get("taxonomy_excel")
        tag    = f" [{Path(excel).name}]" if excel else ""
        line1  = f"  [{i+1}] {name}{tag}{marker}"
        line2  = f"      Folder:   {folder}"
        line3  = f"      Database: {db}"
        sep    = "╠" if i < n - 1 else "╚"
        print(f"  ║{vpad(line1, BW - 2)}║")
        print(f"  ║{vpad(line2, BW - 2)}║")
        print(f"  ║{vpad(line3, BW - 2)}║")
        print(f"  {sep}" + "═" * (BW - 2) + ("╣" if i < n - 1 else "╝"))


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 0 — STARTUP CONFIG WIZARD
# ─────────────────────────────────────────────────────────────────────────────

def _startup_config(args) -> Session:
    """
    Layer 0: încarcă sau creează configurația la pornire.

    Case A: --config PATH explicit → încarcă JSON direct
    Case B: aisa_config.json detectat automat → afișează, întreabă
    Case C: nicio configurație → wizard complet 3 pași
    """
    CONFIG_NAME = "aisa_config.json"

    def _load_json(path: str) -> dict:
        return json.loads(Path(path).read_text(encoding="utf-8"))

    def _build_session_from_json(saved: dict, config_path: str,
                                  cli_input: str, cli_output: str) -> Session:
        """Reconstruiește Session din JSON salvat."""
        if "taxonomies" in saved and saved["taxonomies"]:
            taxonomies = saved["taxonomies"]
        else:
            # Backward compat cu format vechi (fără "taxonomies")
            taxonomies = [_build_taxonomy_entry(
                saved.get("taxonomy_name", "AI_Disclosure"),
                saved.get("taxonomy_excel"),
                saved.get("output_folder", ""),
                saved.get("database_name", ""),
            )]
        pt = taxonomies[0]
        sess = Session(
            config_path   = config_path,
            input_folder  = cli_input  or saved.get("input_folder", ""),
            output_folder = cli_output or pt["output_folder"],
        )
        sess.taxonomies          = taxonomies
        sess.primary_taxonomy_idx = saved.get("primary_taxonomy_idx", 0)
        sess._extra_config = {
            "fortune500_csv": saved.get("fortune500_csv"),
            "database_name":  pt["database_name"],
            "max_workers":    saved.get("max_workers", 4),
        }
        sess.sync_legacy_fields()
        return sess

    # ── Case A: config explicit ───────────────────────────────────────────────
    if args.config and Path(args.config).exists():
        clr()
        print()
        print(f"  ✓ Config încărcat: {args.config}")
        try:
            saved = _load_json(args.config)
            sess  = _build_session_from_json(
                saved, args.config, args.input or "", args.output or ""
            )
            _display_taxonomy_config(sess.taxonomies)
            _setup_file_logging(sess.taxonomies[0]["output_folder"])
            print(f"  📥 Input   : {sess.input_folder or 'Fortune500_PDFs'}")
            print(f"  🪵 Log     : {sess.taxonomies[0]['output_folder']}aisa.log")
        except Exception as e:
            print(f"  ⚠️  Eroare la citire config: {e}")
            sess = Session(
                config_path   = args.config,
                input_folder  = args.input  or "",
                output_folder = args.output or "",
            )
        time.sleep(1.0)
        return sess

    # ── Case B: config detectat automat ──────────────────────────────────────
    auto_cfg = None
    candidates = [CONFIG_NAME]
    # Caută și în sub-foldere *_results/
    for p in glob.glob("*_results/" + CONFIG_NAME):
        candidates.append(p)
    for candidate in candidates:
        if Path(candidate).exists():
            auto_cfg = candidate
            break

    if auto_cfg:
        clr()
        print()
        BW = W - 2
        print("╔" + "═" * BW + "╗")
        print(f"║{vpad('  AISA v' + AISA_VERSION + ' — Configurație detectată automat', BW)}║")
        print("╚" + "═" * BW + "╝")
        try:
            saved = _load_json(auto_cfg)
            print(f"\n  📄 Fișier  : {auto_cfg}")
            print(f"  📥 Input   : {saved.get('input_folder', 'N/A')}")
            if saved.get("fortune500_csv"):
                print(f"  📋 CSV F500: {saved['fortune500_csv']}")
            # Reconstruiește taxonomii pentru afișare
            if "taxonomies" in saved and saved["taxonomies"]:
                _display_taxonomy_config(saved["taxonomies"])
            else:
                _display_taxonomy_config([_build_taxonomy_entry(
                    saved.get("taxonomy_name", "AI_Disclosure"),
                    saved.get("taxonomy_excel"),
                    saved.get("output_folder", ""),
                    saved.get("database_name", ""),
                )])
            print()
            choice = ask("  Folosiți această configurație? [Da/Modificati/Nu]", "D").strip().lower()
            if choice in ("d", "da", "y", "yes", ""):
                sess = _build_session_from_json(
                    saved, auto_cfg, args.input or "", args.output or ""
                )
                _setup_file_logging(sess.taxonomies[0]["output_folder"])
                print(f"  🪵 Log     : {sess.taxonomies[0]['output_folder']}aisa.log")
                time.sleep(0.8)
                return sess
            elif choice in ("n", "nu", "no"):
                pass  # fall through to Case C
            # else "modificati" → fall through to Case C (cu defaults din JSON)
        except Exception as e:
            print(f"  ⚠️  Eroare la citire config: {e}")

    # ── Case C: wizard complet ────────────────────────────────────────────────
    clr()
    print()
    BW = W - 2
    print("╔" + "═" * BW + "╗")
    print(f"║{vpad('  AISA v' + AISA_VERSION + ' — Configurare inițială', BW)}║")
    print(f"║{vpad('  Pasul 1/3: Selecție taxonomii', BW)}║")
    print("╚" + "═" * BW + "╝")
    print()

    builtin = _detect_builtin_taxonomies()
    for i, t in enumerate(builtin, 1):
        print(f"  {i}.  {vpad(t['name'], 35)} {t['label']}")
    print(f"  E.  Încarcă din fișier Excel (.xlsx)")
    print()
    raw_sel = ask(
        "  Introduceți numerele dorite separate prin virgulă (ex: \"1\" sau \"1,2\")",
        "1"
    ).strip()

    selected_taxonomies: List[Dict] = []
    for token in raw_sel.split(","):
        token = token.strip().upper()
        if token == "E":
            xl_path = ask("    Calea la fișierul .xlsx", "").strip()
            if xl_path and Path(xl_path).exists() and xl_path.lower().endswith(".xlsx"):
                xl_name = Path(xl_path).stem
                selected_taxonomies.append(
                    _build_taxonomy_entry(xl_name, xl_path)
                )
                print(f"    ✓ Excel taxonomy: {xl_path}")
            else:
                print("    ⚠️  Fișier invalid sau inexistent — ignorat.")
        else:
            try:
                idx = int(token) - 1
                if 0 <= idx < len(builtin):
                    t = builtin[idx]
                    selected_taxonomies.append(
                        _build_taxonomy_entry(t["name"], None)
                    )
                else:
                    print(f"    ⚠️  Selecție invalidă: {token}")
            except ValueError:
                print(f"    ⚠️  Selecție invalidă: {token}")

    if not selected_taxonomies:
        print("  ⚠️  Nicio taxonomie selectată — folosim AI_Disclosure implicit.")
        selected_taxonomies = [_build_taxonomy_entry("AI_Disclosure", None)]

    # ── Pasul C2: Paths per taxonomie ────────────────────────────────────────
    clr()
    print()
    print("╔" + "═" * BW + "╗")
    print(f"║{vpad('  AISA v' + AISA_VERSION + ' — Configurare inițială', BW)}║")
    print(f"║{vpad('  Pasul 2/3: Configurare foldere și baze de date', BW)}║")
    print("╚" + "═" * BW + "╝")
    print()

    for i, t in enumerate(selected_taxonomies):
        name = t["taxonomy_name"]
        default_folder = f"{name}_results/"
        default_db     = f"{name}_results.db"
        print(f"  Taxonomie: {name}")
        folder = ask(f"    Folder rezultate  [{default_folder}]", default_folder).strip()
        db     = ask(f"    Nume bază de date [{default_db}]", default_db).strip()
        if not folder:
            folder = default_folder
        if not db:
            db = default_db
        t["output_folder"] = folder
        t["database_name"] = db
        Path(folder).mkdir(parents=True, exist_ok=True)
        print(f"    ✓ Folder creat: {folder}")
        print()

    # ── Pasul C3: Configurare globală ─────────────────────────────────────────
    clr()
    print()
    print("╔" + "═" * BW + "╗")
    print(f"║{vpad('  AISA v' + AISA_VERSION + ' — Configurare inițială', BW)}║")
    print(f"║{vpad('  Pasul 3/3: Configurare globală', BW)}║")
    print("╚" + "═" * BW + "╝")
    print()

    inp_folder = ask("  PDF folder (input_folder)", "Fortune500_PDFs").strip()
    csv_path   = ask("  CSV Fortune 500 (opțional)", "").strip() or None
    workers    = ask_int("  Număr workers", 4, lo=1, hi=64)

    Path(inp_folder).mkdir(parents=True, exist_ok=True)

    # ── Salvare config JSON ───────────────────────────────────────────────────
    config_save = Path(selected_taxonomies[0]["output_folder"]) / CONFIG_NAME
    data = {
        "taxonomies":          selected_taxonomies,
        "input_folder":        inp_folder,
        "fortune500_csv":      csv_path,
        "max_workers":         workers,
        "primary_taxonomy_idx": 0,
        # backward compat
        "taxonomy_name":  selected_taxonomies[0]["taxonomy_name"],
        "taxonomy_excel": selected_taxonomies[0].get("taxonomy_excel"),
        "output_folder":  selected_taxonomies[0]["output_folder"],
        "database_name":  selected_taxonomies[0]["database_name"],
    }
    try:
        config_save.parent.mkdir(parents=True, exist_ok=True)
        config_save.write_text(
            json.dumps(data, indent=2, ensure_ascii=False),
            encoding="utf-8"
        )
    except Exception as e:
        print(f"  ⚠️  Nu s-a putut salva config-ul: {e}")

    # ── Construire Session ────────────────────────────────────────────────────
    sess = Session(
        config_path   = str(config_save),
        input_folder  = args.input  or inp_folder,
        output_folder = args.output or selected_taxonomies[0]["output_folder"],
    )
    sess.taxonomies           = selected_taxonomies
    sess.primary_taxonomy_idx = 0
    sess._extra_config = {
        "fortune500_csv": csv_path,
        "database_name":  selected_taxonomies[0]["database_name"],
        "max_workers":    workers,
    }
    sess.sync_legacy_fields()

    log_path = _setup_file_logging(selected_taxonomies[0]["output_folder"])

    print()
    print(f"  ✓ Configurație salvată: {config_save}")
    print(f"  ✓ Taxonomie primară   : {selected_taxonomies[0]['taxonomy_name']}")
    if len(selected_taxonomies) > 1:
        print(f"    + {len(selected_taxonomies)-1} taxonomii secundare")
    print(f"  ✓ Input               : {inp_folder}")
    print(f"  🪵 Log                : {log_path}")

    time.sleep(1.5)
    return sess


# ─────────────────────────────────────────────────────────────────────────────
# CAP 1 — DATA EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def _doc_line(idx: int, total: int, company: str, year: int,
              pages: int, words: int, sentences: int, chars: int,
              refs_new: int, refs_upd: int, text_status: str, elapsed: float):
    pct  = 100 * idx / total if total else 0
    avg  = elapsed / idx if idx else 0
    eta  = _fmt_sec(avg * (total - idx))
    icon = ("✅" if text_status == "valid"
            else "🔵" if "ocr" in text_status
            else "⚠️")
    print(
        f"  [{idx:>4}/{total}] {pct:5.1f}%  "
        f"{icon} {company[:24]:<24} {year}  "
        f"📄{pages:>4}p  💬{sentences:>5}sent  "
        f"🔤{chars/1000:>6.1f}K  "
        f"🔍+{refs_new:<3}~{refs_upd:<3}  "
        f"avg={avg:.1f}s ETA={eta}"
    )


def _run_ingest(sess: Session, pdf_files: List[str], label: str):
    """
    Processes the list of PDFs displaying progress per document.
    Saves ProcessingStats in sess._last_proc_stats.
    """
    cli = _cli()
    m1  = _models()
    try:
        m10 = importlib.import_module("10_pipeline")
    except ModuleNotFoundError:
        print("  ❌ 10_pipeline.py was not found.")
        pause(); return

    config, db_path = sess.resolve()
    total           = len(pdf_files)

    section(f"{label}  —  {total} documents")
    print()

    PS             = m1.ProcessingStats
    proc_stats     = PS()
    proc_stats.start_time      = datetime.now()
    proc_stats.total_documents = total

    doc_stats_list: List[Dict] = []
    start_run = time.time()

    with cli._open_db(db_path) as db:
        processed_sources = {
            Path(r[0]).name for r in db.conn.execute(
                "SELECT source FROM processed_documents"
            ).fetchall()
        }

        for idx, pdf_path in enumerate(pdf_files, 1):
            filename      = Path(pdf_path).name
            is_reanalysis = filename in processed_sources
            doc_start     = time.time()

            result, text_status = None, "error"
            try:
                if hasattr(m10, "process_single"):
                    result, text_status = m10.process_single(
                        pdf_path=pdf_path, config=config, db=db,
                    )
            except Exception as exc:
                print(f"  ❌  [{idx}/{total}] {filename}: {exc}")
                proc_stats.failed_documents += 1
                doc_stats_list.append({
                    "idx": idx, "filename": filename,
                    "company": "—", "year": 0,
                    "pages": 0, "words": 0, "sentences": 0, "chars": 0,
                    "refs_new": 0, "refs_upd": 0,
                    "text_status": "error", "is_reanalysis": is_reanalysis,
                    "elapsed_s": time.time() - doc_start,
                })
                continue

            doc_elapsed   = time.time() - doc_start
            total_elapsed = time.time() - start_run

            if result:
                pages     = getattr(result, "total_pages",    0)
                chars     = getattr(result, "text_length",    0)
                words     = getattr(result, "word_count",     0)
                sentences = getattr(result, "sentence_count", 0)
                refs_new  = len(result.references)
                refs_upd  = 0

                proc_stats.successful_documents += 1
                proc_stats.total_pages          += pages
                proc_stats.total_text_chars     += chars
                proc_stats.total_words          += words
                proc_stats.total_sentences      += sentences
                proc_stats.total_references_raw += refs_new

                ts = text_status or "valid"
                if ts == "valid":        proc_stats.valid_pages     += pages
                elif "ocr" in ts:        proc_stats.ocr_pages       += pages; proc_stats.valid_pages += pages
                elif "corr" in ts:       proc_stats.corrupted_pages += pages

                doc_stats_list.append({
                    "idx": idx, "filename": filename,
                    "company": result.company, "year": result.year,
                    "pages": pages, "words": words,
                    "sentences": sentences, "chars": chars,
                    "refs_new": refs_new, "refs_upd": refs_upd,
                    "text_status": ts, "is_reanalysis": is_reanalysis,
                    "elapsed_s": doc_elapsed,
                })
                _doc_line(idx, total, result.company, result.year,
                          pages, words, sentences, chars,
                          refs_new, refs_upd, ts, total_elapsed)
            else:
                proc_stats.failed_documents += 1
                ts = text_status or "error"
                doc_stats_list.append({
                    "idx": idx, "filename": filename,
                    "company": "—", "year": 0,
                    "pages": 0, "words": 0, "sentences": 0, "chars": 0,
                    "refs_new": 0, "refs_upd": 0,
                    "text_status": ts, "is_reanalysis": is_reanalysis,
                    "elapsed_s": doc_elapsed,
                })
                print(f"  [{idx:>4}/{total}]  ❌  {filename[:55]}  [{ts}]")

    proc_stats.end_time = datetime.now()
    proc_stats.total_processing_time = time.time() - start_run
    proc_stats.calculate_derived_metrics()

    sess._last_proc_stats = proc_stats
    sess._last_doc_stats  = doc_stats_list
    sess.invalidate()

    print(f"\n{SEP2}")
    proc_stats.print_summary()
    print(SEP2)


def _pdfs_all(sess: Session, skip_processed: bool) -> List[str]:
    config, _ = sess.resolve()
    folder = Path(config.input_folder)
    if not folder.exists():
        print(f"  ✗ Folder inexistent: {folder}"); return []

    pdfs = sorted(folder.glob("*.pdf"))

    if skip_processed:
        try:
            with sess.open_db(apply_migrations=False) as db:
                done = {Path(r[0]).name for r in db.conn.execute(
                    "SELECT source FROM processed_documents").fetchall()}
            pdfs = [p for p in pdfs if p.name not in done]
        except Exception:
            pass

    return [str(p) for p in pdfs]


def _pdfs_by_pos(sess: Session, start: int, end: int,
                 skip_processed: bool) -> List[str]:
    result = []
    for p in _pdfs_all(sess, skip_processed):
        try:
            pos = int(Path(p).stem.split(".")[0])
            if start <= pos <= end:
                result.append(p)
        except (ValueError, IndexError):
            pass
    return result


def _pdfs_unreadable(sess: Session) -> List[str]:
    config, _ = sess.resolve()
    folder = Path(config.input_folder)
    try:
        with sess.open_db(apply_migrations=False) as db:
            sources = {r[0] for r in db.conn.execute(
                "SELECT source FROM processed_documents "
                "WHERE text_status IN "
                "('corrupted_ocr_failed','ocr_needed','empty','error')"
            ).fetchall()}
    except Exception:
        return []
    return [str(folder / s) for s in sources if (folder / s).exists()]


def _preview(pdfs: List[str], n: int = 5):
    for p in pdfs[:n]:
        print(f"    • {Path(p).name}")
    if len(pdfs) > n:
        print(f"    ... and {len(pdfs)-n} more")


def _run_multi_ingest(sess: Session, pdf_files: List[str], label: str):
    """
    Run multi-taxonomy pipeline on a given list of PDFs (single PDF-read pass).
    Parallel to _run_ingest() but dispatches to run_multi_taxonomy_pipeline().
    Results go to separate DBs per taxonomy. Saves stats in sess._last_proc_stats
    (primary taxonomy stats) and sess._last_doc_stats.
    """
    try:
        m10 = importlib.import_module("10_pipeline")
    except ModuleNotFoundError:
        print("  ❌ 10_pipeline.py not found."); return

    m5 = importlib.import_module("05_detect")
    m2 = _db_mod()
    m1 = _models()

    base_config, _ = sess.resolve()
    total = len(pdf_files)

    import copy
    run_config = copy.copy(base_config)
    run_config._explicit_pdf_list = [str(p) for p in pdf_files]

    # Build all_tax din sess.taxonomies (v2 — Layer 0)
    all_tax = [dict(t) for t in sess.taxonomies]
    if not all_tax:
        all_tax = [{"taxonomy_name": sess.taxonomy_name,
                    "taxonomy_excel": sess.taxonomy_excel}]

    section(f"{label}  [MULTI-TAXONOMY]  —  {total} documents × {len(all_tax)} taxonomies")
    print()
    for t in all_tax:
        src = t.get("taxonomy_excel") or t.get("taxonomy_name")
        print(f"  • {src}")
    print()

    TaxonomyTarget = m10.TaxonomyTarget
    targets    = []
    opened_dbs = []

    try:
        for t_idx, t in enumerate(all_tax):
            tx_name  = t["taxonomy_name"]
            tx_excel = t.get("taxonomy_excel")

            if tx_excel:
                loader   = importlib.import_module("taxonomy_excel_loader")
                provider = loader.ExcelTaxonomyProvider(tx_excel)
            else:
                provider = m5.load_taxonomy_by_name(tx_name)

            cfg_data = base_config.to_dict()
            cfg_data["taxonomy_name"]  = tx_name
            cfg_data["taxonomy_excel"] = tx_excel or None
            cfg_data["output_folder"]  = t.get("output_folder") or cfg_data.get("output_folder", "")
            cfg_data["database_name"]  = t.get("database_name") or cfg_data.get("database_name", "")
            tax_config = m1.AnalyzerConfig.from_dict(cfg_data)

            db_path = str(Path(tax_config.output_folder) / tax_config.database_name)
            db = m2.DatabaseManager(db_path)
            db.__enter__()
            db.apply_migrations()
            opened_dbs.append(db)

            targets.append(TaxonomyTarget(
                provider = provider,
                db       = db,
                config   = tax_config,
                name     = tx_name,
            ))

        start_t = time.time()

        def progress_cb(current, total_docs, company, year):
            pct     = 100 * current / total_docs if total_docs else 0
            elapsed = time.time() - start_t
            print(
                f"  [{current:>4}/{total_docs}] {pct:5.1f}%  "
                f"{company[:28]:<28} {year}  "
                f"elapsed={_fmt_sec(elapsed)}"
            )

        stats_map = m10.run_multi_taxonomy_pipeline(
            base_config       = run_config,
            targets           = targets,
            progress_callback = progress_cb,
        )

        print(f"\n{SEP}")
        print("  MULTI-TAXONOMY RESULTS")
        print(SEP)
        primary_name = sess.taxonomy_name
        for tx_name, ps in stats_map.items():
            print(f"\n  [{tx_name}]")
            ps.print_summary()
            if tx_name == primary_name:
                sess._last_proc_stats = ps

        sess.invalidate()

    except Exception as exc:
        print(f"\n  ❌ Error: {exc}")
        import traceback; traceback.print_exc()
    finally:
        for db in opened_dbs:
            try:
                db.__exit__(None, None, None)
            except Exception:
                pass


def _dispatch_ingest(sess: Session, pdf_files: List[str], label: str):
    """Alege automat single sau multi-taxonomy din configurația Layer 0."""
    if sess.is_multi_taxonomy:
        _run_multi_ingest(sess, pdf_files, label)
    else:
        _run_ingest(sess, pdf_files, label)


def opt_1_1(sess: Session):
    pdfs = _pdfs_all(sess, skip_processed=False)
    if not pdfs:
        print("  ✓ No documents to process in folder."); pause(); return

    s    = sess.get_stats()
    section("1.1  ALL DOCUMENTS  (new + re-analysis)")
    print(f"  Total:        {len(pdfs)}")
    print(f"  🆕 New:       {s['new']}")
    print(f"  🔄 Existing:  {len(pdfs) - s['new']}  (occurrence++ on re-analysis)")
    if not ask_yn("Continue?"): print("  Cancelled."); pause(); return
    _dispatch_ingest(sess, pdfs, "FULL PROCESSING")
    pause()


def opt_1_2(sess: Session):
    section("1.2  BATCH BY POSITIONS")
    start = ask_int("Start position", 1)
    end   = ask_int("End position",   50, lo=start)
    reana = ask_yn("Include re-analysis of existing documents?", True)
    pdfs  = _pdfs_by_pos(sess, start, end, skip_processed=not reana)
    if not pdfs:
        print(f"  ✓ No documents found in range {start}–{end}."); pause(); return
    print(f"\n  {len(pdfs)} documents found in range {start}–{end}:")
    _preview(pdfs)
    if not ask_yn("Continue?"): print("  Cancelled."); pause(); return
    _dispatch_ingest(sess, pdfs, f"BATCH POSITIONS {start}–{end}")
    pause()


def opt_1_3(sess: Session):
    pdfs = _pdfs_all(sess, skip_processed=True)
    if not pdfs:
        print("  ✓ No NEW documents to process.")
        print("    💡 Use 1.1 or 1.2 for re-analysis."); pause(); return
    section("1.3  NEW DOCUMENTS — ALL")
    print(f"  {len(pdfs)} NEW documents available:")
    _preview(pdfs)
    if not ask_yn("Continue?"): print("  Cancelled."); pause(); return
    _dispatch_ingest(sess, pdfs, "NEW DOCUMENTS — ALL")
    pause()


def opt_1_4(sess: Session):
    pdfs = _pdfs_all(sess, skip_processed=True)
    if not pdfs:
        print("  ✓ No NEW documents to process."); pause(); return
    section("1.4  NEW DOCUMENTS — BATCH OF N")
    print(f"  {len(pdfs)} NEW documents available")
    n     = ask_int("How many documents to process?", 10, lo=1, hi=len(pdfs))
    batch = pdfs[:n]
    print(f"\n  Will process {n} documents:")
    _preview(batch)
    if not ask_yn("Continue?"): print("  Cancelled."); pause(); return
    _dispatch_ingest(sess, batch, f"NEW DOCUMENTS — BATCH {n}")
    remaining = len(pdfs) - n
    if remaining > 0:
        print(f"\n  📌 {remaining} NEW documents still remaining to process.")
    else:
        print("  ✓ All NEW documents have been processed!")
    pause()


def opt_1_5(sess: Session):
    pdfs = _pdfs_unreadable(sess)
    section("1.5  DOCUMENTS WITH CORRUPT / UNREADABLE TEXT")
    if not pdfs:
        print("  ✓ No documents with text issues."); pause(); return
    print(f"  {len(pdfs)} documents require re-processing\n")
    try:
        with sess.open_db(apply_migrations=False) as db:
            rows = db.conn.execute(
                "SELECT text_status, COUNT(*) FROM processed_documents "
                "WHERE text_status IN ('corrupted_ocr_failed','ocr_needed','empty','error') "
                "GROUP BY text_status"
            ).fetchall()
        icons = {"corrupted_ocr_failed": "🔴", "ocr_needed": "🟡", "empty": "⚪", "error": "❌"}
        for status, count in rows:
            print(f"    {icons.get(status,'⚠️')}  {status}: {count}")
    except Exception:
        pass
    print()
    _preview(pdfs)
    if not ask_yn("Re-process documents?"): print("  Cancelled."); pause(); return
    _dispatch_ingest(sess, pdfs, "RE-PROCESSING CORRUPT DOCUMENTS")
    remaining = _pdfs_unreadable(sess)
    if remaining:
        print(f"  ⚠️  {len(remaining)} documents still problematic.")
    else:
        print("  ✓ All documents now have valid text!")
    pause()


def display_cap1_menu(sess: Session):
    s      = sess.get_stats()
    issues = s["docs_with_text_issues"]
    multi  = "[MULTI-TAX]" if sess.is_multi_taxonomy else ""

    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h(f"📥  1.  DATA EXTRACTION  (ingest PDFs) {multi}")
    print(f"     1.1   Toate documentele (new + re-analiză)          [{s['total_pdfs']} docs]")
    print(f"     1.2   Batch by positions  (ex: 1-50)")
    print(f"     1.3   Documente NOI — toate                         [{s['new']} disponibile]")
    print(f"     1.4   Documente NOI — batch de N")
    if issues > 0:
        print(f"     1.5   Documente cu text corupt / necitibil          [{issues} docs] ⚠️")
    else:
        print(f"     1.5   Documente cu text corupt / necitibil")
    if sess.is_multi_taxonomy:
        n = len(sess.taxonomies)
        print(f"           [Multi-taxonomy activ: {n} taxonomii — procesare paralelă automată]")


# ─────────────────────────────────────────────────────────────────────────────
# CAP 2 — DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def _build_taxonomy_db_list(sess: Session) -> List[Dict]:
    """
    Returnează lista tuturor DB-urilor configurate (primary + extra).
    Fiecare entry: {"name": str, "excel": str|None, "db_path": str, "exists": bool}
    """
    # Folosește sess.taxonomies dacă e populat (v2 — Layer 0)
    if sess.taxonomies:
        entries = []
        for t in sess.taxonomies:
            db_path = str(Path(t["output_folder"]) / t["database_name"])
            entries.append({
                "name":    t["taxonomy_name"],
                "excel":   t.get("taxonomy_excel"),
                "db_path": db_path,
                "exists":  Path(db_path).exists(),
            })
        return entries

    # Fallback la logica din main.py (backward compat)
    m1 = _models()
    all_tax = [
        {"taxonomy_name": sess.taxonomy_name, "taxonomy_excel": sess.taxonomy_excel}
    ] + sess.extra_taxonomies

    entries = []
    base, _ = sess.resolve()
    for i, t in enumerate(all_tax):
        tx_name  = t["taxonomy_name"]
        tx_excel = t.get("taxonomy_excel")
        cfg_data = base.to_dict()
        cfg_data["taxonomy_name"]  = tx_name
        cfg_data["taxonomy_excel"] = tx_excel or None
        if i > 0:
            cfg_data["output_folder"] = ""
            cfg_data["database_name"] = ""
        tax_config = m1.AnalyzerConfig.from_dict(cfg_data)
        db_path = str(Path(tax_config.output_folder) / tax_config.database_name)
        entries.append({
            "name":    tx_name,
            "excel":   tx_excel,
            "db_path": db_path,
            "exists":  Path(db_path).exists(),
        })
    return entries


def _dedupe(sess: Session, **kw):
    cli  = _cli()
    args = sess.base_args(**kw)
    cli.cmd_dedupe(args)
    sess.invalidate()


def _dedup_extra_taxonomies(sess: Session):
    """Rulează deduplicarea pe toate taxonomiile secundare (multi-taxonomy)."""
    if not sess.is_multi_taxonomy:
        return
    entries      = _build_taxonomy_db_list(sess)
    extra_entries = [e for e in entries[1:] if e["exists"]]
    if not extra_entries:
        return
    print(f"\n  Multi-taxonomy: deduplicare și pe {len(extra_entries)} taxonomii secundare...")
    for e in extra_entries:
        print(f"  [{e['name']}]")
        try:
            m10 = importlib.import_module("10_pipeline")
            m2  = _db_mod()
            with m2.DatabaseManager(e["db_path"]) as db:
                result = m10.run_dedup_for_db(db, year_filter=None)
            if result:
                for year, clusters in sorted(result.items()):
                    print(f"    {year}: {clusters} clustere deduplicate")
            else:
                print("    Nicio referință găsită.")
        except Exception as exc:
            print(f"    Eroare: {exc}")


def opt_2_1(sess: Session):
    section("2.1  DEDUPLICATION BATCH BY POSITIONS")
    start = ask_int("Start position", 1)
    end   = ask_int("End position",   50, lo=start)
    _dedupe(sess, position_start=start, position_end=end)
    _dedup_extra_taxonomies(sess)
    pause()


def opt_2_2(sess: Session):
    section("2.2  DEDUPLICATE ALL COMPANIES")
    if not ask_yn("Run deduplication on entire database?"): pause(); return
    _dedupe(sess)
    _dedup_extra_taxonomies(sess)
    pause()


def opt_2_3(sess: Session):
    section("2.3  DEDUPLICATE NEW COMPANIES")
    _dedupe(sess, only_new=True)
    _dedup_extra_taxonomies(sess)
    pause()


def display_cap2_menu(sess: Session):
    multi_note = " [toate taxonomiile]" if sess.is_multi_taxonomy else ""

    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h(f"📊  2.  DEDUPLICATION{multi_note}")
    print("     2.1   Deduplicare batch by positions")
    print("     2.2   Deduplicare bulk — toate companiile")
    print("     2.3   Deduplicare only new  (nededuplicate încă)")


# ─────────────────────────────────────────────────────────────────────────────
# CAP 3 — EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _export(sess: Session, fmt: str, year=None):
    cli       = _cli()
    config, _ = sess.resolve()
    args = sess.base_args(
        output=config.output_folder, format=fmt, year=year, base_name="aisa_results"
    )
    cli.cmd_export(args)


def opt_3_1(sess: Session):
    section("3.1  EXPORT BAZĂ BRUTĂ  (referințe raw)")
    config, _ = sess.resolve()
    year_s = ask("An specific (Enter = toți)", "")
    year   = int(year_s) if year_s.isdigit() else None
    fmt    = ask("Format [excel/json/csv/all]", "excel").strip().lower()
    if fmt not in ("excel", "json", "csv", "all"):
        fmt = "excel"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"raw_export_{timestamp}"
    args = sess.base_args(
        output=config.output_folder, format=fmt, year=year,
        base_name=base_name, export_type="raw",
    )
    try:
        cli = _cli()
        cli.cmd_export(args)
        result_ok(f"Export brut salvat în {config.output_folder}/")
    except Exception as e:
        print(f"  Eroare: {e}")
    pause()


def opt_3_2(sess: Session):
    section("3.2  EXPORT BAZĂ DEDUPLICATĂ")
    config, _ = sess.resolve()
    year_s = ask("An specific (Enter = toți)", "")
    year   = int(year_s) if year_s.isdigit() else None
    fmt    = ask("Format [excel/json/csv/all]", "excel").strip().lower()
    if fmt not in ("excel", "json", "csv", "all"):
        fmt = "excel"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"dedup_export_{timestamp}"
    args = sess.base_args(
        output=config.output_folder, format=fmt, year=year,
        base_name=base_name, export_type="dedup",
    )
    try:
        cli = _cli()
        cli.cmd_export(args)
        result_ok(f"Export deduplicat salvat în {config.output_folder}/")
    except Exception as e:
        print(f"  Eroare: {e}")
    pause()


def opt_3_3(sess: Session):
    section("3.3  EXPORT COMPLET  (toate formatele)")
    config, _ = sess.resolve()
    year_s = ask("An specific (Enter = toți)", "")
    year   = int(year_s) if year_s.isdigit() else None
    if not ask_yn("Generează toate rapoartele (Excel + JSON + CSV)?"):
        print("  Anulat."); pause(); return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"full_export_{timestamp}"
    args = sess.base_args(
        output=config.output_folder, format="all", year=year, base_name=base_name,
    )
    try:
        cli = _cli()
        cli.cmd_export(args)
        result_ok(f"Export complet salvat în {config.output_folder}/")
    except Exception as e:
        print(f"  Eroare: {e}")
    pause()


def opt_3_4(sess: Session):
    section("3.4  EXPORT PER TAXONOMIE")

    entries  = _build_taxonomy_db_list(sess)
    existing = [e for e in entries if e["exists"]]
    if not existing:
        print("  Nicio bază de date găsită. Rulează mai întâi pipeline (cap 1).")
        pause(); return

    if len(existing) == 1:
        print("  O singură taxonomie configurată — export direct pentru taxonomia activă.\n")
        selected = existing
        choice = "A"
    else:
        print("  Taxonomii disponibile:\n")
        for i, e in enumerate(existing, 1):
            label = e["excel"] or e["name"]
            print(f"     {i}.  {label:<45}  {e['db_path']}")
        print(f"     A.  TOATE ({len(existing)} taxonomii)")
        print()

        choice = ask(f"Select [1-{len(existing)} / A]", "A").strip().upper()

    year_s = ask("An specific (Enter = toți)", "")
    year   = int(year_s) if year_s.isdigit() else None
    fmt    = ask("Format [excel/json/all]", "excel").strip().lower()

    if choice == "A":
        selected = existing
    else:
        try:
            idx = int(choice) - 1
            if not (0 <= idx < len(existing)):
                raise ValueError
            selected = [existing[idx]]
        except (ValueError, IndexError):
            print("  Selecție invalidă."); pause(); return

    m7  = importlib.import_module("07_export")
    m5  = importlib.import_module("05_detect")
    m2  = _db_mod()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    for e in selected:
        print(f"\n  [{e['name']}]  {e['db_path']}")
        try:
            # Provider opțional (pentru foi Keywords + Taxonomy_Meta)
            provider = None
            try:
                if e["excel"]:
                    loader   = importlib.import_module("taxonomy_excel_loader")
                    provider = loader.ExcelTaxonomyProvider(e["excel"])
                else:
                    provider = m5.load_taxonomy_by_name(e["name"])
            except Exception:
                pass

            out_path = str(Path(e["db_path"]).parent / f"export_{e['name']}_{timestamp}.xlsx")
            with m2.DatabaseManager(e["db_path"]) as db:
                if fmt in ("excel", "all"):
                    m7.export_excel(db, out_path, year=year, taxonomy=provider)
                    print(f"    Excel: {out_path}")
                if fmt in ("json", "all"):
                    json_path = out_path.replace(".xlsx", ".json")
                    m7.export_json(db, json_path, year=year)
                    print(f"    JSON:  {json_path}")
        except Exception as exc:
            print(f"    Eroare: {exc}")
            import traceback; traceback.print_exc()

    pause()


def display_cap3_menu(sess: Session):
    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h("📋  3.  EXPORT")
    print("     3.1   Export bază BRUTĂ          (Excel + JSON, cuvinte cheie colorate)")
    print("     3.2   Export bază DEDUPLICATĂ    (Excel + JSON, cuvinte cheie colorate)")
    print("     3.3   Export COMPLET             (toate formatele: Excel + JSON + CSV)")
    if sess.is_multi_taxonomy:
        n = len(sess.taxonomies)
        print(f"     3.4   Export per taxonomie       [{n} taxonomii configurate]")
    else:
        print("     3.4   Export per taxonomie       [o singură taxonomie activă]")
    print()
    print("     Notă: cuvintele cheie detectate sunt marcate BOLD+ROȘU în Excel")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN MENU (stub — va fi înlocuit complet în pasul 8)
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# CAP 4 — INDICI COMPOZITI (implementare reală PAS 5B)
# ─────────────────────────────────────────────────────────────────────────────

# ── Export helpers ─────────────────────────────────────────────────────────────

def _export_index_to_excel(sess: Session, label: str, output_path: str):
    """Exportă indexul (raw + dedup + buzz) în Excel via 07_export.export_excel."""
    try:
        m7 = importlib.import_module("07_export")
        m2 = _db_mod()
        _, db_path = sess.resolve()
        with m2.DatabaseManager(db_path) as db:
            m7.export_excel(db, output_path)
        result_ok(f"{label} exportat: {output_path}")
    except Exception as e:
        print(f"  ⚠️  Export eșuat: {e}")


def _export_aiti_to_excel(results: dict, output_path: str):
    """Exportă rezultatele AITI (dict company→year→metrics) într-un Excel."""
    try:
        import pandas as pd
        rows = []
        for company, years in results.items():
            for year, metrics in years.items():
                row = {"company": company, "year": year}
                row.update(metrics if isinstance(metrics, dict) else {"aiti": metrics})
                rows.append(row)
        if not rows:
            print("  ⚠️  Nu există date AITI de exportat.")
            return
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(rows).to_excel(output_path, index=False)
        result_ok(f"AITI exportat: {output_path}")
    except ImportError:
        print("  ⚠️  pandas/openpyxl lipsesc — export AITI omis.")
    except Exception as e:
        print(f"  ⚠️  Export AITI eșuat: {e}")


def _export_buzz_index_to_excel(sess: Session, output_path: str):
    """Exportă AI Buzz Index în Excel via 07_export."""
    _export_index_to_excel(sess, "AI Buzz Index", output_path)


def _export_relational_memory_to_excel(sess: Session, output_path: str):
    """Exportă Relational Adoption Memory în Excel via 07_export."""
    _export_index_to_excel(sess, "Relational Memory", output_path)


# ── AI: opt_4_1_ai — Adoption Memory ──────────────────────────────────────────

def opt_4_1_ai(sess: Session):
    section("4.1  ADOPTION MEMORY  (recalculare completă)")
    print("  Această operație recalculează Adoption Memory din referințele deduplicate.")
    print("  Sursa: ai_references_deduplicated → adoption_portfolio_a/b")
    print()
    mode = ask("  Mod [1=incremental / 2=reprocess tot]", "1").strip()
    reprocess_all = (mode == "2")
    if reprocess_all:
        if not ask_yn("  ⚠️  Recalculare completă (resetează flag-urile). Continuați?", False):
            print("  Anulat."); pause(); return
    else:
        if not ask_yn("  Procesează referințe noi (incremental)?", True):
            print("  Anulat."); pause(); return

    try:
        cli  = _cli()
        args = sess.base_args(reprocess_all=reprocess_all)
        cli.cmd_memory(args)
        sess.invalidate()

        # Export Excel cu timestamp
        config, _ = sess.resolve()
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
        out_path   = str(Path(config.output_folder) / f"AdoptionMemory_{timestamp}.xlsx")
        _export_index_to_excel(sess, "Adoption Memory", out_path)
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
        import traceback; traceback.print_exc()
    pause()


# ── AI: opt_4_2_ai — AITI ──────────────────────────────────────────────────────

def opt_4_2_ai(sess: Session):
    section("4.2  AITI — AI Adoption Trajectory Index")
    print("  Calculează AITI pentru toate companiile din adoption_events.")
    print()
    if not ask_yn("  Recalculează AITI (suprascrie valorile existente)?", True):
        print("  Anulat."); pause(); return

    try:
        m12 = importlib.import_module("12_aiti")
        m2  = _db_mod()
        _, db_path = sess.resolve()

        with m2.DatabaseManager(db_path) as db:
            # Verifică dacă există date
            count = db.conn.execute(
                "SELECT COUNT(*) FROM adoption_events"
            ).fetchone()[0]
            if count == 0:
                print("  ⚠️  Nu există date în adoption_events. Rulează 4.1 (Memory) mai întâi.")
                pause(); return

            print(f"  {count} events găsite. Calculez AITI...")
            summary = m12.calculate_aiti_all(db)
            print(f"  ✓ Calculat: {summary.get('calculated', '?')} înregistrări")
            if summary.get('errors'):
                print(f"  ⚠️  Erori: {summary['errors']}")

        # Export Excel
        config, _ = sess.resolve()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        out_path  = str(Path(config.output_folder) / f"AITI_{timestamp}.xlsx")

        # Citim scorurile pentru export
        with m2.DatabaseManager(db_path) as db:
            rows = db.conn.execute(
                "SELECT company, year, aiti_score, new_pairs, continued_pairs, "
                "total_pairs, pair_points, maturity_bonus "
                "FROM aiti_scores ORDER BY company, year"
            ).fetchall()
        if rows:
            try:
                import pandas as pd
                df = pd.DataFrame([dict(r) for r in rows])
                Path(out_path).parent.mkdir(parents=True, exist_ok=True)
                df.to_excel(out_path, index=False)
                result_ok(f"AITI exportat: {out_path}")
            except Exception as e:
                print(f"  ⚠️  Export Excel eșuat: {e}")
        sess.invalidate()
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
        import traceback; traceback.print_exc()
    pause()


# ── AI: opt_4_3_ai — Buzz Index ────────────────────────────────────────────────

def opt_4_3_ai(sess: Session):
    section("4.3  AI BUZZ INDEX")
    print("  Recalculează AI Buzz Index (7 dimensiuni) și exportă Excel.")
    print()
    print("  Mod de calcul:")
    print("    1. Toate companiile")
    print("    2. Batch by positions")
    print("    3. Numai companiile noi (fără index)")
    mode = ask("  Mod [1/2/3]", "1").strip()

    if not ask_yn("  Continuați cu recalcularea Buzz Index?", True):
        print("  Anulat."); pause(); return

    try:
        cli = _cli()
        if mode == "2":
            start = ask_int("  Start position", 1)
            end   = ask_int("  End position",   50, lo=start)
            args  = sess.base_args(year=None, position_start=start, position_end=end)
        elif mode == "3":
            args  = sess.base_args(year=None, only_new=True)
        else:
            year_s = ask("  An specific (Enter = toți)", "")
            args   = sess.base_args(year=int(year_s) if year_s.isdigit() else None)

        cli.cmd_index(args)
        sess.invalidate()

        # Export Excel cu timestamp
        config, _ = sess.resolve()
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
        out_path   = str(Path(config.output_folder) / f"BuzzIndex_{timestamp}.xlsx")
        _export_buzz_index_to_excel(sess, out_path)
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
        import traceback; traceback.print_exc()
    pause()


# ── AI: opt_4_4_ai — TPDI ──────────────────────────────────────────────────────

def opt_4_4_ai(sess: Session):
    section("4.4  TPDI — Technology-Product Diffusion Index")
    print("  Calculează difuzia produs/tehnologie în corpus Fortune 500.")
    print()
    if not ask_yn("  Calculați TPDI?", True):
        print("  Anulat."); pause(); return

    try:
        config, _ = sess.resolve()
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
        out_path   = str(Path(config.output_folder) / f"TPDI_{timestamp}.xlsx")
        top_n_s    = ask("  Top N produse (Enter = toate)", "")
        cli        = _cli()
        args       = sess.base_args(
            tpdi_output = out_path,
            min_adopters = 1,
            no_products  = False,
            top_n        = int(top_n_s) if top_n_s.isdigit() else None,
        )
        cli.cmd_tpdi(args)
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
        import traceback; traceback.print_exc()
    pause()


# ── non-AI: opt_4_1_relational — Relational Adoption Memory ───────────────────

def opt_4_1_relational(sess: Session):
    section("4.1  RELATIONAL ADOPTION MEMORY  (recalculare completă)")
    print("  Recalculează Adoption Memory pentru taxonomia relațională.")
    print()
    mode = ask("  Mod [1=incremental / 2=reprocess tot]", "1").strip()
    reprocess_all = (mode == "2")
    if reprocess_all:
        if not ask_yn("  ⚠️  Recalculare completă. Continuați?", False):
            print("  Anulat."); pause(); return
    else:
        if not ask_yn("  Procesează referințe noi (incremental)?", True):
            print("  Anulat."); pause(); return

    try:
        # Verifică dacă tabelele memory există
        m2 = _db_mod()
        _, db_path = sess.resolve()
        with m2.DatabaseManager(db_path) as db:
            tables = {r[0] for r in db.conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
        if "adoption_portfolio_a" not in tables:
            print("  ⚠️  Tabelele Adoption Memory nu există în acest DB.")
            print("  Asigurați-vă că migrațiile sunt aplicate și că taxonomia suportă Memory.")
            pause(); return

        cli  = _cli()
        args = sess.base_args(reprocess_all=reprocess_all)
        cli.cmd_memory(args)
        sess.invalidate()

        # Export Excel cu timestamp
        config, _ = sess.resolve()
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
        out_path   = str(Path(config.output_folder) / f"RelationalMemory_{timestamp}.xlsx")
        _export_relational_memory_to_excel(sess, out_path)
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
        import traceback; traceback.print_exc()
    pause()


# ── non-AI: opt_4_2_relational — DTI rezervat ─────────────────────────────────

def opt_4_2_relational(sess: Session):
    section("4.2  DTI — Digitalization Trajectory Index  [REZERVAT]")
    print("  Funcționalitate rezervată pentru versiunea viitoare.")
    print("  DTI va măsura traiectoria digitalizării pe dimensiunile D/T/G.")
    pause()


# ── Display ───────────────────────────────────────────────────────────────────

def display_cap4_menu(sess: Session):
    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    if sess.is_ai_taxonomy:
        h("🧠  4.  INDICI COMPOZITI  [AI_Disclosure]")
        print("     4.1   Adoption Memory         — recalc + export")
        print("     4.2   AITI                    — recalc + export")
        print("     4.3   AI Buzz Index           — recalc + export")
        print("     4.4   TPDI                    — recalc + export")
    else:
        tax_name = sess.primary_taxonomy.get("taxonomy_name", "Custom")
        h(f"🧠  4.  INDICI COMPOZITI  [{tax_name}]")
        print("     4.1   Relational Adoption Memory  — recalc + export")
        print("     4.2   DTI — Digitalization Trajectory Index  [rezervat]")
        print()
        print("     Notă: Indici legacy AI (AITI, Buzz, TPDI) nu se aplică")
        print("     pentru taxonomii non-AI.")


# ── Routing ───────────────────────────────────────────────────────────────────

def opt_4_1(sess: Session):
    if sess.is_ai_taxonomy:
        opt_4_1_ai(sess)
    else:
        opt_4_1_relational(sess)


def opt_4_2(sess: Session):
    if sess.is_ai_taxonomy:
        opt_4_2_ai(sess)
    else:
        opt_4_2_relational(sess)


def opt_4_3(sess: Session):
    if sess.is_ai_taxonomy:
        opt_4_3_ai(sess)
    else:
        print("  Opțiunea 4.3 (Buzz Index) nu se aplică pentru taxonomii non-AI.")
        pause()


def opt_4_4(sess: Session):
    if sess.is_ai_taxonomy:
        opt_4_4_ai(sess)
    else:
        print("  Opțiunea 4.4 (TPDI) nu se aplică pentru taxonomii non-AI.")
        pause()


# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# CAP 6/7 — VIZUALIZĂRI & GRAFICE
# ─────────────────────────────────────────────────────────────────────────────

def _viz(sess: Session, mode: str, year=None):
    """Run a 16_viz export (excel / html / pdf)."""
    try:
        m16 = importlib.import_module("16_viz")
    except ModuleNotFoundError:
        print("  ❌  16_viz.py not found."); pause(); return

    config, db_path = sess.resolve()
    out = config.output_folder

    fn_map = {
        "excel": (m16.export_excel_charts,   f"{out}/aisa_charts.xlsx"),
        "html":  (m16.export_html_dashboard, f"{out}/aisa_dashboard.html"),
        "pdf":   (m16.export_pdf_report,     f"{out}/aisa_report.pdf"),
    }
    fn, default_path = fn_map[mode]
    try:
        path = fn(db_path, default_path, year)
        result_ok(f"Saved: {path}")
    except ImportError as e:
        print(f"  ❌  Missing dependency: {e}")
    except Exception as e:
        print(f"  ❌  Error: {e}")
        import traceback; traceback.print_exc()
    pause()


def opt_6_1(sess: Session):
    section("6.1  CHARTS EXPORT  (Excel embedded charts)")
    year_s = ask("Specific year (Enter = all)", "")
    year   = int(year_s) if year_s.isdigit() else None
    _viz(sess, "excel", year)


def opt_6_2(sess: Session):
    section("6.2  INTERACTIVE DASHBOARD  (HTML / Plotly)")
    year_s = ask("Reference year for static charts (Enter = all)", "")
    year   = int(year_s) if year_s.isdigit() else None
    _viz(sess, "html", year)


def opt_6_3(sess: Session):
    section("6.3  RAPORT PDF  (publication-quality)")
    year_s = ask("Specific year (Enter = latest)", "")
    year   = int(year_s) if year_s.isdigit() else None
    _viz(sess, "pdf", year)


# ─────────────────────────────────────────────────────────────────────────────
# CAP 8 — COVERAGE & STATISTICS
# ─────────────────────────────────────────────────────────────────────────────

def opt_8_1(sess: Session):
    """8.1 — Statistici run curent (din sess._last_proc_stats)."""
    section("8.1  STATISTICI RUN CURENT")
    ps = sess._last_proc_stats
    if ps is None:
        print("  ⚠️  Nicio statistică disponibilă. Rulează ingest mai întâi (1.x).")
        pause(); return
    ps.print_summary()
    pause()


def opt_8_2(sess: Session):
    """8.2 — Statistici per document (tabel detaliat ultima sesiune)."""
    section("8.2  STATISTICI PER DOCUMENT  (ultima sesiune)")
    docs = sess._last_doc_stats
    if not docs:
        print("  ⚠️  Nicio statistică per document. Rulează ingest mai întâi (1.x).")
        pause(); return

    has_para = any("paragraph_count" in d for d in docs)
    para_col = f"  {'Para':>5}" if has_para else ""
    hdr = (f"  {'No':>4}  {'Company':<24} {'Year':>4}  "
           f"{'Pgs':>4}  {'Wds':>7}  {'Snt':>6}{para_col}  {'Chr':>9}  "
           f"{'Ref+':>4}  {'Status':<23}  {'Time':>6}")
    print(f"\n{hdr}")
    print(f"  {SEP2}")

    for d in docs:
        icon = ("✅" if d["text_status"] == "valid"
                else "🔵" if "ocr" in d["text_status"] else "⚠️")
        para_val = (f"  {d.get('paragraph_count', 0):>5}" if has_para else "")
        print(
            f"  {d['idx']:>4}  {d['company'][:24]:<24} {d['year'] or 0:>4}  "
            f"{d['pages']:>4}  {d['words']:>7,}  {d['sentences']:>6,}{para_val}  "
            f"{d['chars']:>9,}  "
            f"{d['refs_new']:>4}  "
            f"{icon} {d['text_status'][:21]:<21}  "
            f"{d['elapsed_s']:>5.1f}s"
        )

    print(f"\n  {SEP2}")
    ok    = sum(1 for d in docs if d["text_status"] == "valid")
    ocr   = sum(1 for d in docs if "ocr" in d["text_status"] and "failed" not in d["text_status"])
    err   = sum(1 for d in docs if d["text_status"] in ("error", "corrupted_ocr_failed", "empty"))
    tot_p = sum(d["pages"]     for d in docs)
    tot_w = sum(d["words"]     for d in docs)
    tot_s = sum(d["sentences"] for d in docs)
    tot_c = sum(d["chars"]     for d in docs)
    tot_r = sum(d["refs_new"]  for d in docs)
    print(f"  Documents: {len(docs)}  ✅ {ok}  🔵 {ocr}  ❌ {err}")
    line = (f"  Totals:  {tot_p:,} pgs  |  {tot_w:,} wds  |  "
            f"{tot_s:,} sent  |  {tot_c:,} chr  |  {tot_r:,} refs")
    if has_para:
        tot_para = sum(d.get("paragraph_count", 0) for d in docs)
        line += f"  |  {tot_para:,} para"
    print(line)
    pause()


def opt_8_3(sess: Session):
    """8.3 — Status procesare documente (breakdown by text_status)."""
    section("8.3  STATUS PROCESARE DOCUMENTE")

    # Detectează dacă coloana last_processed_at există în DB
    has_last_proc = False
    try:
        with sess.open_db(apply_migrations=False) as db:
            cols = {r[1] for r in db.conn.execute(
                "PRAGMA table_info(processed_documents)"
            ).fetchall()}
            has_last_proc = "last_processed_at" in cols
    except Exception:
        pass

    if not has_last_proc:
        print("  Notă: coloana last_processed_at nu există în DB curent.")
        print("  Rulează o migrare sau re-procesează documentele pentru a popula această coloană.")
        print()

    try:
        with sess.open_db(apply_migrations=False) as db:
            if has_last_proc:
                rows = db.conn.execute(
                    "SELECT source, company, year, doc_type, refs_found, "
                    "text_status, process_count, last_processed_at "
                    "FROM processed_documents ORDER BY last_processed_at DESC"
                ).fetchall()
            else:
                rows = db.conn.execute(
                    "SELECT source, company, year, doc_type, refs_found, "
                    "text_status, process_count "
                    "FROM processed_documents ORDER BY company, year"
                ).fetchall()
    except Exception as e:
        print(f"  ❌ DB Error: {e}"); pause(); return

    if not rows:
        print("  ⚠️  Niciun document procesat în DB."); pause(); return

    icons = {
        "valid": "✅", "corrupted_ocr_success": "🔵",
        "corrupted_ocr_failed": "🔴", "ocr_needed": "🟡",
        "empty": "⚪", "error": "❌", "unknown": "❓",
    }

    by_status: Dict = defaultdict(list)
    for r in rows:
        by_status[r[5] or "unknown"].append(r)

    print(f"\n  {'Status':<26} {'Count':>6}")
    print(f"  {'─'*36}")
    for status, lst in sorted(by_status.items()):
        print(f"  {icons.get(status,'⚠️')} {status:<24} {len(lst):>6}")
    print(f"  {'─'*36}")
    print(f"  {'TOTAL':<26} {len(rows):>6}")

    problematic = [r for r in rows if (r[5] or "") in
                   ("corrupted_ocr_failed", "ocr_needed", "empty", "error")]
    if problematic:
        print(f"\n  ⚠️  Documente problematice ({len(problematic)}):")
        print(f"\n  {'File':<44} {'Company':<24} {'Year':>4}  {'Status'}")
        print(f"  {'─'*90}")
        for r in problematic[:25]:
            print(f"  {(r[0] or '')[:43]:<44} {(r[1] or '—')[:23]:<24} "
                  f"{r[2] or 0:>4}  {icons.get(r[5],'⚠️')} {r[5]}")
        if len(problematic) > 25:
            print(f"  ... și {len(problematic)-25} mai multe")
    pause()


def opt_8_4(sess: Session):
    """8.4 — Coverage matrix (companie × an × tip document) → Excel."""
    section("8.4  COVERAGE MATRIX  (companie × an × tip document)  → Excel")
    config, _ = sess.resolve()
    output = ask("Output folder", config.output_folder)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  ❌ openpyxl nu este instalat:  pip install openpyxl"); pause(); return

    folder = Path(config.input_folder)
    if not folder.exists():
        print(f"  ❌ Folder inexistent: {folder}"); pause(); return

    all_pdfs = list(folder.glob("*.pdf"))
    if not all_pdfs:
        print("  ⚠️  Niciun PDF în folder."); pause(); return

    doc_data, doc_types, years_found = [], set(), set()
    for pdf in all_pdfs:
        parts = pdf.stem.split(" - ")
        try:
            pc  = parts[0].split(".", 1)
            pos = int(pc[0].strip())
            co  = pc[1].strip() if len(pc) > 1 else pdf.stem[:30]
            yr  = int(parts[1].strip()) if len(parts) > 1 else 0
            dt  = parts[2].strip() if len(parts) > 2 else "unknown"
        except (ValueError, IndexError):
            pos, co, yr, dt = 0, pdf.stem[:30], 0, "unknown"
        doc_data.append({"pos": pos, "company": co, "year": yr, "doc_type": dt})
        doc_types.add(dt)
        if yr > 0: years_found.add(yr)

    dts, yrs = sorted(doc_types), sorted(years_found)
    cov: Dict = defaultdict(dict)
    for d in doc_data:
        cov[(d["pos"], d["company"])][(d["doc_type"], d["year"])] = 1

    wb = Workbook(); ws = wb.active; ws.title = "Coverage"
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    g_fill = PatternFill("solid", fgColor="C6EFCE")
    r_fill = PatternFill("solid", fgColor="FFC7CE")
    bdr    = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    ctr    = Alignment(horizontal="center")

    hdrs = ["Pos", "Company"] + [f"{dt} {y}" for dt in dts for y in yrs]
    for col, hd in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=col, value=hd)
        cell.fill = h_fill; cell.font = h_font
        cell.border = bdr;  cell.alignment = ctr

    cos = sorted(cov.keys(), key=lambda x: x[0])
    tots: Dict = defaultdict(int)
    for ri, (pos, co) in enumerate(cos, 2):
        ws.cell(row=ri, column=1, value=pos).border = bdr
        ws.cell(row=ri, column=2, value=co).border  = bdr
        col = 3
        for dt in dts:
            for yr in yrs:
                val  = cov[(pos, co)].get((dt, yr), 0)
                cell = ws.cell(row=ri, column=col, value=val)
                cell.border = bdr; cell.alignment = ctr
                cell.fill   = g_fill if val else r_fill
                if val: tots[(dt, yr)] += 1
                col += 1

    tr = len(cos) + 2
    ws.cell(row=tr, column=2, value="TOTAL").font = Font(bold=True)
    col = 3
    for dt in dts:
        for yr in yrs:
            c = ws.cell(row=tr, column=col, value=tots[(dt, yr)])
            c.font = Font(bold=True); c.alignment = ctr; col += 1

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30

    out = Path(output) / "document_coverage_matrix.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    result_ok("COVERAGE MATRIX GENERAT")
    print(f"  Companies : {len(cos)}")
    print(f"  Types     : {', '.join(dts)}")
    print(f"  Years     : {', '.join(map(str, yrs))}")
    print(f"  📄 Output  : {out}")
    pause()


def display_cap8_menu(sess: Session):
    ps     = sess._last_proc_stats
    docs   = sess._last_doc_stats
    n_docs = len(docs) if docs else 0
    n_pgs  = sum(d.get("pages", 0) for d in docs) if docs else 0

    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h("📂  8.  COVERAGE & STATISTICS")
    if ps:
        print(f"     Ultima sesiune: {n_docs} documente, {n_pgs:,} pagini procesate")
    else:
        print("     [nicio sesiune recentă — rulează mai întâi cap 1]")
    print()
    print("     8.1   Statistici run curent        (sumar sesiune: pages/words/refs)")
    print("     8.2   Statistici per document      (tabel detaliat ultima sesiune)")
    print("     8.3   Status procesare             (breakdown by text_status)")
    print("     8.4   Coverage matrix → Excel      (companie × an × tip document)")
    print()
    print("     Metrici disponibili: documente, pagini, cuvinte, fraze, refs detectate")


# ─────────────────────────────────────────────────────────────────────────────

def display_cap67_menu():
    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h("📊  6/7.  VIZUALIZĂRI & GRAFICE")
    print("     6.1   Charts Excel          (grafice embedded în workbook)")
    print("     6.2   Dashboard HTML        (interactiv, Plotly)")
    print("     6.3   Raport PDF            (publication-quality)")
    print()
    print("     Notă: necesită 16_viz.py și dependințe (plotly, reportlab).")


# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# CAP 9 — CONFIGURAȚIE & STATUS
# ─────────────────────────────────────────────────────────────────────────────

def opt_9_1(sess: Session):
    section("9.1  CONFIGURAȚIE CURENTĂ")
    cli  = _cli()
    args = sess.base_args(show=True, save=None, set_key=None, set_value=None, reset=False)
    try:
        cli.cmd_config(args)
    except Exception as e:
        print(f"  ❌ Eroare: {e}")
    # Afișează și taxonomiile v2
    if sess.taxonomies:
        print(f"\n  Taxonomii configurate ({len(sess.taxonomies)}):")
        for i, t in enumerate(sess.taxonomies):
            marker = "  [PRIMAR]" if i == 0 else ""
            print(f"    [{i+1}] {t['taxonomy_name']}{marker}")
            print(f"         Folder: {t['output_folder']}")
            print(f"         DB:     {t['database_name']}")
    pause()


def opt_9_2(sess: Session):
    section("9.2  SALVARE CONFIGURAȚIE")
    try:
        config, _ = sess.resolve()
        default_path = str(Path(config.output_folder) / "aisa_config.json")
    except Exception:
        default_path = "aisa_config.json"

    print(f"  Default location: {default_path}")
    path = ask("JSON file path", default_path).strip()

    if not path:
        print("  ✗ Cale invalidă. Operație anulată.")
        pause(); return
    p = Path(path)
    if p.is_dir():
        path = str(p / "aisa_config.json")
        print(f"  → Salvare în: {path}")
    if not path.lower().endswith(".json"):
        path += ".json"

    try:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"  ❌ Nu se poate crea directorul: {e}")
        pause(); return

    cli  = _cli()
    args = sess.base_args(show=False, save=path, set_key=None, set_value=None, reset=False)
    try:
        cli.cmd_config(args)
        # Suprascrie cu format extins v2 (include taxonomies)
        data = json.loads(Path(path).read_text(encoding="utf-8")) if Path(path).exists() else {}
        if sess.taxonomies:
            data["taxonomies"]           = sess.taxonomies
            data["primary_taxonomy_idx"] = sess.primary_taxonomy_idx
        Path(path).write_text(
            json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        sess.config_path = path
        print(f"  ✓ Salvat în: {path}")
    except Exception as e:
        print(f"  ❌ Eroare la salvare: {e}")
    pause()


def opt_9_3(sess: Session):
    """9.3 — Setează căi input/output pentru sesiunea curentă."""
    section("9.3  SETARE CĂI INPUT / OUTPUT")

    # Afișează info taxonomii (întărire PAS 8)
    n_tax = len(sess.taxonomies) if sess.taxonomies else 1
    pt    = sess.primary_taxonomy
    print(f"  Taxonomii configurate : {n_tax}")
    print(f"  Taxonomie primară     : {pt.get('taxonomy_name', 'N/A')}")

    try:
        config, _ = sess.resolve()
        cur_input  = config.input_folder
        cur_output = config.output_folder
        cur_csv    = config.fortune500_csv or ""
        cur_db     = config.database_name or "aisa_results.db"
    except Exception:
        cur_input  = sess.input_folder  or "Fortune500_PDFs"
        cur_output = sess.output_folder or "Results_AISA"
        cur_csv    = ""
        cur_db     = "aisa_results.db"

    # Arată unde va fi salvat JSON-ul (întărire PAS 8)
    save_location = str(Path(pt.get("output_folder", cur_output)) / "aisa_config.json")
    print(f"  Config va fi salvat în: {save_location}")
    print()
    print(f"  Input curent   : {cur_input}")
    print(f"  Output curent  : {cur_output}")
    print(f"  CSV F500       : {cur_csv or '(negăsit)'}")
    print(f"  Database       : {cur_db}")
    print()

    new_input  = ask("PDF folder           (Enter = păstrează)", cur_input)
    new_output = ask("Results folder       (Enter = păstrează)", cur_output)
    new_csv    = ask("CSV Fortune 500      (Enter = păstrează)", cur_csv).strip() or cur_csv
    new_db     = ask("Database name        (Enter = păstrează)", cur_db).strip()   or cur_db

    changed = False
    if new_input != cur_input:
        if not Path(new_input).exists():
            print(f"  ⚠️  Folder '{new_input}' inexistent. Se creează...")
            Path(new_input).mkdir(parents=True, exist_ok=True)
        sess.input_folder = new_input
        changed = True
        print(f"  ✓ Input setat  : {new_input}")

    if new_output != cur_output:
        Path(new_output).mkdir(parents=True, exist_ok=True)
        sess.output_folder = new_output
        changed = True
        print(f"  ✓ Output setat : {new_output}")

    extra = sess.get_extra_config()
    if new_csv != cur_csv:
        extra["fortune500_csv"] = new_csv or None
        sess._extra_config = extra
        changed = True
        print(f"  ✓ CSV F500 setat: {new_csv or '(gol)'}")
    if new_db != cur_db:
        extra["database_name"] = new_db
        sess._extra_config = extra
        # Actualizează și în taxonomia primară dacă e v2
        if sess.taxonomies:
            sess.taxonomies[0]["database_name"] = new_db
            sess.sync_legacy_fields()
        changed = True
        print(f"  ✓ Database setat: {new_db}")

    if changed:
        sess.invalidate()
        if ask_yn("\n  Salvezi configurația în JSON?", False):
            cfg_path = ask("JSON file path", sess.config_path or save_location).strip()
            if not cfg_path:
                print("  ✗ Cale invalidă.")
            else:
                if Path(cfg_path).is_dir():
                    cfg_path = str(Path(cfg_path) / "aisa_config.json")
                if not cfg_path.lower().endswith(".json"):
                    cfg_path += ".json"
                try:
                    Path(cfg_path).parent.mkdir(parents=True, exist_ok=True)
                    config2, _ = sess.resolve()
                    cfg_data = config2.to_dict()
                    if sess.taxonomies:
                        cfg_data["taxonomies"]           = sess.taxonomies
                        cfg_data["primary_taxonomy_idx"] = sess.primary_taxonomy_idx
                    Path(cfg_path).write_text(
                        json.dumps(cfg_data, indent=2, ensure_ascii=False),
                        encoding="utf-8"
                    )
                    sess.config_path = cfg_path
                    print(f"  ✓ Salvat în {cfg_path}")
                except Exception as e:
                    print(f"  ❌ Eroare salvare: {e}")
    else:
        print("  ✓ Nicio modificare.")
    pause()


def opt_9_4(sess: Session):
    """9.4 — Translate non-English references + context into English."""
    section("9.4  TRADUCERE REFERINȚE + CONTEXT ÎN ENGLEZĂ")
    print("  Traduce fragmentele non-EN din baza curentă și salvează:")
    print("    - text_translated")
    print("    - context_translated")
    print()
    print("  Engine-uri suportate: deepl / llm / google / helsinki")
    print("  Recomandat: DeepL ca primar, LLM ca fallback.")
    print()

    try:
        _, db_path = sess.resolve()
        print(f"  DB curentă : {db_path}")
    except Exception as e:
        print(f"  ❌ Nu pot rezolva DB-ul curent: {e}")
        pause(); return

    default_engine = "deepl"
    engine = ask("Engine [deepl/llm/google/helsinki]", default_engine).strip().lower() or default_engine
    if engine not in ("deepl", "llm", "google", "helsinki"):
        print(f"  ⚠️  Engine invalid: {engine}. Folosesc '{default_engine}'.")
        engine = default_engine

    batch_size = ask_int("Batch size", 50, lo=1, hi=5000)
    only_missing = ask_yn("Tradu doar rândurile fără traducere?", True)
    dry_run = ask_yn("Dry run (fără write în DB)?", False)

    print()
    print("  Variabile utile de mediu:")
    print("    DEEPL_API_KEY")
    print("    OPENAI_API_KEY")
    print("    OPENAI_BASE_URL        [opțional]")
    print("    TRANSLATION_LLM_MODEL  [opțional]")
    print()

    if not ask_yn("Continui cu traducerea?", True):
        print("  Anulat.")
        pause(); return

    try:
        m15 = importlib.import_module("15_translate")

        avail, msg = m15.check_engine_availability(engine)
        if not avail:
            best = m15.get_best_available_engine()
            if not best:
                print(f"  ❌ Engine '{engine}' indisponibil: {msg}")
                print("  Niciun engine disponibil. Configurează cheile/API-urile și încearcă din nou.")
                pause(); return
            print(f"  ⚠️  Engine '{engine}' indisponibil: {msg}")
            print(f"  → Folosesc fallback automat: {best}")
            engine = best

        stats = m15.update_translations_in_db(
            db_path=db_path,
            engine=engine,
            batch_size=batch_size,
            only_missing=only_missing,
            dry_run=dry_run,
        )

        print()
        print(SEP2)
        print("  REZULTAT TRADUCERE")
        print(SEP2)
        if isinstance(stats, dict):
            for key in ("rows_found", "processed", "translated", "skipped", "errors", "engine_used", "dry_run"):
                if key in stats:
                    print(f"  {key:<12}: {stats[key]}")
        else:
            print(f"  Rezultat: {stats}")

        if not dry_run and ask_yn("Generezi export Excel după traducere?", True):
            try:
                config, _ = sess.resolve()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                base_name = f"translated_export_{timestamp}"
                cli = _cli()
                args = sess.base_args(
                    output=config.output_folder,
                    format="excel",
                    year=None,
                    base_name=base_name,
                    export_type="raw",
                )
                cli.cmd_export(args)
                print(f"  ✓ Export generat în {config.output_folder}/")
            except Exception as e:
                print(f"  ⚠️  Traducerea s-a făcut, dar exportul a eșuat: {e}")

    except Exception as e:
        print(f"  ❌ Eroare la traducere: {e}")
        import traceback; traceback.print_exc()

    pause()


def display_cap9_menu(sess: Session):
    def h(title):
        BW = W - 4
        print(f"\n  ╔{'═'*BW}╗")
        print(f"  ║  {vpad(title, BW-2)}║")
        print(f"  ╚{'═'*BW}╝")

    h("🔧  9.  CONFIGURAȚIE & STATUS")
    print("     9.1   Afișează configurația curentă")
    print("     9.2   Salvează configurația în JSON")
    print("     9.3   Setează căi input / output")
    print("     9.4   Tradu referințe + context în engleză")
    print()
    print("     T.1   Taxonomie built-in Python         (schimbă taxonomia activă)")
    print("     T.2   Încarcă taxonomie din Excel        (.xlsx)")
    print("     T.3   Afișează taxonomii disponibile")
    print("     T.4   Configurează multi-taxonomy        (pipeline paralel)")


# ─────────────────────────────────────────────────────────────────────────────
# CAP T — TAXONOMY (copiat și adaptat pentru v2 sess.taxonomies)
# ─────────────────────────────────────────────────────────────────────────────

def opt_T_1(sess: Session):
    """T.1 — Switch between built-in Python taxonomies."""
    section("T.1  SELECTARE TAXONOMIE BUILT-IN")
    try:
        from version import TAXONOMY_REGISTRY
        registry = TAXONOMY_REGISTRY
    except (ImportError, AttributeError):
        registry = {"AI_Disclosure":        "AI Disclosure (default)",
                    "Digitalization_Relational_v2": "Digitalization Relational v2.2",
                    "Digitalization_Relational_v2_2_ZH": "Digitalization Relational v2.2 ZH (bilingual)"}

    print("  Taxonomii disponibile:\n")
    names = list(registry.keys()) if isinstance(registry, dict) else list(registry)
    for i, name in enumerate(names, 1):
        label  = registry[name] if isinstance(registry, dict) else name
        marker = "  <- ACTIV" if name == sess.taxonomy_name else ""
        print(f"     {i}.  {name:<35}  {label}{marker}")

    print()
    choice = ask(f"Select [1-{len(names)}] (Enter = păstrează)", "").strip()
    if not choice:
        print("  Neschimbat."); pause(); return

    try:
        idx = int(choice) - 1
        if not (0 <= idx < len(names)):
            raise ValueError
    except ValueError:
        print("  Selecție invalidă."); pause(); return

    new_name = names[idx]
    # Actualizează sess.taxonomies (v2) și câmpurile legacy
    if sess.taxonomies:
        sess.taxonomies[0] = _build_taxonomy_entry(new_name, None)
        sess.sync_legacy_fields()
    else:
        sess.taxonomy_name  = new_name
        sess.taxonomy_excel = None
    sess.invalidate()
    print(f"  ✓ Taxonomie setată: {new_name}")
    print(f"  Output folder: {new_name}_results/")
    print(f"  Database:      {new_name}_results.db")
    pause()


def opt_T_2(sess: Session):
    """T.2 — Load taxonomy from Excel file."""
    section("T.2  ÎNCARCĂ TAXONOMIE DIN EXCEL")
    print("  Fișierul Excel trebuie să aibă foile: Categories, Keywords (obligatorii);")
    print("  Patterns, FP_Patterns, Taxonomy_Meta (opționale).")
    print()

    current = sess.taxonomy_excel or ""
    path    = ask("Calea la fișierul .xlsx (Enter = anulează)", current).strip()
    if not path:
        print("  Anulat."); pause(); return

    if not Path(path).exists():
        print(f"  Fișier negăsit: {path}"); pause(); return
    if not path.lower().endswith(".xlsx"):
        print("  Fișierul trebuie să fie .xlsx."); pause(); return

    try:
        m_loader = importlib.import_module("taxonomy_excel_loader")
        provider = m_loader.ExcelTaxonomyProvider(path)
        dims     = provider.get_dimensions()
        print(f"\n  Fișier valid  --  v{provider.get_version()}")
        for dim_name, cats in dims.items():
            total_kw = sum(len(c.keywords) for c in cats.values())
            print(f"    [{dim_name}]  {len(cats)} categorii  |  {total_kw} cuvinte cheie")
        fp = provider.get_fp_patterns()
        if fp:
            print(f"    [FP]  {sum(len(v) for v in fp.values())} FP patterns")
    except Exception as e:
        print(f"  Eroare la încărcare taxonomie: {e}"); pause(); return

    try:
        tax_name = provider._name
    except Exception:
        tax_name = Path(path).stem

    # Actualizează sess.taxonomies (v2) și câmpurile legacy
    if sess.taxonomies:
        sess.taxonomies[0] = _build_taxonomy_entry(tax_name, path)
        sess.sync_legacy_fields()
    else:
        sess.taxonomy_excel = path
        sess.taxonomy_name  = tax_name
    sess.invalidate()

    print(f"\n  ✓ Taxonomie încărcată: {tax_name}")
    print(f"  Output folder: {tax_name}_results/")
    print(f"  Database:      {tax_name}_results.db")
    pause()


def opt_T_3(sess: Session):
    """T.3 — Show all available taxonomies."""
    section("T.3  TAXONOMII DISPONIBILE")

    try:
        from version import TAXONOMY_REGISTRY
        registry = TAXONOMY_REGISTRY
    except (ImportError, AttributeError):
        registry = {"AI_Disclosure":        "AI Disclosure",
                    "Digitalization_Relational_v2": "Digitalization Relational v2.2",
                    "Digitalization_Relational_v2_2_ZH": "Digitalization Relational v2.2 ZH (bilingual)"}

    print("  Taxonomii Python built-in:")
    names = list(registry.keys()) if isinstance(registry, dict) else list(registry)
    for name in names:
        label  = registry[name] if isinstance(registry, dict) else name
        marker = "  <- ACTIV" if (name == sess.taxonomy_name and not sess.taxonomy_excel) else ""
        print(f"     * {name:<35} {label}{marker}")

    print("\n  Fișiere Excel găsite pe disc:")
    xlsx_files = (
        glob.glob("*.xlsx") +
        glob.glob("taxonomies/*.xlsx") +
        glob.glob("*_results/*.xlsx")
    )
    tax_files = [f for f in xlsx_files if "taxonom" in f.lower()]
    if not tax_files:
        print("     (niciun fișier Excel găsit)")
    else:
        for f in tax_files:
            marker = "  <- ACTIV" if f == sess.taxonomy_excel else ""
            print(f"     * {f}{marker}")

    print()
    print("  Folosește T.1 pentru built-in, T.2 pentru Excel.")
    pause()


def opt_T_4(sess: Session):
    """T.4 — Configure extra taxonomies for multi-taxonomy pipeline."""
    while True:
        section("T.4  CONFIGURARE MULTI-TAXONOMY")

        primary_src = sess.taxonomy_excel or sess.taxonomy_name
        print(f"  Taxonomie primară (mereu inclusă): {primary_src}")
        print()

        if sess.extra_taxonomies:
            print(f"  Taxonomii extra ({len(sess.extra_taxonomies)} configurate):")
            for i, t in enumerate(sess.extra_taxonomies, 1):
                src = t.get("taxonomy_excel") or t.get("taxonomy_name")
                print(f"    {i}.  {src}")
        else:
            print("  Taxonomii extra: (niciuna configurată)")

        print()
        print("  A.  Adaugă taxonomie Python built-in")
        print("  B.  Adaugă taxonomie din Excel")
        print("  C.  Șterge toate taxonomiile extra")
        print("  0.  Înapoi")
        print()

        choice = ask("Select", "0").strip().upper()

        if choice == "0":
            # Sincronizează sess.taxonomies din extra_taxonomies
            all_tax = [{"taxonomy_name": sess.taxonomy_name,
                        "taxonomy_excel": sess.taxonomy_excel}] + sess.extra_taxonomies
            sess.taxonomies = [
                _build_taxonomy_entry(t["taxonomy_name"], t.get("taxonomy_excel"))
                for t in all_tax
            ]
            sess.sync_legacy_fields()
            break

        elif choice == "A":
            try:
                from version import TAXONOMY_REGISTRY
                registry = TAXONOMY_REGISTRY
            except (ImportError, AttributeError):
                registry = {"AI_Disclosure": "AI Disclosure",
                            "Digitalization_Relational_v2": "Digitalization Relational v2.2",
                    "Digitalization_Relational_v2_2_ZH": "Digitalization Relational v2.2 ZH (bilingual)"}

            names    = list(registry.keys()) if isinstance(registry, dict) else list(registry)
            already  = {t["taxonomy_name"] for t in sess.extra_taxonomies
                        if not t.get("taxonomy_excel")}
            primary_name = sess.taxonomy_name if not sess.taxonomy_excel else None
            available    = [n for n in names if n != primary_name and n not in already]

            if not available:
                print("  Toate taxonomiile Python sunt deja configurate."); pause(); continue

            print("\n  Taxonomii Python disponibile de adăugat:\n")
            for i, name in enumerate(available, 1):
                label = registry[name] if isinstance(registry, dict) else name
                print(f"     {i}.  {name:<35}  {label}")
            print()
            choice2 = ask(f"Select [1-{len(available)}] (Enter = anulează)", "").strip()
            if not choice2:
                print("  Anulat."); continue
            try:
                idx = int(choice2) - 1
                if not (0 <= idx < len(available)):
                    raise ValueError
            except ValueError:
                print("  Selecție invalidă."); continue

            chosen = available[idx]
            sess.extra_taxonomies.append({"taxonomy_name": chosen, "taxonomy_excel": None})
            print(f"  ✓ Adăugat: {chosen}")
            pause()

        elif choice == "B":
            path = ask("Calea la fișierul .xlsx (Enter = anulează)", "").strip()
            if not path:
                print("  Anulat."); continue
            if not Path(path).exists():
                print(f"  Fișier negăsit: {path}"); pause(); continue
            if not path.lower().endswith(".xlsx"):
                print("  Fișierul trebuie să fie .xlsx."); pause(); continue
            if sess.taxonomy_excel == path:
                print("  ⚠️  Acesta este deja taxonomia primară."); pause(); continue
            if any(t.get("taxonomy_excel") == path for t in sess.extra_taxonomies):
                print("  ⚠️  Fișierul e deja în lista extra."); pause(); continue

            try:
                loader   = importlib.import_module("taxonomy_excel_loader")
                provider = loader.ExcelTaxonomyProvider(path)
                dims     = provider.get_dimensions()
                print(f"\n  Fișier valid  --  v{provider.get_version()}")
                for dim_name, cats in dims.items():
                    total_kw = sum(len(c.keywords) for c in cats.values())
                    print(f"    [{dim_name}]  {len(cats)} categorii  |  {total_kw} kw")
            except Exception as e:
                print(f"  Eroare: {e}"); pause(); continue

            try:
                tax_name = provider._name
            except Exception:
                tax_name = Path(path).stem

            sess.extra_taxonomies.append({"taxonomy_name": tax_name, "taxonomy_excel": path})
            print(f"  ✓ Adăugat: {path}  [{tax_name}]")
            pause()

        elif choice == "C":
            if not sess.extra_taxonomies:
                print("  Nimic de șters."); continue
            if ask_yn(f"Șterge toate {len(sess.extra_taxonomies)} taxonomii extra?", False):
                sess.extra_taxonomies = []
                if sess.taxonomies and len(sess.taxonomies) > 1:
                    sess.taxonomies = sess.taxonomies[:1]
                    sess.sync_legacy_fields()
                print("  ✓ Șterse.")
                pause()
        else:
            print("  Opțiune invalidă.")


# ─────────────────────────────────────────────────────────────────────────────
# MENIU PRINCIPAL — versiunea finală (PAS 8)
# ─────────────────────────────────────────────────────────────────────────────

def display_main_menu(sess: Session):
    clr()
    display_header(sess)
    display_cap1_menu(sess)
    display_cap2_menu(sess)
    display_cap3_menu(sess)
    display_cap4_menu(sess)
    display_cap67_menu()
    display_cap8_menu(sess)
    display_cap9_menu(sess)
    print(f"\n  {SEP2}")
    print("     0.    IEȘIRE")
    print(f"  {SEP2}\n")


# ─────────────────────────────────────────────────────────────────────────────
# HANDLERS — versiunea finală (PAS 8)
# ─────────────────────────────────────────────────────────────────────────────

_HANDLERS: Dict[str, object] = {
    "1.1": opt_1_1, "1.2": opt_1_2, "1.3": opt_1_3,
    "1.4": opt_1_4, "1.5": opt_1_5,
    "2.1": opt_2_1, "2.2": opt_2_2, "2.3": opt_2_3,
    "3.1": opt_3_1, "3.2": opt_3_2, "3.3": opt_3_3, "3.4": opt_3_4,
    "4.1": opt_4_1, "4.2": opt_4_2, "4.3": opt_4_3, "4.4": opt_4_4,
    "6.1": opt_6_1, "6.2": opt_6_2, "6.3": opt_6_3,
    "8.1": opt_8_1, "8.2": opt_8_2, "8.3": opt_8_3, "8.4": opt_8_4,
    "9.1": opt_9_1, "9.2": opt_9_2, "9.3": opt_9_3, "9.4": opt_9_4,
    "t.1": opt_T_1, "t.2": opt_T_2, "t.3": opt_T_3, "t.4": opt_T_4,
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN LOOP — versiunea finală (PAS 8)
# ─────────────────────────────────────────────────────────────────────────────

def run(sess: Session):
    while True:
        try:
            display_main_menu(sess)
            choice = input("  🔹 Opțiune: ").strip().lower()

            if choice in ("0", "exit", "quit", "q"):
                clr()
                print("\n  👋 La revedere!\n")
                sys.exit(0)

            if choice in _HANDLERS:
                try:
                    _HANDLERS[choice](sess)
                except KeyboardInterrupt:
                    print("\n  ⚠️  Întrerupt.")
                    pause()
                except Exception as exc:
                    print(f"\n  ❌ Eroare neașteptată în opțiunea '{choice}': {exc}")
                    import traceback; traceback.print_exc()
                    pause()
            else:
                print(f"  ✗ Opțiune invalidă: '{choice}'")
                time.sleep(1)

        except KeyboardInterrupt:
            clr()
            print("\n  ⚠️  Întrerupt de utilizator. La revedere!\n")
            sys.exit(0)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT — versiunea finală (PAS 8)
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="main_new",
        description=f"AISA v{AISA_VERSION} — Interactive menu v2",
    )
    parser.add_argument("--config", "-c", metavar="PATH",
                        help="Config JSON (sare peste wizard)")
    parser.add_argument("--db",           metavar="PATH",
                        help="Override cale SQLite DB")
    parser.add_argument("--input",  "-i", metavar="DIR",
                        help="Override folder PDF-uri")
    parser.add_argument("--output", "-o", metavar="DIR",
                        help="Override folder rezultate")
    args = parser.parse_args()

    sess = _startup_config(args)

    # Aplică override-urile CLI explicit (întărire PAS 8)
    if args.db:
        sess.db_override = args.db
    if args.input:
        sess.input_folder = args.input
        sess.invalidate()
    if args.output:
        sess.output_folder = args.output
        # Actualizează și în taxonomia primară dacă există
        if sess.taxonomies:
            sess.taxonomies[0]["output_folder"] = args.output
            sess.sync_legacy_fields()
        sess.invalidate()

    run(sess)


if __name__ == "__main__":
    main()